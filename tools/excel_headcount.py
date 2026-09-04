"""Excel com TODOS os animais que entram no headcount da semana, um por linha.

Serve para auditar o número: em vez de discutir se o total está certo, abre-se a
lista e confere-se nome por nome. Vem também a lista do que ficou FORA, com o
motivo de cada exclusão, que é onde as divergências com o haras sempre moraram
(status de saída, embrião, linha repetida por cotista, Mato Grosso).

As regras NÃO são reimplementadas aqui: a contagem sai das mesmas funções do
fechamento (`_plantel_por_status`, `_receptoras_info`), então o Excel e o
relatório publicado não podem discordar.

Abas:
  CONTAGEM   animais do plantel + receptoras, com o bucket de cada um
  FORA       linhas descartadas, com o motivo
  RESUMO     total por bucket, fechando com o headcount da semana

Uso:
    python tools/excel_headcount.py                    # semana que fecha hoje
    python tools/excel_headcount.py 04/09/2026
    python tools/excel_headcount.py 04/09/2026 "C:/caminho/arquivo.xlsx"
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import PGSemanalReport as R  # noqa: E402

NAVY = "FF04223B"
AMBER = "FFCA9703"


def _cabecalho(ws, titulos):
    ws.append(titulos)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _larguras(ws, larguras):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def monta(fim: date) -> Workbook:
    ini = R._inicio_da_semana(fim)
    print(f"janela {ini} a {fim}")

    plantel = R._plantel_por_status()
    receptoras = R._receptoras_info()
    contados = set(plantel["roster"])

    wb = Workbook()

    # ---------------- CONTAGEM ----------------
    ws = wb.active
    ws.title = "CONTAGEM"
    # MÃE e PAI não são enfeite: dois potros diferentes aparecem no plantel só como
    # 'MACHO', e é a filiação que os separa (é assim que a chave do roster os
    # distingue). Sem as duas colunas, a lista parece ter linha repetida.
    _cabecalho(ws, ["#", "ANIMAL", "MÃE", "PAI", "TIPO", "CATEGORIA", "STATUS",
                    "LOCAL", "BUCKET"])
    n = 0
    por_bucket = {}
    for l in sorted(plantel["linhas"], key=lambda x: (R._norm(x.get("local")), R._norm(x.get("nome")))):
        local = R._norm(l.get("local"))
        if local in R.HEADCOUNT_LOCAIS_FORA or local not in R.HEADCOUNT_BUCKETS:
            continue                      # Mato Grosso e local sem bucket não contam
        rot = R.HEADCOUNT_BUCKETS[local][0]
        n += 1
        por_bucket[rot] = por_bucket.get(rot, 0) + 1
        ws.append([n, l.get("nome"), l.get("mae"), l.get("pai"), "ANIMAL",
                   l.get("categoria"), l.get("status_plantel"), l.get("local"), rot])
    animais = n

    for nome, info in sorted(receptoras.items()):
        if info.get("local") not in R.RECEPTORAS_LOCAIS_ATIVOS:
            continue
        st = R._norm(info.get("status"))
        if not (st.startswith("PRENHA") or st.startswith("VAZIA")):
            continue                      # POTRA/MATRIZ/POTRO na planilha de receptoras
        bucket = R.HEADCOUNT_BUCKETS[R.RECEPTORAS_PARA_BUCKET[info["local"]]][0]
        n += 1
        por_bucket[bucket] = por_bucket.get(bucket, 0) + 1
        ws.append([n, f"RECEPTORA {nome}", None, None, "RECEPTORA", "RECEPTORA",
                   info.get("status"), info.get("local"), bucket])
    recept = n - animais
    _larguras(ws, [5, 46, 30, 26, 11, 14, 22, 28, 14])

    # ---------------- FORA ----------------
    wf = wb.create_sheet("FORA")
    _cabecalho(wf, ["ANIMAL", "CATEGORIA", "STATUS", "LOCAL", "COTA", "CONDIÇÃO ATUAL",
                    "MOTIVO DE FICAR FORA", "OBS DA PLANILHA"])
    fora = list(plantel.get("descartadas") or [])
    # Mato Grosso não é descarte do roster, é local fora da contagem — entra aqui
    # porque na conversa com o haras ele sempre aparece como "e esses?"
    for l in plantel["linhas"]:
        if R._norm(l.get("local")) in R.HEADCOUNT_LOCAIS_FORA:
            fora.append({"nome": l.get("nome"), "categoria": l.get("categoria"),
                         "status": l.get("status_plantel"), "local": l.get("local"),
                         "cota": None, "condicao": None, "obs": None,
                         "motivo": "local fora da contagem (decisão de 31/07/2026)"})
    for l in sorted(fora, key=lambda x: (x.get("motivo") or "", R._norm(x.get("nome")))):
        wf.append([l.get("nome"), l.get("categoria"), l.get("status"), l.get("local"),
                   l.get("cota"), l.get("condicao"), l.get("motivo"), l.get("obs")])
    _larguras(wf, [46, 14, 22, 28, 8, 20, 38, 60])

    # ---------------- RESUMO ----------------
    wr = wb.create_sheet("RESUMO")
    _cabecalho(wr, ["LOCAL", "ANIMAIS + RECEPTORAS"])
    for rot in ("FAZENDA", "ARRENDAMENTO", "CTE", "SOCIO"):
        wr.append([rot, por_bucket.get(rot, 0)])
    wr.append(["TOTAL", n])
    wr.append([])
    wr.append(["animais do plantel", animais])
    wr.append(["receptoras contadas", recept])
    wr.append([])
    wr.append(["semana", f"{ini} a {fim}"])
    wr.append(["fonte do plantel", plantel["fonte"]])
    wr.append(["linhas fora da contagem", len(fora)])
    for c in wr["A"]:
        c.font = Font(bold=(c.row in (1, 6)))
    wr["B6"].font = Font(bold=True, color=AMBER)
    _larguras(wr, [30, 24])

    # ---------------- CONFERIR x HARAS ----------------
    # A lista serve para achar A LINHA que difere, em vez de discutir o total: com
    # 193 do nosso lado e 194 do lado deles, é UM nome. As duas abas se olham por
    # COUNTIF nos dois sentidos, então basta colar a lista do haras e as duas
    # colunas acendem sozinhas.
    # Nomes de aba SEM espaço, de propósito: fórmula com nome de aba espaçado
    # precisa de apóstrofo e vira armadilha de escape na hora de gerar.
    wc = wb.create_sheet("CONFERIR")
    _cabecalho(wc, ["#", "ANIMAL", "LOCAL", "BUCKET", "ESTÁ NA LISTA DO HARAS?"])
    nossos = [(r[1], r[7], r[8]) for r in ws.iter_rows(min_row=2, values_only=True)]
    for i, (nome, local, bucket) in enumerate(sorted(nossos, key=lambda x: str(x[0])), start=1):
        lin = i + 1
        wc.append([i, nome, local, bucket,
                   '=IF(COUNTIF(HARAS!$A:$A,B{0})>0,"sim","NAO - so na nossa")'.format(lin)])
    _larguras(wc, [5, 52, 28, 14, 30])

    wh = wb.create_sheet("HARAS")
    _cabecalho(wh, ["ANIMAL (cole aqui a lista do haras, a partir de A2)", "ESTÁ NA NOSSA?"])
    for lin in range(2, 402):
        wh.cell(row=lin, column=2).value = (
            '=IF(A{0}="","",IF(COUNTIF(CONFERIR!$B:$B,A{0})>0,'
            '"sim","NAO - so na lista do haras"))'.format(lin))
    _larguras(wh, [56, 30])

    print(f"contados: {n}  (animais {animais} + receptoras {recept})")
    print("por bucket:", por_bucket)
    print(f"fora: {len(fora)} linha(s)")
    return wb


def main():
    args = [a for a in sys.argv[1:]]
    fim = R._parse_d(args[0]) if args and "/" in args[0] else date.today()
    destino = Path(args[1]) if len(args) > 1 else (
        Path.home() / "Downloads" / f"headcount_{fim.strftime('%d-%m-%Y')}.xlsx")
    wb = monta(fim)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    print(f"-> {destino}")


if __name__ == "__main__":
    main()
