"""Monta a especificação do Comitê Mensal HPG a partir das bases reais.

O deck é a saída, não a fonte: cada slide vira um objeto {t: <tipo>, ...} num
spec JSON, e daí saem as DUAS renderizações — o HTML (hub/comite.html) e o PPTX
(botão "Exportar PPTX"). Fonte única, dois formatos.

Mapa completo de slide × fonte: `_docs/COMITE_MAPEAMENTO.md`.

**Um deck por mês.** O DRE não é lido pela aba "Real x Orçado", que mostra só o
mês em que o operador deixou o arquivo — é lido da aba `DRE-Compet`, que tem
TODOS os meses lado a lado (a linha 6 marca, em cada bloco, o número do mês na
coluna do Realizado). Assim o deck tem seletor de mês e nunca fica preso num mês
velho. As bases não-DRE (plantel, estação, vendas) entram no mês que elas têm;
quando o mês pedido não existe na base, o slide vira `pendente` dizendo isso.

Uso:
    python tools/build_comite.py          # todos os meses com dado
    python tools/build_comite.py 06/2026  # só esse mês
"""
import json
import base64
import io
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# tools/ vive na raiz do repo desde a reorganizacao: a raiz E o site.
REPO = Path(__file__).resolve().parent.parent
HUB = REPO
OUT = HUB / "assets" / "comite"
sys.path.insert(0, str(REPO / "scripts"))

# Caminhos do Drive e helpers vêm do pipeline que já roda — não duplicar.
from PGSemanalReport import (                                    # noqa: E402
    EMB_COMERCIAIS, MAPA_VENDAS_DIR, SAFRA_ATUAL, _controle_plantel,
    _latest_by_yymmdd, _latest_estacao_master, _load, _norm, _s, _to_num,
    caminho_curto,
)
# os resolvedores de fonte compartilhados anotam ali o arquivo que escolheram
from PGSemanalReport import _FONTES_USADAS as _FONTES_COMPARTILHADAS   # noqa: E402

DRE_DIR = Path(r"G:/Drives compartilhados/Luxor Controladoria/Ambiente de testes/DRE Data")
DRE_HARAS = DRE_DIR / "DRE 2026 HPG - HARAS.xlsx"
# NÃO é fonte do comitê — a seção CASA/FPG sai do DRE_Historico, igual ao resto do
# financeiro (ver _docs/COMITE_MAPEAMENTO.md, pendência 3). Fica declarada só para
# quem for procurar o arquivo do ano não concluir que ele foi esquecido.
DRE_CASA = DRE_DIR / "DRE 2026 FPG - CASA.xlsx"  # noqa: F401
PLANTEL_DIR = Path(r"C:/Users/Arthur/repos/LuxorMonthlyP-CRoutines/PlantelHPG")
BASE_BI = REPO / "bases" / "base_bi.parquet"

MESES = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
ABR = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

avisos = []
def aviso(m):
    if m not in avisos:
        avisos.append(m)
        print(f"  [aviso] {m}")


def _json_default(o):
    if isinstance(o, np.generic):
        return o.item()
    if isinstance(o, (datetime, date)):
        return o.isoformat()[:10]
    raise TypeError(f"tipo não serializável: {type(o).__name__}")


def num(v):
    if v is None or isinstance(v, str):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if pd.isna(f) else f


def pend(n, titulo, sub, fonte, motivo):
    return {"t": "pendente", "n": n, "titulo": titulo, "sub": sub, "fonte": fonte, "motivo": motivo}


def brl_curto(v):
    if v is None:
        return "—"
    s, a = ("-" if v < 0 else ""), abs(v)
    if a >= 1e6:
        return f"{s}R$ {a/1e6:.1f}M".replace(".", ",")
    if a >= 1e3:
        return f"{s}R$ {a/1e3:.0f}k"
    return f"{s}R$ {a:.0f}"


# ==================================================================== DRE
# Fonte: DRE_Historico.xlsx — a base consolidada que o LxDREdataExtractor gera e
# que o LuxorP&CHub já lê. É ela, e NÃO o "DRE 2026 HPG - HARAS.xlsx":
#   - o arquivo do ano tem só o mês em que o operador deixou a aba de resumo, e
#     em 03/08/2026 estava em fevereiro, com a coluna Orçado zerada;
#   - o histórico traz os 12 meses, Haras (HPG) e Casa (FPG), Competência e
#     Caixa, com Orçado e Realizado de verdade (jun/26 bate com o deck oficial).
# O DRE_Historico é DERIVADO: quem o gera é o LxDREdataExtractor, e ele grava o
# resultado AO LADO DE SI MESMO (OUTPUT_PATH = pasta do próprio script). Existem duas
# cópias do extractor — uma no repo de rotinas mensais, outra na pasta do Drive —, então
# existem duas saídas, e quem rodar a do repo não atualiza a do Drive.
# Foi o que aconteceu: em 18/08/2026 a cópia do repo estava em 11/08 com julho fechado
# (R$ 4,7M) e a do Drive em 15/07 com julho zerado. O comitê lia a do Drive e parava em
# junho, sem nada indicando que existia versão mais nova.
# Aqui vale a MAIS RECENTE, não uma pasta canônica: as duas são o mesmo artefato
# derivado e o que importa é o fechamento mais novo.
DRE_HIST_CANDIDATOS = (
    PLANTEL_DIR.parent / "DRE Data" / "DRE_Historico.xlsx",   # repo de rotinas mensais
    DRE_DIR / "DRE_Historico.xlsx",                            # Drive da Controladoria
)
_hist_cache = {}

# rótulo -> arquivo efetivamente aberto neste run. Vai para o spec e de lá para a
# auditoria: o que o build LEU, não o que alguém escreveu que ele lê.
_FONTES: dict = {}


def _registra(rotulo, caminho):
    """Anota a fonte e devolve o caminho, para caber em linha de chamada."""
    if caminho is not None:
        _FONTES[rotulo] = Path(caminho)
    return caminho


_dre_hist_cache = []


def _dre_hist():
    """Cópia mais recente do DRE_Historico, dizendo qual usou. O aviso sai UMA vez por
    run: a função é chamada tanto na checagem de existência quanto na leitura."""
    if _dre_hist_cache:
        return _dre_hist_cache[0]
    existentes = [p for p in DRE_HIST_CANDIDATOS if p.exists()]
    if not existentes:
        return None
    d = lambda p: datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y")
    escolhido = max(existentes, key=lambda p: p.stat().st_mtime)
    atrasadas = [o for o in existentes if o != escolhido]
    msg = f"  [dre] lendo a cópia de {escolhido.parents[1].name} ({d(escolhido)})"
    if atrasadas:
        msg += " — mais nova que " + ", ".join(f"{o.parents[1].name} ({d(o)})" for o in atrasadas)
    print(msg)
    _dre_hist_cache.append(_registra("DRE histórico", escolhido))
    return escolhido


def le_historico():
    """Base DRE Geral (mês) + Base YTD (acumulado), já filtradas para 2026."""
    if _hist_cache:
        return _hist_cache
    DRE_HIST = _dre_hist()
    if DRE_HIST is None:
        return {}
    geral = pd.read_excel(DRE_HIST, sheet_name="Base DRE Geral")
    ytd = pd.read_excel(DRE_HIST, sheet_name="Base YTD")
    for d in (geral, ytd):
        d["dt"] = pd.to_datetime(d["Data de Fechamento"])
        d["mes"] = d["dt"].dt.month
        d["ano"] = d["dt"].dt.year
    _hist_cache["geral"], _hist_cache["ytd"] = geral, ytd
    return _hist_cache


def meses_fechados(cc="HPG", modelo="Competência", ano=2026):
    """Mês só entra no deck quando tem realizado lançado — mês futuro tem só
    orçado, e um deck com realizado zerado engana mais do que informa."""
    h = le_historico()
    if not h:
        return []
    g = h["geral"]
    g = g[(g["Centro de Custo"] == cc) & (g["Modelo"] == modelo) & (g["ano"] == ano)]
    return sorted(int(m) for m, s in g.groupby("mes")["Realizado"].apply(lambda x: x.abs().sum()).items() if s)


def pct(orc, real):
    """∆ % = (realizado − orçado) / |orçado|. Sem orçado não existe percentual."""
    if not orc:
        return None
    return (real - orc) / abs(orc)


def _linhas_dre(df, col_orc, col_real, so_subtotal=False, so_com_valor=False):
    """Três níveis, tirados de Grupo/Subgrupo/É Subtotal:
        0 = linha do GRUPO (Receita Bruta, Custos e Despesas, Resultado…)
        1 = SUBGRUPO (Volumoso e Concentrado, Despesas com Pessoal…)
        2 = natureza folha
    Antes tudo que era `É Subtotal` virava nível 0, e o slide saía com quase toda
    linha em dourado e negrito — sem hierarquia nenhuma pra ler.

    Nome repetido em grupos diferentes (Sanidade aparece em Custos e em Despesas)
    ganha o grupo entre parênteses; sem isso a mesma palavra aparecia duas vezes
    com números diferentes.
    """
    linhas = []
    for _, r in df.sort_values("Ordem").iterrows():
        sub = bool(r["É Subtotal"])
        if so_subtotal and not sub:
            continue
        orc, real = num(r[col_orc]) or 0.0, num(r[col_real]) or 0.0
        if so_com_valor and not orc and not real and not sub:
            continue
        nome = str(r["Natureza de Lançamento"]).strip()
        grupo = "" if pd.isna(r["Grupo"]) else str(r["Grupo"]).strip()
        subg = "" if pd.isna(r["Subgrupo"]) else str(r["Subgrupo"]).strip()
        nivel = 0 if (not subg or subg == nome and not grupo) else (1 if sub else 2)
        if subg == nome and grupo and grupo != nome:
            nivel = 1
        if not subg or grupo == nome:
            nivel = 0
        linhas.append({"nome": nome.title(), "nivel": nivel, "grupo": grupo.title(),
                       "v": [orc, real, (real - orc) / 1000.0, pct(orc, real)]})
    vistos = {}
    for l in linhas:
        vistos[l["nome"]] = vistos.get(l["nome"], 0) + 1
    for l in linhas:
        if vistos[l["nome"]] > 1 and l["grupo"] and l["grupo"] != l["nome"]:
            l["nome"] = f"{l['nome']} ({l['grupo']})"
        l["total"] = l["nivel"] == 0        # compat: o render antigo lia `total`
        del l["grupo"]
    return linhas


def dre_mes(cc, modelo, ano, m, **kw):
    h = le_historico()
    if not h:
        return []
    g = h["geral"]
    df = g[(g["Centro de Custo"] == cc) & (g["Modelo"] == modelo)
           & (g["ano"] == ano) & (g["mes"] == m)]
    return _linhas_dre(df, "Orçado", "Realizado", **kw)


def dre_ytd(cc, modelo, ano, m, **kw):
    h = le_historico()
    if not h:
        return []
    y = h["ytd"]
    faixa = f"{m:02d}-Jan a {ABR[m-1]}"
    df = y[(y["Centro de Custo"] == cc) & (y["Modelo"] == modelo)
           & (y["ano"] == ano) & (y["Acumulado"] == faixa)]
    if df.empty:                       # a faixa é rotulada pelo mês final
        df = y[(y["Centro de Custo"] == cc) & (y["Modelo"] == modelo)
               & (y["ano"] == ano) & (y["mes"] == m) & (y["Acumulado"].str.startswith(f"{m:02d}-"))]
    return _linhas_dre(df, "Orçado YTD", "Realizado YTD", **kw)


def dre_grupo(cc, modelo, ano, m, grupo):
    h = le_historico()
    if not h:
        return []
    g = h["geral"]
    df = g[(g["Centro de Custo"] == cc) & (g["Modelo"] == modelo) & (g["ano"] == ano)
           & (g["mes"] == m) & (g["Grupo"] == grupo)]
    return _linhas_dre(df, "Orçado", "Realizado", so_com_valor=True)


# =========================================================== Investimentos (S09)
def slide_investimentos(m, ano):
    if not DRE_HARAS.exists():
        return pend(9, f"INVESTIMENTOS — COMENTÁRIOS {ano}", "", DRE_HARAS.name, "arquivo não encontrado")
    import openpyxl
    wb = openpyxl.load_workbook(DRE_HARAS, data_only=True, read_only=True)
    ws = wb["Investimentos"]
    BLOCOS = ("INFRAESTRUTURA", "COMPRA DE ANIMAIS E PRODUTOS", "MÁQUINAS E EQUIPAMENTOS",
              "INSTALAÇÕES", "FORMAÇÃO DE PASTAGEM")
    meses, atual, bl = [], None, None
    for r in ws.iter_rows(values_only=True):
        a = str(r[0]).strip().upper() if r[0] is not None else ""
        b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        v = num(r[2]) if len(r) > 2 else None
        if a.startswith("INVESTIMENTOS -"):
            nome = a.split("-", 1)[1].strip().split("/")[0].title()
            atual = {"mes": nome, "total": 0.0, "itens": []}
            meses.append(atual); bl = None
            continue
        if atual is None:
            continue
        if a in BLOCOS:
            bl = a
            if bl == "COMPRA DE ANIMAIS E PRODUTOS" and v:
                atual["total"] = v
            continue
        if bl == "COMPRA DE ANIMAIS E PRODUTOS" and v is not None:
            atual["itens"].append({"desc": b or a.title(), "valor": v})
    wb.close()
    # o slide é acumulado do ano: mostra de janeiro até o mês do deck
    idx = {nm.lower(): i + 1 for i, nm in enumerate(MESES)}
    meses = [x for x in meses if idx.get(x["mes"].lower(), 99) <= m]
    for x in meses:
        if not x["itens"]:
            x["itens"] = [{"desc": "Sem compra de animais e produtos registrada no mês", "valor": 0.0}]
    return {"t": "lista_mes", "n": 9, "titulo": f"INVESTIMENTOS — COMENTÁRIOS {ano}",
            "sub": f"Compra de animais e produtos · Jan–{ABR[m-1]}", "meses": meses}


# ============================================================ Plantel (S11/S12/S37)
# S11 — estoque em equinos. Regra do guia: status PLANTEL e sufixo EXATO
# 'DA PAO GRANDE' ou 'OUTRO'; variação com percentual ficava fora porque o
# animal dividido já aparece pela cota e contá-lo de novo o duplicaria.
SUFIXOS_S11 = ("DA PAO GRANDE", "OUTRO")
# Mudança pedida em 04/08/2026: o headcount passa a contar também o que é
# **100% do Eduardo**. No sufixo, o "E nn%" é a fatia dele — então entra só o
# `E 100%`, que é dele inteiro (cota 1,0, sem divisão). As fatias parciais
# (E 50%, E 25%, …) continuam fora, pelo mesmo motivo de sempre.
SUFIXOS_EXTRA_S11 = ("DA PAO GRANDE - E 100%",)
_base_bi_cache = None


def base_bi():
    global _base_bi_cache
    if _base_bi_cache is None and BASE_BI.exists():
        d = pd.read_parquet(BASE_BI)
        d["mes"] = pd.to_datetime(d["mes_referencia"]).dt.strftime("%Y-%m")
        _base_bi_cache = d
    return _base_bi_cache


def slide_estoque(m, ano):
    d = base_bi()
    if d is None:
        return pend(11, "ESTOQUE EM EQUINOS — FAZENDA PAO GRANDE", "", "bases/base_bi.parquet",
                    "rode python scripts/PGBaseBI.py")
    alvo = f"{ano}-{m:02d}"
    if alvo not in set(d["mes"]):
        ult = sorted(d["mes"].unique())[-1]
        aviso(f"base_bi vai até {ult} — S11 fica pendente nos meses seguintes")
        return pend(11, "ESTOQUE EM EQUINOS — FAZENDA PAO GRANDE", "", "bases/base_bi.parquet",
                    f"a base vai até {ult}; sem o mês {alvo}. Rode scripts/PGDataExtractor.py + PGBaseBI.py")
    x = d[(d["mes"] == alvo) & (d["status_plantel"] == "PLANTEL")
          & (d["sufixo_grupo"].isin(SUFIXOS_S11) | d["sufixo"].isin(SUFIXOS_EXTRA_S11))]
    n_eduardo = int((x["sufixo"].isin(SUFIXOS_EXTRA_S11)).sum())
    patrim = float(x["patrimonio_proporcional"].sum())
    aval = int(x["valor_100"].notna().sum())
    medio = float(x["valor_100"].mean()) if aval else 0.0
    cat = x["categoria"].value_counts()
    return {"t": "kpis_tabela", "n": 11, "titulo": "ESTOQUE EM EQUINOS — FAZENDA PAO GRANDE",
            "sub": (f"Composição patrimonial do plantel · {MESES[m-1].upper()} {ano} · {len(x)} animais"
                    f" · Status PLANTEL · Sufixo: Da PG / Outros"
                    + (f" / E 100% (Eduardo)" if n_eduardo else "")),
            "kpis": [{"v": f"{len(x)}", "l": "Animais Ativos",
                      "s": (f"Da PG + Outros + {n_eduardo} do Eduardo" if n_eduardo
                            else "DA PAO GRANDE + OUTROS")},
                     {"v": brl_curto(patrim), "l": "Patrimônio HPG", "s": "patrimônio proporcional"},
                     {"v": brl_curto(medio), "l": "Valor Médio", "s": f"{aval} animais avaliados"}],
            "tabela": {"cols": ["CATEGORIA", "Nº", "%"],
                       "rows": [[k.title(), int(v), f"{v/len(x)*100:.0f}%"] for k, v in cat.items()]}}


MOV_LINHAS = [("saldo_ini", "Saldo Inicial"), ("compra", "(+) Compras"), ("producao", "(+) Prod. Emb."),
              ("venda", "(-) Baixa Vendas"), ("morte", "(-) Baixa Mortes"), ("doacao", "(-) Doações"),
              ("reaval", "(±) Reavaliação"), ("saiu_controle", "(-) Saiu do Controle"),
              ("saldo_fim", "Saldo Final")]


def slide_movimentacao(m, ano):
    f = PLANTEL_DIR / "mov_cascata.parquet"
    if not f.exists():
        return pend(12, f"RESUMO DA MOVIMENTAÇÃO DO PLANTEL — {ano}", "", f.name,
                    "rode o LxMovimentacao.py no repo LuxorMonthlyP-CRoutines")
    c = pd.read_parquet(f)
    a = c[(c["mes"] >= f"{ano}-01") & (c["mes"] <= f"{ano}-{m:02d}")].sort_values("mes")
    if a.empty:
        ult = c["mes"].max()
        return pend(12, f"RESUMO DA MOVIMENTAÇÃO DO PLANTEL — {ano}", "", f.name,
                    f"a cascata vai até {ult}; sem meses de {ano} até {ABR[m-1]}")
    cols = ["TÍTULO"] + [ABR[int(x.split("-")[1]) - 1].upper() for x in a["mes"]]
    rows = [[rot] + [a[k].tolist()[i] for i in range(len(a))] for k, rot in MOV_LINHAS]
    u = a.iloc[-1]
    return {"t": "matriz", "n": 12, "titulo": f"RESUMO DA MOVIMENTAÇÃO DO PLANTEL — {ano}",
            "sub": "Saldo mensal · compras, produções, vendas e baixas",
            "kpis": [{"v": brl_curto(u["producao"]), "l": "Produção Emb.", "s": cols[-1]},
                     {"v": brl_curto(u["venda"]), "l": "Baixa Vendas", "s": cols[-1]},
                     {"v": brl_curto(u["morte"] + u["doacao"]), "l": "Mortes/Doações", "s": cols[-1]},
                     {"v": brl_curto(u["saldo_fim"]), "l": "Saldo Final", "s": "Haras PG"}],
            "cols": cols, "rows": rows}


def slide_contagem(m, ano):
    """S37 — contagem por local do MÊS do deck.

    A aba CONTAGEM do CONTROLE PLANTEL não serve aqui: ela é um retrato AO VIVO, sem
    dimensão de mês. Lendo direto dela, o deck de JUNHO/2026 exibia 203 animais
    (100/44/1/58) — a contagem de 14/08 — enquanto junho fechou com 206 (104/43/1/58).
    O rótulo dizia junho e o número era de agosto.

    O `base_bi` também não resolve: ele vem do CONTROLE_DE_PLANTEL mensal, cujo roster
    não reconcilia com a CONTAGEM (junho dá 221, com 88 sócios contra 58) e que quase
    não tem receptora, porque receptora mora no arquivo de ARRENDAMENTOS E RECEPTORAS.

    Quem tem o número certo E datado é o snapshot do fechamento semanal: usamos o
    último snapshot do mês pedido. Antes de 06/2026 não existe snapshot — aí a
    pendência é explícita, em vez de mostrar o número de outro mês."""
    from PGSemanalReport import HIST_SNAPSHOTS, HIST_HEADCOUNT
    prefixo = f"{ano}-{m:02d}"

    def _ler(f):
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            return {}

    # snapshot completo primeiro (tem a abertura animais/receptoras); o
    # headcount_history é mais magro mas cobre meses anteriores — junho/2026 só existe lá
    snaps, leves = _ler(HIST_SNAPSHOTS), _ler(HIST_HEADCOUNT)
    wids = sorted(w for w in snaps if w.startswith(prefixo))
    if wids:
        wid = wids[-1]
        det = snaps[wid].get("headcount_detalhe") or {}
        hc = snaps[wid].get("headcount") or {}
    else:
        wids = sorted(w for w in leves if w.startswith(prefixo))
        if not wids:
            todos = sorted(set(snaps) | set(leves))
            return pend(37, "PLANTEL — PAO GRANDE, ARRENDAMENTO E SÓCIOS",
                        f"{MESES[m-1].upper()} {ano}", "snapshot do fechamento semanal",
                        f"nenhum fechamento semanal guardado em {prefixo}; o histórico "
                        f"começa em {todos[0] if todos else '—'}")
        wid = wids[-1]
        det = {}
        # headcount_history usa nomes curtos
        hc = {"total": leves[wid].get("total"), "fazenda_pg": leves[wid].get("fpg"),
              "arrendamento": leves[wid].get("arr"), "cte": leves[wid].get("cte"),
              "socio": leves[wid].get("soc")}
    if det:
        ordem = [k for k in ("FAZENDA", "ARRENDAMENTO", "CTE", "SOCIO") if k in det]
        rows = [[k.title(), int(det[k]["animais"]), int(det[k]["receptoras"]),
                 int(det[k]["total"])] for k in ordem]
    else:
        # snapshot antigo, sem a abertura animais/receptoras. Preencher essas colunas
        # com zero afirmaria "nenhuma receptora"; melhor a tabela ter só o total.
        rows = [[l, int(hc.get(k) or 0)] for l, k in
                (("Fazenda", "fazenda_pg"), ("Arrendamento", "arrendamento"),
                 ("Cte", "cte"), ("Socio", "socio")) if hc.get(k) is not None]
    total = int(hc.get("total") or sum(r[-1] for r in rows))
    kp = [{"v": f"{int(r[-1])}", "l": r[0].upper(), "s": f"{r[-1]/total*100:.0f}% do total"}
          for r in rows[:3]]
    kp.append({"v": f"{total}", "l": "TOTAL GERAL", "s": "sob responsabilidade da PG"})
    cols = (["LOCAL", "ANIMAIS", "RECEPTORAS", "TOTAL"] if det else ["LOCAL", "TOTAL"])
    d, mm, aa = wid[8:10], wid[5:7], wid[:4]
    return {"t": "kpis_tabela", "n": 37, "titulo": "PLANTEL — PAO GRANDE, ARRENDAMENTO E SÓCIOS",
            "sub": f"{MESES[m-1].upper()} {ano} · fechamento semanal de {d}/{mm}/{aa}",
            "kpis": kp,
            "tabela": {"cols": cols, "rows": rows}}


# ============================================================ Estação (S16–S20)
def _estacao_wb():
    return _load(_latest_estacao_master())


def slides_estacao():
    """S16 funil, S17 garanhões, S18 comparativo, S19/S20 doadoras A e B.

    A estação é da SAFRA, não do mês — o mesmo conteúdo vale para qualquer mês do
    deck. Definições (do guia): absorção = perda antes dos 60d; aborto = embrião
    confirmado que não nasceu; óbito = nasceu e morreu.
    """
    try:
        src = _latest_estacao_master()
        wb = _load(src)
    except Exception as e:
        p = pend(16, "ESTAÇÃO DE MONTA — EMBRIÕES E PRENHEZES", "", "ESTACAO DE MONTA.xlsx",
                 f"não consegui abrir: {e}")
        return [p, dict(p, n=17, titulo="ESTAÇÃO DE MONTA — GARANHÕES"),
                dict(p, n=18, titulo="ESTAÇÃO DE MONTA — COMPARATIVO COM ANOS ANTERIORES"),
                dict(p, n=19, titulo="ESTAÇÃO DE MONTA — DOADORAS TIME A"),
                dict(p, n=20, titulo="ESTAÇÃO DE MONTA — DOADORAS TIME B")]
    out = [funil(wb), garanhoes(wb), comparativo(wb)] + doadoras(wb)
    wb.close()
    return out


def funil(wb):
    """S16 — aba ESTAÇÃO. Colunas (1-based): 11 LAVADO, 13 15D, 14 30D, 15 45D,
    16 60D, 17 ABORTO, 36 ESTAÇÃO. Confirmado = lavado+ e 15d+ e (30/45/60 '+' ou
    vazio), menos aborto=SIM."""
    ws = wb["ESTAÇÃO"]
    tent = lav = p15 = p30 = p45 = p60 = ab = 0
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 3 or r[0] is None or _s(r[35]) != SAFRA_ATUAL:
            continue
        tent += 1
        if _norm(r[10]) != "+":
            continue
        lav += 1
        # o funil é ENCADEADO: só chega em 30d quem passou em 15d, e assim por
        # diante. Contar cada coluna solta dava 85 confirmados contra 56 do
        # relatório oficial, porque a linha que nunca chegou aos 60 dias tem a
        # célula vazia e entrava como se tivesse passado.
        if _norm(r[12]) != "+":
            continue
        p15 += 1
        if _norm(r[13]) not in ("+", ""):
            continue
        p30 += 1
        if _norm(r[14]) not in ("+", ""):
            continue
        p45 += 1
        if _norm(r[15]) not in ("+", ""):
            continue
        p60 += 1
        if _norm(r[16]) == "SIM":     # aborto só conta depois de confirmado >60d
            ab += 1
    conf = p60 - ab
    ref = lambda v: f"{v/lav*100:.0f}% dos lavados" if lav else "—"
    return {"t": "kpis_tabela", "n": 16, "titulo": f"ESTAÇÃO DE MONTA {SAFRA_ATUAL} — EMBRIÕES E PRENHEZES",
            "sub": (f"{conf} embriões confirmados · taxa de recuperação {lav/tent*100:.0f}%"
                    f" · {tent} tentativas" if tent else "sem tentativas na safra"),
            "kpis": [{"v": str(conf), "l": "Embriões Conf.", "s": f"Estação {SAFRA_ATUAL}"},
                     {"v": str(lav), "l": "Lavados (+)", "s": f"{lav/tent*100:.0f}% de positivos" if tent else "—"},
                     {"v": f"{lav/tent*100:.0f}%" if tent else "—", "l": "Taxa Recup.", "s": SAFRA_ATUAL},
                     {"v": str(ab), "l": "Abortos", "s": "confirmados > 60d"}],
            "tabela": {"cols": [f"FUNIL DE PRENHEZ — ESTAÇÃO {SAFRA_ATUAL}", "Nº", "REFERÊNCIA"],
                       "rows": [["Tentativas", tent, "100%"],
                                ["Lavados (+)", lav, f"{lav/tent*100:.0f}%" if tent else "—"],
                                ["Prenhez 15d", p15, ref(p15)], ["Prenhez 30d", p30, ref(p30)],
                                ["Prenhez 45d", p45, ref(p45)], ["Prenhez 60d", p60, ref(p60)],
                                ["(−) Abortos", ab, "> 60 dias confirmados"],
                                ["Confirmados", conf, ref(conf)]]}}


def garanhoes(wb):
    """S17 — aba GARANHOES: 3 garanhão, 4 tipo de sêmen, 5 total lavados,
    6 lavados positivos, 7 %, 8 embriões confirmados, 9 prenhez, 10 aborto,
    11 total confirmados."""
    ws = wb["GARANHOES"]
    # no fim da aba há um bloco de legenda com os tipos de sêmen — ele entrava na
    # lista como se fosse garanhão ("Fresco", "Refrigerado", "Congelado")
    LEGENDA = {"FRESCO", "REFRIGERADO", "CONGELADO", "TIPO DE SEMEN", "TIPO DE SÊMEN"}
    rows, sem_uso = [], []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 4 or len(r) < 11 or r[2] is None:
            continue
        nome = _s(r[2])
        if not nome or _norm(r[2]).startswith("TOTAL") or _norm(r[2]) in LEGENDA:
            continue
        tot, conf = int(_to_num(r[4]) or 0), int(_to_num(r[10]) or 0)
        linha_ = [nome.title(), (_s(r[3]) or "")[:1].upper(), tot,
                  int(_to_num(r[5]) or 0), conf, f"{conf/tot*100:.0f}%" if tot else "—"]
        # garanhão sem nenhum lavado só ocupa espaço na tabela; vira nota de rodapé
        (rows if tot else sem_uso).append(linha_)
    rows.sort(key=lambda x: (-x[4], -x[2]))
    somas = [sum(r[i] for r in rows) for i in (2, 3, 4)]
    por_tipo = {}
    for r in rows:
        if not r[1]:
            continue
        a, b = por_tipo.get(r[1], (0, 0))
        por_tipo[r[1]] = (a + r[3], b + r[2])
    nome_tipo = {"R": "Refrigerado", "C": "Congelado", "F": "Fresco"}
    return {"t": "kpis_tabela", "n": 17, "titulo": f"ESTAÇÃO DE MONTA {SAFRA_ATUAL} — GARANHÕES",
            "sub": (f"{len(rows)} garanhões usados · {somas[0]} lavados · {somas[1]} positivos"
                    f" · {somas[2]} embriões confirmados · fonte: aba GARANHOES"),
            "kpis": [{"v": f"{p/t*100:.0f}%", "l": nome_tipo.get(k, k), "s": f"{p} de {t} lavados"}
                     for k, (p, t) in sorted(por_tipo.items(), key=lambda kv: -kv[1][1]) if t],
            "tabela": {"cols": ["GARANHÃO", "SÊMEN", "LAVADOS", "POSITIVOS", "CONFIRMADOS", "ÍND. %"],
                       "rows": rows},
            "obs": f"{len(sem_uso)} garanhões cadastrados sem lavado na safra ficaram fora da lista"
                   if sem_uso else None}


# Mês da estação de monta: começa em agosto e fecha em julho.
MESES_ESTACAO = [8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6, 7]


def comparativo(wb):
    """S18 — comparativo entre estações, CALCULADO da aba ESTAÇÃO linha a linha.

    A aba `COMPARATIVO` da planilha não serve: está congelada em 20/21–23/24 e o
    haras parou de manter. Como a aba ESTAÇÃO guarda a safra de cada embrião
    (coluna 36) e as datas, o comparativo sai da mesma base do funil (S16) — e
    então nunca fica velho, nem depende de alguém atualizar um resumo à mão.

    Mês de referência do embrião = IA + 60 dias, que é quando a prenhez é
    confirmada (mesma conta do fechamento semanal).
    """
    ws = wb["ESTAÇÃO"]
    por_safra = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 3 or r[0] is None or len(r) < 36:
            continue
        safra = _s(r[35])
        if not safra or "/" not in safra:
            continue
        d = por_safra.setdefault(safra, {"conf": {}, "aborto": 0, "absorcao": 0,
                                         "obito": 0, "lavados": 0, "tent": 0})
        d["tent"] += 1
        if _norm(r[10]) != "+":
            continue
        d["lavados"] += 1
        if _norm(r[12]) != "+":
            d["absorcao"] += 1        # perdeu antes da confirmação de 60 dias
            continue
        passou = all(_norm(r[j]) in ("+", "") for j in (13, 14, 15))
        if not passou:
            d["absorcao"] += 1
            continue
        if _norm(r[16]) == "SIM":
            d["aborto"] += 1          # confirmado > 60d que não nasceu
            continue
        if r[28] is not None:
            d["obito"] += 1           # nasceu e morreu — conta como confirmado
        ia = r[7]
        if not hasattr(ia, "month"):
            continue
        conf = ia + timedelta(days=60)
        d["conf"][conf.month] = d["conf"].get(conf.month, 0) + 1

    safras = sorted(por_safra, key=lambda s: s.split("/")[0])[-4:]
    if not safras:
        return pend(18, "ESTAÇÃO DE MONTA — COMPARATIVO ENTRE ESTAÇÕES", "",
                    "ESTACAO DE MONTA.xlsx, aba ESTAÇÃO", "nenhuma safra encontrada na coluna ESTAÇÃO")
    curto = [f"{s[2:4]}/{s[-2:]}" for s in safras]
    rows = []
    for m in MESES_ESTACAO:
        rows.append([ABR[m - 1]] + [por_safra[s]["conf"].get(m, 0) for s in safras])
    tot = {s: sum(por_safra[s]["conf"].values()) for s in safras}
    rows.append(["Confirmados", *[tot[s] for s in safras]])
    rows.append(["Abortos", *[por_safra[s]["aborto"] for s in safras]])
    rows.append(["Absorções", *[por_safra[s]["absorcao"] for s in safras]])
    rows.append(["Lavados (+)", *[por_safra[s]["lavados"] for s in safras]])
    rows.append(["Tentativas", *[por_safra[s]["tent"] for s in safras]])
    return {"t": "kpis_tabela", "n": 18, "titulo": "ESTAÇÃO DE MONTA — COMPARATIVO ENTRE ESTAÇÕES",
            "sub": ("Embriões confirmados por mês (IA + 60 dias) · calculado da aba ESTAÇÃO, "
                    "não da aba COMPARATIVO"),
            "kpis": [{"v": str(tot[s]), "l": f"Estação {c}", "s": f"{por_safra[s]['lavados']} lavados (+)"}
                     for s, c in zip(safras, curto)],
            "tabela": {"cols": ["MÊS"] + curto, "rows": rows}}


def doadoras(wb):
    """S19/S20 — meta × realizado por doadora. Meta e Time vêm do PLANEJAMENTO
    (7 TIME, 8 META TOTAL, 9 TOTAL EMBRIÕES); REC. EMBR. traz os lavados+."""
    ws = wb["PLANEJAMENTO"]
    por_time = {"A": [], "B": []}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 4 or r[1] is None:
            continue
        nome = _s(r[1])
        if not nome or _norm(r[1]).startswith("TOTAL"):
            continue
        time = _norm(r[6])
        if time not in ("A", "B"):
            continue
        meta, real = _to_num(r[7]) or 0, _to_num(r[8]) or 0
        por_time[time].append([nome.title(), int(meta), int(real),
                               f"{real/meta*100:.0f}%" if meta else "—"])
    lav = {}
    ws2 = wb["REC. EMBR."]
    for i, r in enumerate(ws2.iter_rows(values_only=True), 1):
        if i < 3 or r[2] is None:
            continue
        lav[_norm(r[2])] = _to_num(r[5]) or 0
    out = []
    for time in ("A", "B"):
        rows = por_time[time]
        for row in rows:
            row.insert(3, int(lav.get(_norm(row[0]), 0)))
        meta = sum(r[1] for r in rows)
        real = sum(r[2] for r in rows)
        lavp = sum(r[3] for r in rows)
        rows.sort(key=lambda x: -x[2])
        out.append({"t": "kpis_tabela", "n": 19 if time == "A" else 20,
                    "titulo": f"ESTAÇÃO DE MONTA {SAFRA_ATUAL} — DOADORAS TIME {time}",
                    "sub": f"{len(rows)} doadoras · meta {meta} embriões · realizado {real}",
                    "kpis": [{"v": str(meta), "l": "Meta", "s": f"Time {time}"},
                             {"v": str(real), "l": "Realizado", "s": "embriões confirmados"},
                             {"v": f"{real/meta*100:.0f}%" if meta else "—", "l": "Atingimento", "s": "real ÷ meta"},
                             {"v": str(lavp), "l": "Lavados (+)", "s": "aba REC. EMBR."}],
                    "tabela": {"cols": ["DOADORA", "META", "REAL", "LAV +", "%"],
                               "rows": [[r[0], r[1], r[2], r[3], r[4]] for r in rows]}})
    return out


# ========================================================= Coberturas (S21)
# O guia chama de "COBERTURAS_CAVALOS_FORA.xlsx"; no Drive o arquivo é
# REPRODUÇÃO/COBERTURAS - CAVALOS DE FORA NÃO USADAS.xlsx.
COBERTURAS = None            # resolvido em tempo de execução (ver _coberturas_path)
# Excluídos por decisão do haras (constam no guia).
COBERTURAS_FORA = ("TRILHO DA ZIZICA", "QUANTUM DE ALCATEIA")


def _coberturas_path():
    from _pg_common import DRIVE_ROOT
    p = DRIVE_ROOT / "REPRODUÇÃO" / "COBERTURAS - CAVALOS DE FORA NÃO USADAS.xlsx"
    return p if p.exists() else None


def slide_coberturas():
    """S21 — aba Planilha2: 2 garanhão, 3 qtd comprada, 4 utilizadas, 5 saldo.
    A Planilha1 é o log de compra (uma linha por negócio); a Planilha2 é o
    consolidado por garanhão, que é o que o slide mostra."""
    f = _coberturas_path()
    if f is None:
        return pend(21, f"ESTAÇÃO DE MONTA {SAFRA_ATUAL} — COBERTURAS DISPONÍVEIS",
                    "Saldo por garanhão de fora",
                    "REPRODUÇÃO/COBERTURAS - CAVALOS DE FORA NÃO USADAS.xlsx",
                    "arquivo não encontrado no Drive")
    wb = _load(_registra("coberturas de fora", f))
    ws = wb["Planilha2"]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 3 or r[1] is None:
            continue
        nome = _s(r[1])
        if not nome or _norm(r[1]).startswith("TOTAL") or _norm(r[1]) in COBERTURAS_FORA:
            continue
        qtd, usa, saldo = (int(_to_num(r[j]) or 0) for j in (2, 3, 4))
        rows.append([nome.title(), qtd, usa, saldo])
    wb.close()
    rows.sort(key=lambda x: -x[3])
    com_saldo = [r for r in rows if r[3] > 0]
    return {"t": "kpis_tabela", "n": 21,
            "titulo": f"ESTAÇÃO DE MONTA {SAFRA_ATUAL} — COBERTURAS DISPONÍVEIS",
            "sub": (f"Coberturas de garanhões de fora · {len(com_saldo)} garanhões com saldo"
                    f" · exclui {', '.join(x.title() for x in COBERTURAS_FORA)}"),
            "kpis": [{"v": str(sum(r[3] for r in rows)), "l": "Saldo Total", "s": "coberturas a usar"},
                     {"v": str(sum(r[1] for r in rows)), "l": "Compradas", "s": "no acumulado"},
                     {"v": str(sum(r[2] for r in rows)), "l": "Utilizadas", "s": "aceites"},
                     {"v": str(len(com_saldo)), "l": "Garanhões", "s": "com saldo disponível"}],
            "tabela": {"cols": ["GARANHÃO", "COMPRADAS", "UTILIZADAS", "SALDO"], "rows": rows}}


# ====================================================== Inadimplência (S31)
# O ControleInadimplencia.py já grava os agregados; não é preciso print do
# dashboard nem tocar em linha identificável — aqui só entram KPI e faixa etária,
# que não têm nome de devedor.
INAD_DIR = Path(r"G:/Drives compartilhados/Luxor Controladoria/Ambiente de testes"
                r"/Controle de inadimplência/output_pbi")


def slide_inadimplencia(m, ano):
    kpi_f, faixa_f = INAD_DIR / "indicadores_kpi.xlsx", INAD_DIR / "resumo_por_faixa.xlsx"
    if not kpi_f.exists() or not faixa_f.exists():
        return pend(31, "VENDAS — INADIMPLÊNCIAS E RECEBÍVEIS", f"Posição {ABR[m-1]}/{str(ano)[2:]}",
                    "controle-de-inadimplencia → output_pbi/indicadores_kpi.xlsx",
                    f"saída não encontrada em {INAD_DIR}; rode o ControleInadimplencia.py")
    k = pd.read_excel(_registra("inadimplência (KPI)", kpi_f)).iloc[0]
    fx = pd.read_excel(_registra("inadimplência (faixas)", faixa_f))
    ref = pd.to_datetime(k["data_referencia"]).strftime("%d/%m/%Y")
    venc = fx[fx["status_titulo"] == "Vencido"].groupby("faixa_atraso")[["valor_total", "qtd_titulos"]].sum()
    tot_venc = float(venc["valor_total"].sum()) or 1.0
    rows = [[str(i).split(" - ", 1)[-1].title(), int(r["qtd_titulos"]),
             brl_curto(r["valor_total"]), f"{r['valor_total']/tot_venc*100:.0f}%"]
            for i, r in venc.iterrows()]
    return {"t": "kpis_tabela", "n": 31, "titulo": "VENDAS — INADIMPLÊNCIAS E RECEBÍVEIS",
            "sub": (f"Posição de {ref} · agregados do ControleInadimplencia.py"
                    f" · sem dado identificável de devedor"),
            "kpis": [{"v": brl_curto(k["total_em_aberto"]), "l": "Em Aberto", "s": f"{int(k['qtd_clientes_total'])} clientes"},
                     {"v": brl_curto(k["total_vencido"]), "l": "Vencido",
                      "s": f"{k['percentual_vencido']:.1f}% da carteira"},
                     {"v": brl_curto(k["total_a_vencer"]), "l": "A Vencer", "s": "em dia"},
                     {"v": brl_curto(k["acao_judicial_total"]), "l": "Ação Judicial", "s": "total em cobrança"},
                     {"v": f"{int(k['qtd_clientes_vencidos'])}", "l": "Clientes Vencidos",
                      "s": f"ticket médio {brl_curto(k['ticket_medio'])}"}],
            "tabela": {"cols": ["FAIXA DE ATRASO", "TÍTULOS", "VALOR", "% DO VENCIDO"], "rows": rows},
            "obs": None if pd.to_datetime(k["data_referencia"]).month == m
                   else f"a base de cobrança está posicionada em {ref}"}


# ============================================================== Vendas (S29–S35)
VENDEDOR_COMITE = "CARLA"


def slides_vendas(m, ano, meta_anual=4_500_000):
    """S29 KPI e S30 detalhamento. Filtro obrigatório: VENDEDOR = CARLA, sem
    CANCELADO (regra do guia). Colunas do MAPA VENDAS (1-based): 7 valor produto,
    8 valor da venda, 11 tipo de evento, 12 nome do evento, 15 vendedor,
    19 status contrato, 22 ano, 23 mês."""
    try:
        src = _latest_by_yymmdd(MAPA_VENDAS_DIR, "*_PG_Mapa Vendas.xlsx", "mapa de vendas")
        wb = _load(src)
    except Exception as e:
        p = pend(29, "VENDAS — RESULTADO ACUMULADO", "", "PG_Mapa Vendas.xlsx, aba MAPA VENDAS",
                 f"não consegui abrir: {e}")
        return [p, dict(p, n=30, titulo="VENDAS — DETALHAMENTO POR MÊS E EVENTO")]
    ws = wb["MAPA VENDAS"]
    por_mes = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 3 or len(r) < 23 or r[21] is None:
            continue
        # a coluna traz o nome completo do vendedor ("CARLA ...") — comparar por
        # igualdade zerava o slide inteiro
        if VENDEDOR_COMITE not in _norm(r[14]):
            continue
        if "CANCELAD" in _norm(r[18]):
            continue
        a, mm = _to_num(r[21]), _to_num(r[22])
        if a != ano or not mm or mm > m:
            continue
        v = _to_num(r[7]) or 0
        ev = _s(r[10]) or _s(r[11]) or "—"
        por_mes.setdefault(int(mm), {}).setdefault(ev.title(), 0)
        por_mes[int(mm)][ev.title()] += v
    wb.close()
    ytd = sum(sum(d.values()) for d in por_mes.values())
    mes_v = sum(por_mes.get(m, {}).values())
    rows = []
    for mm in sorted(por_mes):
        tot = sum(por_mes[mm].values())
        rows.append([MESES[mm - 1].upper(), "", tot])
        for ev, v in sorted(por_mes[mm].items(), key=lambda kv: -kv[1]):
            rows.append(["", ev, v])
    return [
        {"t": "kpis_tabela", "n": 29, "titulo": f"VENDAS {ano} — RESULTADO ACUMULADO — {VENDEDOR_COMITE.title()}",
         "sub": f"Meta anual {brl_curto(meta_anual)} · acumulado Jan–{ABR[m-1]} {brl_curto(ytd)}",
         "kpis": [{"v": brl_curto(mes_v), "l": f"Vendas {ABR[m-1]}", "s": "realizado no mês"},
                  {"v": brl_curto(ytd), "l": "Acumulado YTD", "s": f"Jan–{ABR[m-1]}"},
                  {"v": brl_curto(meta_anual), "l": "Meta Anual", "s": "parâmetro do comitê"},
                  {"v": brl_curto(meta_anual - ytd), "l": "Saldo para Meta",
                   "s": f"{ytd/meta_anual*100:.0f}% atingido"}],
         "tabela": {"cols": ["MÊS", "TOTAL"],
                    "rows": [[MESES[mm - 1].title(), brl_curto(sum(por_mes[mm].values()))]
                             for mm in sorted(por_mes)]}},
        {"t": "tabela", "n": 30, "titulo": f"VENDAS — JANEIRO A {MESES[m-1].upper()}/{str(ano)[2:]} — {VENDEDOR_COMITE.title()}",
         "sub": f"Detalhamento por mês e evento · filtro: vendedor = {VENDEDOR_COMITE}, sem cancelados",
         "cols": ["MÊS", "EVENTO / ORIGEM", "VALOR"], "moeda": [2],
         "rows": rows},
    ]


# ENTREGAR (1-based): 2 doadora, 3 garanhão, 4 data venda, 6 comprador, 7 cota,
# 11 valor, 12 status pgto, 13 status embrião.  RECEBER: 6 vendedor.
S32 = ("QUITADO", "PAGANDO")
S33 = ("PAUSAD", "APOS CONF", "APÓS CONF")
S34 = ("DIREITO", "TROCA")


def slides_embrioes():
    try:
        wb = _load(_registra("embriões a entregar", EMB_COMERCIAIS))
    except Exception as e:
        base = pend(32, "VENDAS — EMBRIÕES VENDIDOS A FAZER", "", EMB_COMERCIAIS.name,
                    f"não consegui abrir: {e}")
        return [base, dict(base, n=33), dict(base, n=34),
                dict(base, n=35, titulo="ESTAÇÃO DE MONTA — EMBRIÕES COMPRADOS A RECEBER")]

    def linhas(aba, col_contraparte):
        out = []
        for i, r in enumerate(wb[aba].iter_rows(values_only=True), 1):
            if i < 4 or len(r) < 13 or r[1] is None:
                continue
            out.append({"doadora": _s(r[1]), "garanhao": _s(r[2]), "data": r[3],
                        "contraparte": _s(r[col_contraparte]), "cota": _to_num(r[6]),
                        "valor": _to_num(r[10]), "pgto": _s(r[11]) or "",
                        "status": _s(r[12]) or ""})
        return out

    ent = linhas("ENTREGAR", 5)
    rec = linhas("RECEBER", 5)
    wb.close()
    af = lambda x: _norm(x["status"]).startswith("A FAZER")
    tem = lambda x, ks: any(k in _norm(x["pgto"]) for k in ks)

    def tabela(n, titulo, sel, itens, rotulo_contra="COMPRADOR"):
        lst = [x for x in itens if sel(x)]
        tot = sum(x["valor"] or 0 for x in lst)
        return {"t": "tabela", "n": n, "titulo": titulo,
                "sub": f"{len(lst)} contrato(s) · {brl_curto(tot)}",
                "cols": ["DOADORA", "GARANHÃO", "DATA", rotulo_contra, "CT", "VALOR", "PGTO"],
                "moeda": [5], "data": [2],
                "rows": [[x["doadora"], x["garanhao"], x["data"], x["contraparte"],
                          f"{x['cota']*100:.0f}%" if x["cota"] else "—", x["valor"], x["pgto"]]
                         for x in sorted(lst, key=lambda y: -(y["valor"] or 0))]}

    return [
        tabela(32, "VENDAS — EMBRIÕES VENDIDOS A FAZER (QUITADO / PAGANDO)",
               lambda x: af(x) and tem(x, S32), ent),
        tabela(33, "VENDAS — EMBRIÕES VENDIDOS A FAZER (PGTO PAUSADO / APÓS CONF.)",
               lambda x: af(x) and tem(x, S33), ent),
        tabela(34, "VENDAS — EMBRIÕES DE DIREITO / REPOSIÇÃO",
               lambda x: _norm(x["status"]).startswith("REPOSI") or (af(x) and tem(x, S34)), ent),
        tabela(35, "ESTAÇÃO DE MONTA — EMBRIÕES COMPRADOS A RECEBER", af, rec, "VENDEDOR"),
    ]


# ============================= Conteúdo escrito à mão (S08, S23–S27, S38, S39)
# Comentário do DRE, exposição, manejo e foto não saem de planilha: são escritos
# todo mês. Ficam em `_docs/comite_conteudo.json`, semeado do último deck
# aprovado por `tools/extrair_conteudo.py`. Sem isso esses slides seriam
# placeholder pra sempre.
CONTEUDO = REPO / "_docs" / "comite_conteudo.json"
FALTA_CONTEUDO = "escreva o conteúdo desse mês em _docs/comite_conteudo.json"


def le_conteudo():
    if not CONTEUDO.exists():
        return {}
    d = json.loads(CONTEUDO.read_text(encoding="utf-8"))
    return {k: v for k, v in d.items() if re.fullmatch(r"\d{4}-\d{2}", k)}


def conteudo_do_mes(todos, chave):
    """Só conteúdo do próprio mês — puxar de um mês futuro colocaria no deck de
    janeiro a exposição que ainda não tinha acontecido."""
    return todos.get(chave, {})


def slide_comentarios(c, m, ano):
    itens = c.get("comentarios") or []
    if not itens:
        return pend(8, f"COMENTÁRIOS — VARIAÇÕES YTD JAN–{ABR[m-1].upper()} {ano}",
                    "Principais destaques acumulados por categoria",
                    "_docs/comite_conteudo.json → comentarios", FALTA_CONTEUDO)
    return {"t": "comentarios", "n": 8,
            "titulo": f"COMENTÁRIOS — VARIAÇÕES YTD JAN–{ABR[m-1].upper()} {ano}",
            "sub": "DRE 2026 | HPG · principais destaques acumulados por categoria",
            "itens": itens}


def slides_exposicoes(c, ano):
    exp = c.get("exposicoes") or {}
    prog, res = exp.get("programacao") or [], exp.get("resultados") or []
    out = []
    if prog:
        out.append({"t": "tabela", "n": 23, "titulo": f"EXPOSIÇÕES {ano} — PROGRAMAÇÃO",
                    "sub": "Calendário de participações previstas",
                    "cols": ["EVENTO", "DATA", "LOCAL", "STATUS"], "rows": prog})
    else:
        out.append(pend(23, f"EXPOSIÇÕES {ano} — PROGRAMAÇÃO", "Calendário de participações",
                        "_docs/comite_conteudo.json → exposicoes.programacao", FALTA_CONTEUDO))
    if res:
        for k, r in enumerate(res):
            out.append({"t": "resultados", "n": 24 + k, "titulo": r["titulo"],
                        "sub": r.get("sub", ""), "animais": r["animais"]})
    else:
        out.append(pend(24, "RESULTADOS DAS EXPOSIÇÕES", "Animais, títulos e colocações",
                        "_docs/comite_conteudo.json → exposicoes.resultados", FALTA_CONTEUDO))
    return out


def slide_manejo(c, m, ano):
    itens = c.get("manejo") or []
    if not itens:
        return pend(38, "MANEJO — PONTOS DE MELHORIA E DECISÕES", "Histórico de intervenções",
                    "_docs/comite_conteudo.json → manejo", FALTA_CONTEUDO)
    return {"t": "manejo", "n": 38, "titulo": "MANEJO — PONTOS DE MELHORIA E DECISÕES",
            "sub": f"Histórico de intervenções Jan–{ABR[m-1]} {ano}", "itens": itens}


# 6 fotos por slide: mais que isso e cada foto vira selo; menos, sobra tela.
FOTOS_POR_SLIDE = 6
# As fotos NAO ficam no repo (publico) nem no site (publico): entram embutidas no
# spec.json, que sai pelo bucket privado. Cru sao 12 MB em 28 arquivos; reduzidas
# pro tamanho em que o deck as mostra (6 por slide, ~1/3 de tela) dao ~2 MB.
FOTO_LADO_MAX = 760
FOTO_QUALIDADE = 68


def _foto_embutida(caminho: Path) -> str | None:
    """Foto como data URI reduzida. None se nao der pra ler — slide vira pendencia
    em vez de imagem quebrada, que e o que acontecia com caminho relativo depois
    que as fotos sairam do repo."""
    try:
        from PIL import Image
    except ImportError:
        print("  [fotos] Pillow ausente (pip install Pillow) — fotos ficam de fora")
        return None
    try:
        im = Image.open(caminho)
        im.thumbnail((FOTO_LADO_MAX, FOTO_LADO_MAX))
        if im.mode != "RGB":
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, "JPEG", quality=FOTO_QUALIDADE, optimize=True)
        return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
    except Exception as exc:
        print(f"  [fotos] {caminho.name}: {exc!r}")
        return None


def slides_fotos(c, m, ano):
    fs = c.get("fotos") or []
    if not fs:
        return [pend(39, "MANEJO — FOTOS E REGISTROS", f"Registros de {MESES[m-1]}",
                     "_docs/comite_conteudo.json → fotos (assets/comite/fotos/)",
                     "rode tools/extrair_conteudo.py ou solte as fotos do mês na pasta")]
    # troca o nome do arquivo pela imagem embutida; some quem nao carregou
    embutidas, kb = [], 0
    for f in fs:
        uri = _foto_embutida(OUT / f)   # f ja vem como "fotos/fotoNN.jpg"
        if uri:
            embutidas.append(uri)
            kb += len(uri) // 1024
    if not embutidas:
        return [pend(39, "MANEJO — FOTOS E REGISTROS", f"Registros de {MESES[m-1]}",
                     "assets/comite/fotos/ (fora do repo, por serem imagens da fazenda)",
                     "solte as fotos do mes na pasta local")]
    print(f"  [fotos] {len(embutidas)} embutidas no spec ({kb // 1024 or 1} MB)")
    fs = embutidas
    n = (len(fs) + FOTOS_POR_SLIDE - 1) // FOTOS_POR_SLIDE
    return [{"t": "fotos", "n": 39, "titulo": "MANEJO — FOTOS E REGISTROS",
             "sub": f"Registros de {MESES[m-1]} {ano}" + (f" · {k+1}/{n}" if n > 1 else ""),
             "fotos": fs[k * FOTOS_POR_SLIDE:(k + 1) * FOTOS_POR_SLIDE]} for k in range(n)]


# ==================================================================== deck
def divisor(n, titulo, sub):
    return {"t": "divisor", "n": n, "titulo": titulo, "sub": sub}


# Linhas por slide. Com 26 a linha fica em ~20px na tela de 720 — dá pra ler
# sentado longe, que é o ponto de uma apresentação. Passou disso, o slide QUEBRA
# em continuação, como o PowerPoint faria. Espremer tudo numa página só foi o
# que deixou a lista ilegível e vazando por cima do rodapé.
MAX_LINHAS = 26


def divide_tab(slide):
    """Mesma quebra, para o slide de KPIs + tabela (as linhas moram em .tabela)."""
    t = slide.get("tabela")
    if not t or len(t["rows"]) <= MAX_LINHAS - 6:
        return [slide]
    linhas, out = t["rows"], []
    passo = MAX_LINHAS - 6
    n = (len(linhas) + passo - 1) // passo
    for k in range(n):
        p = dict(slide)
        p["tabela"] = dict(t, rows=linhas[k*passo:(k+1)*passo])
        p["titulo"] = f"{slide['titulo']} ({k+1}/{n})"
        out.append(p)
    return out


def divide(slide, campo="linhas"):
    """Devolve [slide] ou a lista de slides '(cont.)' quando a tabela é longa."""
    linhas = slide.get(campo) or []
    if len(linhas) <= MAX_LINHAS:
        return [slide]
    partes, n = [], (len(linhas) + MAX_LINHAS - 1) // MAX_LINHAS
    for k in range(n):
        p = dict(slide)
        p[campo] = linhas[k * MAX_LINHAS:(k + 1) * MAX_LINHAS]
        if k:
            p["titulo"] = f"{slide['titulo']} (cont. {k + 1}/{n})"
        else:
            p["titulo"] = f"{slide['titulo']} (1/{n})"
        partes.append(p)
    return partes


def monta_deck(m, ano, ctx):
    cont = conteudo_do_mes(ctx["conteudo"], f"{ano}-{m:02d}")
    s = [
        {"t": "capa", "titulo": "RELATÓRIO DE DESEMPENHO ESTRATÉGICO",
         "mes": f"{MESES[m-1].upper()} / {ano}", "org": "HARAS PAO GRANDE"},
        {"t": "agenda", "titulo": "AGENDA",
         "sub": f"RELATÓRIO DESEMPENHO ESTRATÉGICO — {ABR[m-1].upper()}/{str(ano)[2:]}",
         "itens": [{"n": "01", "titulo": "FINANCEIRO", "sub": "DRE Haras · Caixa · Plantel"},
                   {"n": "02", "titulo": "ESTAÇÃO DE MONTA", "sub": "Embriões · Doadoras · Garanhões"},
                   {"n": "03", "titulo": "EXPOSIÇÕES", "sub": "Programação e resultados"},
                   {"n": "04", "titulo": "VENDAS", "sub": "Pipeline e contratos"},
                   {"n": "05", "titulo": "DECISÕES E MANEJO", "sub": "Plantel · Obras · Casa"}]},
        divisor(1, "FINANCEIRO", f"DRE Haras · Caixa · Plantel | {MESES[m-1].upper()} {ano}"),
    ]
    FONTE = "Fonte: DRE_Historico.xlsx (Base DRE Geral)"
    mesano = f"{ABR[m-1].upper()}/{str(ano)[2:]}"
    dre = lambda n, t, sub, lin, fonte=None: (
        {"t": "dre", "n": n, "titulo": t, "sub": f"{sub} · {fonte or FONTE}", "linhas": lin}
        if lin else pend(n, t, sub, "DRE_Historico.xlsx", "sem linha para esse recorte no histórico"))

    s += divide(dre(4, f"RESUMO FINANCEIRO — HARAS COMPETÊNCIA — ORÇADO X REALIZADO {mesano}",
                    "DRE 2026 | HPG · competência mensal",
                    dre_mes("HPG", "Competência", ano, m, so_subtotal=True)))
    s += divide(dre(5, f"ANÁLISE DE CUSTOS — {MESES[m-1].upper()} {ano}",
                    "Custos indiretos de produção · linhas zeradas no mês omitidas",
                    dre_grupo("HPG", "Competência", ano, m, "CUSTOS E DESPESAS OPERACIONAIS")))
    s += divide(dre(6, f"ANÁLISE DE DESPESAS — {MESES[m-1].upper()} {ano}",
                    "Despesas do mês · linhas zeradas no mês omitidas",
                    dre_grupo("HPG", "Competência", ano, m, "DESPESAS")))
    s += divide(dre(7, f"HARAS COMPETÊNCIA — ACUMULADO JAN–{ABR[m-1].upper()} {ano} (YTD)",
                    "DRE 2026 | HPG · acumulado no ano",
                    dre_ytd("HPG", "Competência", ano, m, so_subtotal=True),
                    "Fonte: DRE_Historico.xlsx (Base YTD)"))
    s.append(slide_comentarios(cont, m, ano))
    s.append(slide_investimentos(m, ano))
    s += divide(dre(10, f"HARAS CAIXA — ORÇADO X REALIZADO {mesano}", "FC 2026 | HPG · caixa mensal",
                    dre_mes("HPG", "Caixa", ano, m, so_subtotal=True)))
    s.append(slide_estoque(m, ano))
    s.append(slide_movimentacao(m, ano))
    s += divide(dre(13, f"RESUMO FINANCEIRO — CASA/FPG — ORÇADO X REALIZADO {mesano}",
                    "FPG | Casa · caixa mensal", dre_mes("FPG", "Caixa", ano, m, so_com_valor=True)))
    s += divide(dre(14, f"CASA/FPG — ORÇADO X REALIZADO ACUMULADO JAN–{ABR[m-1].upper()} {ano}",
                    "FPG | Casa · acumulado no ano", dre_ytd("FPG", "Caixa", ano, m, so_com_valor=True),
                    "Fonte: DRE_Historico.xlsx (Base YTD)"))

    s.append(divisor(2, "ESTAÇÃO DE MONTA", "Embriões · Doadoras · Garanhões"))
    for x in ctx["estacao"]:
        s += divide_tab(x)
    s += divide_tab(ctx["coberturas"])

    s.append(divisor(3, "EXPOSIÇÕES", "Programação e resultados"))
    for x in slides_exposicoes(cont, ano):
        s += divide(x, 'rows') if x["t"] == "tabela" else [x]

    s.append(divisor(4, "VENDAS", "Pipeline e contratos"))
    for x in slides_vendas(m, ano):
        s += divide(x, 'rows')
    s += divide_tab(slide_inadimplencia(m, ano))
    for x in ctx["embrioes"]:
        s += divide(x, 'rows')

    s.append(divisor(5, "DECISÕES E MANEJO", "Plantel · Obras · Casa"))
    s.append(slide_contagem(m, ano))
    s.append(slide_manejo(cont, m, ano))
    s += slides_fotos(cont, m, ano)
    s.append({"t": "encerramento", "titulo": "HARAS PAO GRANDE"})
    return s


def build(so_mes=None):
    ano = so_mes.year if so_mes else 2026
    ctx = {}
    if _dre_hist() is None:
        aviso("DRE_Historico.xlsx não encontrado em nenhuma das cópias conhecidas "
              f"({' | '.join(str(c.parent) for c in DRE_HIST_CANDIDATOS)}) "
              "— seção financeira fica pendente")
        meses = []
    else:
        meses = meses_fechados(ano=ano)
        if meses:
            print(f"  [dre] meses fechados em {ano}: "
                  f"{', '.join(ABR[x-1] for x in meses)}")
    if not meses:
        meses = [so_mes.month] if so_mes else [date.today().month]
        aviso("nenhum mês com realizado no DRE — deck sai só com as bases não-financeiras")

    ctx["estacao"] = slides_estacao()
    ctx["coberturas"] = slide_coberturas()
    ctx["conteudo"] = le_conteudo()
    ctx["embrioes"] = slides_embrioes()

    alvo = [so_mes.month] if so_mes else meses
    decks = {}
    for m in alvo:
        decks[f"{ano}-{m:02d}"] = monta_deck(m, ano, ctx)

    chaves = sorted(decks)
    # As fontes que os resolvedores compartilhados registram (roster, receptoras,
    # controle mensal) ficam em PGSemanalReport._FONTES_USADAS; as deste módulo, em
    # _FONTES. A auditoria quer as duas na mesma lista.
    fontes = {}
    for rotulo, caminho in list(_FONTES_COMPARTILHADAS.items()) + list(_FONTES.items()):
        p = Path(caminho)
        try:
            quando = datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="minutes")
        except OSError:
            quando = None
        fontes[rotulo] = {"arquivo": p.name, "caminho": caminho_curto(p),
                          "pasta": p.parent.name, "modificado": quando}

    payload = {"meses": chaves, "padrao": chaves[-1], "avisos": avisos,
               "labels": {k: f"{MESES[int(k[5:]) - 1]} {k[:4]}" for k in chaves},
               "fontes": fontes, "decks": decks}
    OUT.mkdir(parents=True, exist_ok=True)
    js = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=_json_default)
    (OUT / "spec.json").write_text(js, encoding="utf-8")
    (OUT / "spec.js").write_text(f"window.COMITE_SPEC = {js};\n", encoding="utf-8")
    n = len(decks[chaves[-1]])
    p = sum(1 for x in decks[chaves[-1]] if x["t"] == "pendente")
    print(f"[comite] {len(chaves)} meses ({chaves[0]} … {chaves[-1]}) · {n} slides "
          f"({n - p} com conteúdo, {p} pendentes) · {len(js)//1024} KB -> assets/comite/spec.js")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            mm, aaaa = sys.argv[1].split("/")
            build(date(int(aaaa), int(mm), 1))
        except ValueError:
            sys.exit(f"mês inválido: {sys.argv[1]!r} — use MM/AAAA (ex.: 06/2026)")
    else:
        build()
