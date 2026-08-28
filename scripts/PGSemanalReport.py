"""
PGSemanalReport — gera a ATUALIZAÇÃO SEMANAL do Haras Pão Grande a partir das
fontes REORGANIZADAS do Drive (layout 2026-07).

Substitui o pipeline antigo (PGSemanalExtractor/_pg_semanal), que apontava pra
paths e modelos de planilha que não existem mais.

Fontes (ver memória project-hpg-semanal):
  - Headcount   : ATUALIZACAO SEMANAL/CONTROLE PLANTEL.xlsx  aba CONTAGEM (pré-agregada)
  - Receptoras  : idem, aba 'RECEPTORAS ' (censo; filtro do report a definir)
  - Produção    : REPRODUÇÃO/ESTAÇÃO DE MONTA/Estação 2025-2026/{YYMMDD} ESTACAO DE MONTA.xlsx
                  aba ESTAÇÃO (embrião confirmado = coluna 60D == '+')
  - Movimentação: PLANILHAS SEMANAIS/SAIDA E ENTRADA DE ANIMAIS - MODELO ENVIAR NO GRUPO.xlsx
  - Comerciais  : REPRODUÇÃO/EMBRIOES A ENTREGAR - A RECEBER.xlsx  abas PAINEL/ENTREGAR/RECEBER

Uso:
    python PGSemanalReport.py                 # semana termina hoje
    python PGSemanalReport.py 17/07/2026       # semana termina nessa data
    python PGSemanalReport.py 03/07/2026 17/07/2026   # janela explícita

Saída:
    - imprime as 5 seções no console, comparadas aos alvos do docx 17-07 (quando aplicável)
    - grava semanal_data.json (consumido pelo dashboard HTML)
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from _pg_common import DRIVE_ROOT, ensure_cache

BASE_DIR = Path(__file__).resolve().parent.parent  # raiz do projeto (scripts/ fica 1 nível abaixo)

# ------------------------------------------------------------------
# Localização das fontes (layout novo do Drive)
# ------------------------------------------------------------------
# ATUALIZACAO SEMANAL é a pasta de SAÍDA do fechamento — é onde o relatório em Word é
# publicado. Puxar dado de lá é ler do lugar errado: as três planilhas que moram ali
# pertencem, por assunto, a PLANTEL, REPRODUÇÃO e VENDAS.
#
# Foi feita a varredura por substituto de CONTEÚDO e não existe: o acumulado de 61 não
# aparece em nenhuma aba da estação de monta (PLANEJAMENTO e REC. EMBR. dão 56, RESUMO
# está com as fórmulas em #REF!) nem no controle mensal (aba EMBRIOES PG é financeira;
# os 65 embriões do roster estão com a coluna SAFRA vazia); e pendência de sociedade
# não é marcada em lugar nenhum — o MAPA VENDAS é histórico de vendas, sem status.
#
# Então o que falta é MUDAR OS ARQUIVOS DE PASTA, não trocar de fonte. Cada caminho
# abaixo é uma LISTA em ordem de preferência: a pasta canônica primeiro, ATUALIZACAO
# SEMANAL só como último recurso e com aviso no run. No dia em que o haras mover os
# arquivos, o pipeline segue sozinho — e enquanto não mover, o run diz o que falta.
FALLBACK_DIR = DRIVE_ROOT / "ATUALIZACAO SEMANAL"

# Roster do plantel: assunto é PLANTEL.
CONTROLE_PLANTEL_DIRS = (DRIVE_ROOT / "PLANTEL", FALLBACK_DIR)
CONTROLE_PLANTEL_GLOB = "CONTROLE PLANTEL.xlsx"
# "EMBRIÕES E MATRIZES - MODELO ENVIAR NO GRUPO 3 <DD-MM>.xlsx" — fonte do acumulado na
# estação. Assunto é REPRODUÇÃO; já morou em PLANILHAS SEMANAIS. Sufixo de data muda
# (e mente: a cópia viva se chama "30-05" e é de agosto), então resolve por glob+mtime.
EMB_MATRIZES_DIRS = (
    DRIVE_ROOT / "REPRODUÇÃO" / "ESTAÇÃO DE MONTA",
    DRIVE_ROOT / "REPRODUÇÃO",
    FALLBACK_DIR,
)
EMB_MATRIZES_GLOB = "EMBRI*E MATRIZES*.xlsx"
# Animais para sair (hoje só sociedade pendente) — assunto é VENDAS. A pasta
# PLANILHAS PARA O EDUARDO saiu da lista: não tem nenhum arquivo desse nome.
ANIMAIS_SAIR_DIRS = (
    DRIVE_ROOT / "VENDAS" / "SAIDA DE ANIMAIS VENDIDOS",
    FALLBACK_DIR,
)
ANIMAIS_SAIR_GLOB = "Animais para sair*.xlsx"

EMB_COMERCIAIS = DRIVE_ROOT / "REPRODUÇÃO" / "EMBRIOES A ENTREGAR - A RECEBER.xlsx"
ESTACAO_MONTA_BASE = DRIVE_ROOT / "REPRODUÇÃO" / "ESTAÇÃO DE MONTA"
# O plantel e as receptoras vivem em PLANTEL/Estação <ano>-<ano>, e a copia de
# trabalho MUDA DE PASTA quando a estacao vira: em 21/08/2026 os arquivos "EDITAR
# SETEMBRO" passaram para "Estação 2026-2027", porque setembro abre estacao nova.
# Olhar so a pasta da estacao corrente fez o pipeline concluir que os arquivos tinham
# sido apagados e cair numa copia congelada de 05/08. Varremos TODAS as pastas de
# estacao e ficamos com o mais recente — aqui frescor e o que importa, e a guarda de
# fonte velha cobre o resto.
PLANTEL_DIR_BASE = DRIVE_ROOT / "PLANTEL"
PLANTEL_ESTACAO_GLOB = "Estação *"
RECEPTORAS_DIR = PLANTEL_DIR_BASE / "Estação 2025-2026"   # so p/ mensagens de erro
# Mapa de Vendas — quem consome é o deck do comitê (tools/build_comite.py),
# não o fechamento semanal. Cheguei a tratar como constante morta por não achar
# uso neste módulo; o uso está no outro.
MAPA_VENDAS_DIR = DRIVE_ROOT / "VENDAS" / "MAPAS DE VENDAS" / "Estação 2025-2026"
# CONTROLE_DE_PLANTEL mensal (STATUS PLANTEL, SAIDAS-ENTRADAS, MOVIMENTAÇÕES)
CONTROLE_MENSAL_DIR = RECEPTORAS_DIR

HIST_HEADCOUNT = BASE_DIR / "_cache" / "headcount_history.json"
HIST_SNAPSHOTS = BASE_DIR / "_cache" / "semanal_snapshots.json"

SAFRA_ATUAL = "2025/2026"
# Transicao de estacao: o relatorio publica as duas linhas enquanto a safra nova
# nao anda. Nao ha nada de 2026/2027 nas planilhas ainda — o numero nasce zerado e
# comeca a andar sozinho quando o haras lancar a primeira IA.
SAFRA_PROXIMA = "2026/2027"


def _rotulo_safra(safra: str) -> str:
    """'2025/2026' -> '25/26', que e como o relatorio escreve."""
    a, b = safra.split("/")
    return f"{a[-2:]}/{b[-2:]}"
BASES_DIR = BASE_DIR / "bases"
JSON_OUT = BASES_DIR / "semanal_data.json"

# Alvos do docx 17-07-26 (semana 03/07-17/07) — só pra validação em tela
DOCX_1707 = {
    "acumulado_estacao": 61,
    "confirmados_semana": 2,
    "nascimentos": 1,
    "abortos_obitos": 1,
    "receptoras_total": 63,
    "receptoras_prenhas": 36,
    "receptoras_vazias": 27,
    "headcount_total": 206,
    "headcount_fpg": 104,
    "headcount_arr": 43,
    "headcount_cte": 1,
    "headcount_soc": 58,
    "saidas_semana": 8,
    "vendidos_pendentes": 2,
    "sociedade_pendentes": 2,
}


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def _norm(v) -> str:
    return str(v).strip().upper() if v is not None else ""


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _dt(v):
    if isinstance(v, datetime):
        return v.date()
    return None


def _to_num(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _load(path: Path):
    cached = ensure_cache(path)
    return openpyxl.load_workbook(cached, data_only=True, read_only=True)


# {rotulo: Path} das fontes efetivamente abertas neste run — base do aviso de fonte
# velha. Sem isso, apagar a copia de trabalho faz o pipeline cair numa versao
# congelada semanas atras SEM UM RUIDO: em 21/08/2026 os dois arquivos "EDITAR
# SETEMBRO" foram apagados no meio do dia e o fechamento passou a ler receptoras de
# 05/08 e roster mensal de 13/08, publicando headcount 205 no lugar de 202.
_FONTES_USADAS: dict = {}


# Pasta de DIVULGAÇÃO, não de dado: é onde o haras deixa o que foi enviado ao grupo,
# e serve para conferir o calculado contra o publicado. Três arquivos vivos só existem
# lá hoje, então ler de lá é a situação atual — mas tem de doer, não passar batido.
FONTES_FORA_DE_LUGAR: set = set()


def _registra_fonte(rotulo: str, f: Path) -> Path:
    _FONTES_USADAS[rotulo] = f
    try:
        if FALLBACK_DIR in Path(f).parents:
            if rotulo not in FONTES_FORA_DE_LUGAR:
                print(f"  [fonte] {rotulo}: lendo de '{FALLBACK_DIR.name}', que é pasta de "
                      f"divulgação, não de dado ({Path(f).name}). Fora dela a cópia mais "
                      f"nova é de 2025 — o arquivo precisa ser movido na origem.")
            FONTES_FORA_DE_LUGAR.add(rotulo)
    except Exception:
        pass
    return f


def caminho_curto(f) -> str:
    """Caminho legível: relativo a 'PLANILHAS DE CONTROLE' quando vem do Drive.

    A raiz é um atalho (`G:\\.shortcut-targets-by-id\\1mBrSez...`) — mostrar isso
    inteiro só polui. O que identifica a fonte é a pasta dentro do Drive, porque o
    mesmo nome de arquivo existe em mais de uma (a estação de monta muda de pasta a
    cada safra). Fora do Drive, devolve o caminho como está."""
    p = Path(f)
    try:
        return str(p.relative_to(DRIVE_ROOT)).replace("\\", "/")
    except ValueError:
        try:
            return str(p.relative_to(BASE_DIR)).replace("\\", "/")
        except ValueError:
            return str(p)


# so o orquestrador libera, via --forcar
PERMITIR_FONTE_VELHA = False
# Fontes que TEM de ser da semana: descrevem estado que muda toda semana (quem esta
# onde, quem saiu, quantos embrioes em pe). Parada, a fonte esta perdida.
# As demais — 'Animais para sair', embrioes a entregar — sao de baixa rotatividade:
# so mudam quando ha venda ou entrega nova, e ficar parado e o normal. Para essas o
# aviso sai, mas nao bloqueia: tratar "nada aconteceu" como "fonte perdida" travava o
# fechamento sem motivo.
FONTES_SEMANAIS = ("receptoras", "controle mensal", "roster do plantel",
                   "estacao de monta", "acumulado na estação")


def _avisar_fontes_velhas(ini: date, fim: date):
    """Fonte salva antes do inicio da janela nao pode descrever esta semana.

    BLOQUEIA antes de congelar o snapshot: avisar depois nao serve, porque o numero
    errado ja foi publicado. Em 21/08/2026 isso aconteceu duas vezes seguidas."""
    velhas = []
    for rotulo, f in sorted(_FONTES_USADAS.items()):
        try:
            m = datetime.fromtimestamp(f.stat().st_mtime).date()
        except OSError:
            continue
        if m < ini:
            velhas.append((rotulo, f.name, m))
    if not velhas:
        return
    bloqueiam = [v for v in velhas if v[0] in FONTES_SEMANAIS]
    print(f"  [fontes] !! {len(velhas)} fonte(s) mais VELHAS que a janela "
          f"({ini.strftime('%d/%m')}-{fim.strftime('%d/%m')}) — o que sai delas nao "
          f"descreve esta semana:")
    for rotulo, nome, m in velhas:
        marca = "  <- semanal, BLOQUEIA" if rotulo in FONTES_SEMANAIS else "  (baixa rotatividade)"
        print(f"    - {rotulo}: {nome} (salvo em {m.strftime('%d/%m/%Y')}){marca}")
    if not bloqueiam:
        print("    Nenhuma delas deveria mudar toda semana — segue.")
        return
    print("    Conferir se a copia de trabalho mudou de pasta, foi apagada ou renomeada.")
    if not PERMITIR_FONTE_VELHA:
        raise RuntimeError(
            "fonte(s) semanal(is) mais velha(s) que a janela: "
            + "; ".join(f"{r} ({n}, {m:%d/%m})" for r, n, m in bloqueiam)
            + " — snapshot NAO congelado. Use --forcar para gravar assim mesmo.")


def _latest_by_mtime(folder: Path, pattern: str) -> Path:
    """Arquivo mais recente por MTIME (data de modificação real). Inclui as cópias
    'EDITAR ...' — o operador trabalha nelas (versão viva), então são as MAIS frescas.
    Os arquivos {YYMMDD} congelados são snapshots antigos."""
    cands = [f for f in folder.glob(pattern) if not f.name.startswith("~$")]
    if not cands:
        raise FileNotFoundError(f"Nenhum arquivo {pattern} em {folder}")
    return max(cands, key=lambda f: f.stat().st_mtime)


# Arquivos ainda resolvidos na pasta de saída — o run avisa no fim.
_NA_PASTA_DE_SAIDA: list = []


def _tem_aba(f: Path, aba: str) -> bool:
    """Só os nomes das abas — não carrega célula nenhuma nem passa pelo cache."""
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
    except Exception:
        return False
    try:
        return aba in wb.sheetnames
    finally:
        wb.close()


def _resolver(pattern: str, dirs, rotulo: str, requer_aba: str | None = None) -> Path:
    """Primeiro diretório da lista que tenha o arquivo; dentro dele, o mais recente por
    mtime. A ordem é intencional (pasta canônica antes do fallback), então NÃO compare
    mtime entre diretórios: a cópia velha no lugar certo ganha da nova no lugar errado
    — é assim que a migração acontece sozinha quando alguém move o arquivo.

    `requer_aba` existe porque casar o nome não basta. O glob do Windows é
    case-insensitive e 'Animais para sair*.xlsx' casa o legado 'ANIMAIS PARA SAIR OU
    BUSCAR - ATUALIZADA 02-01-25.xlsx', de 2025 e com outro layout — a pasta canônica
    tinha um homônimo velho. Candidato sem a aba esperada é descartado, com aviso: um
    arquivo com o nome certo e a estrutura errada não pode virar fonte em silêncio."""
    tentadas, recusados = [], []
    for d in dirs:
        tentadas.append(str(d))
        if not d.exists():
            continue
        cands = sorted((f for f in d.glob(pattern) if not f.name.startswith("~$")),
                       key=lambda f: f.stat().st_mtime, reverse=True)
        for f in cands:
            if requer_aba and not _tem_aba(f, requer_aba):
                recusados.append(f"{f.name} (sem aba {requer_aba!r})")
                continue
            if recusados:
                print(f"  [fontes] {rotulo}: ignorado(s) {'; '.join(recusados)}")
            # o roster é resolvido 5x no run (headcount, doadoras, conferência...);
            # o aviso é sobre o ARQUIVO, então registra uma vez só
            if d == FALLBACK_DIR and (rotulo, f.name) not in {
                    (r, n) for r, n, _ in _NA_PASTA_DE_SAIDA}:
                _NA_PASTA_DE_SAIDA.append((rotulo, f.name, dirs[0]))
            return _registra_fonte(rotulo, f)
    detalhe = (" | recusados: " + "; ".join(recusados)) if recusados else ""
    raise FileNotFoundError(
        f"Nenhum {pattern!r} ({rotulo}) em: " + " | ".join(tentadas) + detalhe)


def _avisar_pasta_de_saida():
    """Fecha o run listando o que ainda sai da pasta de publicação e pra onde deveria
    ir. Sem isso a dependência some de vista e ninguém move o arquivo."""
    if not _NA_PASTA_DE_SAIDA:
        return
    print(f"  [fontes] {len(_NA_PASTA_DE_SAIDA)} arquivo(s) ainda lidos de "
          f"{FALLBACK_DIR.name} (pasta de publicação, não de dado):")
    for rotulo, nome, destino in _NA_PASTA_DE_SAIDA:
        print(f"    - {rotulo}: {nome}  ->  mover para {destino.name}")


def _controle_plantel() -> Path:
    return _resolver(CONTROLE_PLANTEL_GLOB, CONTROLE_PLANTEL_DIRS, "roster do plantel",
                     requer_aba="PLANTEL")


def _estacao_dirs() -> list:
    """Pastas de estacao do PLANTEL, da mais nova para a mais velha."""
    if not PLANTEL_DIR_BASE.exists():
        return []
    return sorted((d for d in PLANTEL_DIR_BASE.glob(PLANTEL_ESTACAO_GLOB) if d.is_dir()),
                  reverse=True)


def _latest_no_plantel(pattern: str, rotulo: str) -> Path:
    """Arquivo mais recente que casa o padrao em QUALQUER pasta de estacao."""
    cands = [f for d in _estacao_dirs() for f in d.glob(pattern)
             if not f.name.startswith("~$")]
    if not cands:
        raise FileNotFoundError(
            f"Nenhum {pattern!r} ({rotulo}) em: "
            + " | ".join(str(d) for d in _estacao_dirs()))
    return _registra_fonte(rotulo, max(cands, key=lambda f: f.stat().st_mtime))


def _latest_estacao_master() -> Path:
    """Master da estação de monta, na pasta 'Estação <ano>-<ano>' certa.

    Era fixo em 'Estação 2025-2026'. Na virada de safra o haras passa a atualizar o
    arquivo em 'Estação 2026-2027', e esse caminho fixo nunca ia ver — igual ao bug
    já corrigido pro roster/receptoras (ver comentário de PLANTEL_ESTACAO_GLOB).
    Mesma solução: varre TODAS as pastas de safra e fica com a mais nova por mtime."""
    if not ESTACAO_MONTA_BASE.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {ESTACAO_MONTA_BASE}")
    dirs = sorted((d for d in ESTACAO_MONTA_BASE.glob("Estação *") if d.is_dir()),
                  reverse=True)
    cands = [f for d in dirs for f in d.glob("*ESTACAO DE MONTA.xlsx")
             if not f.name.startswith("~$")]
    if not cands:
        raise FileNotFoundError(
            "Nenhum '*ESTACAO DE MONTA.xlsx' (estacao de monta) em: "
            + " | ".join(str(d) for d in dirs))
    return _registra_fonte("estacao de monta", max(cands, key=lambda f: f.stat().st_mtime))


def _latest_by_yymmdd(folder: Path, pattern: str, rotulo: str | None = None) -> Path:
    """Mais recente por mtime, registrado sob `rotulo`.

    O rótulo era adivinhado pelo padrão — RECEPTORAS virava 'receptoras', o resto
    virava 'controle mensal'. O comitê chama esta mesma função para o mapa de vendas,
    e o registro passava a dizer que o controle mensal era o mapa de vendas. Rótulo
    errado não muda número, mas vai direto para a auditoria, que serve exatamente
    para dizer de onde veio cada dado."""
    if rotulo is None:
        rotulo = "receptoras" if "RECEPTORAS" in pattern.upper() else "controle mensal"
    return _registra_fonte(rotulo, _latest_by_mtime(folder, pattern))


# ------------------------------------------------------------------
# Estrutura do relatório
# ------------------------------------------------------------------
@dataclass
class Report:
    semana_inicio: str
    semana_fim: str
    fontes: dict = field(default_factory=dict)
    fontes_caminhos: dict = field(default_factory=dict)  # rótulo -> caminho no Drive
    fontes_fora_de_lugar: list = field(default_factory=list)  # lidas da pasta de divulgação
    producao: dict = field(default_factory=dict)
    receptoras: dict = field(default_factory=dict)
    headcount: dict = field(default_factory=dict)
    terceiros: dict = field(default_factory=dict)
    saidas: dict = field(default_factory=dict)
    detalhe: dict = field(default_factory=dict)  # tabelas de detalhe p/ o dashboard
    eventos: dict = field(default_factory=dict)  # listas COMPLETAS datadas (filtro client-side)
    calendario: list = field(default_factory=list)  # semanas travadas p/ o seletor
    semana_atual: str = ""  # id (segunda) da semana selecionada por padrão
    snapshots: dict = field(default_factory=dict)  # snapshot por semana (histórico local)
    roster: list = field(default_factory=list)  # nomes do plantel (p/ diff saídas/entradas)
    receptoras_locais: dict = field(default_factory=dict)  # {animal: LOCAL} p/ diff de transferência
    populacao: list = field(default_factory=list)  # plantel + receptoras contadas (p/ diff saídas/entradas)
    saidas_planilha: dict | None = None  # aba SAIDAS-ENTRADAS, quando o haras preencher
    confirmed: list = field(default_factory=list)  # embriões confirmados (+/-=OK) p/ diff semanal
    docx_ref: dict = field(default_factory=dict)  # números dos relatórios oficiais (SÓ p/ validar)


# ------------------------------------------------------------------
# Seção 1 — PRODUÇÃO (master da estação de monta)
# ------------------------------------------------------------------
# Abas da planilha do grupo que compõem o acumulado.
#   (aba, fatia fixa ou None = deduzir do STATUS)
# As COLUNAS são resolvidas pelo CABEÇALHO (linha 3), nunca por índice fixo. Em
# 07/08/2026 a coluna STATUS foi apagada da aba de PAO GRANDE, tudo à direita
# andou uma casa e o índice de ESTAÇÃO passou a cair em COTA PG: nenhuma linha
# batia a safra, a fatia "pg" zerou e o acumulado caiu de 61 pra 27 sem erro
# nenhum. Cabeçalho por nome faz a planilha poder mexer coluna sem quebrar.
ABAS_ACUMULADO = (
    ("EMBRIÕES PAO GRANDE", "pg"),
    ("EMBRIOES SOCIOS - VENDIDOS", None),
)
HDR_ROW_ACUMULADO = 3  # linha do cabeçalho nas duas abas


def _sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn")


def _col_idx(hdr, *nomes) -> int:
    """Índice da coluna pelo nome do cabeçalho (sem acento, case-insensitive).
    Explode se não achar — coluna que sumiu tem de virar erro, não zero calado."""
    alvo = {_sem_acento(n).strip().upper() for n in nomes}
    for i, h in enumerate(hdr):
        if h is not None and _sem_acento(h).strip().upper() in alvo:
            return i
    achados = [str(h) for h in hdr if h is not None]
    raise KeyError(f"coluna {nomes[0]!r} não está no cabeçalho: {achados}")


def _latest_emb_matrizes() -> Path:
    return _resolver(EMB_MATRIZES_GLOB, EMB_MATRIZES_DIRS, "acumulado na estação",
                     requer_aba="EMBRIÕES PAO GRANDE")


def _acumulado_grupo(safra: str = SAFRA_ATUAL) -> dict:
    """Embriões vivos da safra na planilha 'EMBRIÕES E MATRIZES' — a que o haras
    manda no grupo e usa como fonte oficial do acumulado.

    Ela é um retrato do que está EM PÉ: a linha é apagada quando o embrião sai da
    conta, seja por parição (vira potro) ou por aborto. Por isso o acumulado da
    estação = linhas vivas + parições da safra. Aborto NÃO volta — é baixa
    definitiva, confirmado com o haras em 2026-07-31 (os 3 abortos da safra também
    sumiram da planilha e não entram no número oficial).

    A aba ICSI fica de fora: hoje só tem safra 2023/2024."""
    src = _latest_emb_matrizes()
    wb = _load(src)
    out = {"pg": 0, "socio": 0, "vendido": 0, "fonte": src.name,
           "chaves": set(), "linhas": []}
    for aba, fatia_fixa in ABAS_ACUMULADO:
        if aba not in wb.sheetnames:
            print(f"  [acumulado] aba {aba!r} não existe em {src.name} — fatia zerada")
            continue
        iest = ist = idoa = igar = irec = icot = isoc = None
        for i, r in enumerate(wb[aba].iter_rows(values_only=True), start=1):
            if i == HDR_ROW_ACUMULADO:
                iest = _col_idx(r, "ESTACAO")
                ist = None if fatia_fixa else _col_idx(r, "STATUS")
                # as duas abas têm as mesmas colunas em posições diferentes
                # (RECEPTORA é a 5ª numa e a 4ª na outra) — resolver por nome
                idoa, igar, irec = (_col_idx(r, "DOADORA"), _col_idx(r, "GARANHÃO"),
                                    _col_idx(r, "RECEPTORA"))
                # cota e sócio não existem nas duas abas; ausentes viram None
                def _opt(*nomes):
                    try:
                        return _col_idx(r, *nomes)
                    except KeyError:
                        return None
                icot, isoc = _opt("COTA PG"), _opt("COMPRADOR/SOCIO")
                continue
            if i <= HDR_ROW_ACUMULADO or r[1] is None or not str(r[1]).strip():
                continue
            if _s(r[iest]) != safra:
                continue
            fatia = fatia_fixa or ("vendido" if _norm(r[ist]) == "VENDIDO" else "socio")
            out[fatia] += 1
            chave = _chave_embriao(r[idoa], r[igar], r[irec])
            out["chaves"].add(chave)
            # a LINHA, não só a chave: quando o embrião pare, o haras apaga a linha e a
            # cota vai com ela. Guardada aqui, a fatia da parição é recuperável.
            out["linhas"].append({
                "chave": list(chave), "aba": aba, "fatia": fatia,
                "doadora": _s(r[idoa]), "garanhao": _s(r[igar]),
                "receptora": _s(r[irec]), "cota": (r[icot] if icot is not None else None),
                "socio": (_s(r[isoc]) if isoc is not None else None),
            })
    wb.close()
    out["total"] = out["pg"] + out["socio"] + out["vendido"]
    return out


# Animal bloqueado por pendência documental continua "pendente de saída" no sentido
# literal, mas o relatório NÃO o conta — 14/08 e 21/08 dizem "01 animal" onde temos
# MUSICA e NOBRE, e NOBRE é o único com essa observação nas duas semanas.
OBS_BLOQUEIA_SAIDA = ("FALTANDO EXAME",)


def _sem_cotista(n) -> str:
    """'PARIS DA PAO GRANDE (EDUARDO)' -> 'PARIS DA PAO GRANDE'.

    O roster mensal repete a mesma linha por cotista, com o nome do cotista no fim
    do nome. Sem tirar isso, um animal conta duas ou três vezes."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", _norm(n)).strip()


def _chave_animal(nome, mae, pai) -> tuple:
    """Identidade do animal no roster mensal: nome sem cotista + filiação.

    Nome sozinho junta dois potros distintos chamados `MACHO`. Com a data de
    nascimento na chave, `XARDA DO SALTO` viraria dois animais por causa de um dia
    de diferença digitado errado (20/09 e 21/09 de 2021)."""
    return (_sem_cotista(nome), _norm(mae), _norm(pai))


def _nucleo_nome(n) -> str:
    """Nome comparável entre planilhas. O roster chama o mesmo animal de
    'POTRA MORENA L2 X DAMASCO DA PAO GRANDE - 07/03/2025 RECEP 07 V' e o Animais para
    sair de 'FEMEA MORENA L2 X DAMASCO DA PAO GRANDE': muda o prefixo de sexo e sobra
    a data/receptora no fim. Sem normalizar, a reposição não é reconhecida e entra na
    conta de vendidos."""
    t = _norm(n)
    t = re.sub(r"^(POTRA|POTRO|FEMEA|FEMA|MACHO)\s+", "", t)
    t = re.sub(r"\s*-\s*\d.*$", "", t)
    t = re.sub(r"\s+RECEP.*$", "", t)
    return t.strip()


def _chave_embriao(doadora, garanhao, receptora) -> tuple:
    """Identidade do embrião entre a planilha do grupo e a aba ESTAÇÃO. Receptora
    normalizada porque vem '309' numa e 309.0 na outra."""
    def _rec(v):
        try:
            return str(int(float(v)))
        except (TypeError, ValueError):
            return _norm(v)
    return (_norm(doadora), _norm(garanhao), _rec(receptora))


def _mortes_do_plantel(ini: date, fim: date) -> list:
    """Óbitos registrados na aba 'CONFIRMAÇÕES, ABORTOS, MORTES' do controle mensal.

    Reunião 2026-07-31: a estação de monta só registra aborto/absorção, morte de
    receptora prenha e potro que nasce e morre em seguida. Óbito de animal já no
    plantel só aparece aqui. Colunas: B produto, C data, D observação (texto livre,
    por isso a classificação é por palavra-chave).
    """
    try:
        src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    except FileNotFoundError:
        return []
    wb = _load(src)
    aba = "CONFIRMAÇÕES, ABORTOS, MORTES"
    if aba not in wb.sheetnames:
        wb.close()
        return []
    achados = []
    for i, r in enumerate(wb[aba].iter_rows(values_only=True), start=1):
        if i < 3 or r[1] is None:
            continue
        d = _dt(r[2])
        if not d or not (ini <= d <= fim):
            continue
        obs = _norm(r[3])
        if "ABORT" in obs:            # aborto já vem da estação; aqui só morte
            continue
        if any(p in obs for p in ("MORREU", "MORTE", "OBITO", "MORTO")):
            achados.append({"animal": _s(r[1]), "data": d.isoformat(),
                            "ocorrencia": _s(r[3]), "origem": "plantel"})
    wb.close()
    return achados


def _acumulado_planejamento(wb) -> int:
    """Total da estação de monta = soma de 'TOTAL EMBRIÕES' (REAL, idx8) da aba
    PLANEJAMENTO, linhas de doadora (col0 numérica). NÃO é mais o acumulado
    publicado — é referência de conferência: a estação só registra o que passou
    pela FPG, então fica abaixo do oficial."""
    if "PLANEJAMENTO" not in wb.sheetnames:
        return 0
    ws = wb["PLANEJAMENTO"]
    total = 0.0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 4 or row[0] is None:
            continue
        if not str(row[0]).strip().isdigit():
            continue
        v = _to_num(row[8])
        if v:
            total += v
    return int(round(total))


def build_producao(rep: Report, ini: date, fim: date):
    master = _latest_estacao_master()
    rep.fontes["estacao_master"] = master.name
    wb = _load(master)
    ws = wb["ESTAÇÃO"]
    embrioes = []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 3 or r[0] is None:
            continue
        safra_linha = _s(r[35])
        if safra_linha not in (SAFRA_ATUAL, SAFRA_PROXIMA):
            continue
        ia = _dt(r[7])
        # confirmação oficial = coluna +/- (idx17) == 'OK' (validado: os embriões
        # confirmados no docx têm +/-='OK', 60D às vezes nem marcado). Conta = PLANEJAMENTO.
        confirmado = _norm(r[17]) == "OK"
        data_conf = (ia + timedelta(days=60)) if ia else None
        cotas = _to_num(r[5])
        # split PG / sócio / vendido (defs do comitê: total produzido aberto nessas 3 fatias)
        comprador = _s(r[34])
        if cotas is not None and cotas == 0:
            fatia = "vendido"
        elif (cotas is not None and 0 < cotas < 1) or _s(r[6]):
            fatia = "socio"
        else:
            fatia = "pg"
        embrioes.append({
            "doadora": _s(r[2]), "garanhao": _s(r[3]), "local": _s(r[4]),
            "cotas_pg": cotas, "socio": _s(r[6]), "fatia": fatia,
            "data_ia": ia.isoformat() if ia else None,
            "receptora": _s(r[11]),
            "confirmado": confirmado,
            "data_confirmacao": data_conf.isoformat() if data_conf else None,
            "sexo_potro": _s(r[24]), "nome_potro": _s(r[25]),
            "data_paricao": _dt(r[23]).isoformat() if _dt(r[23]) else None,
            "data_aborto": _dt(r[21]).isoformat() if _dt(r[21]) else None,
            "data_obito": _dt(r[28]).isoformat() if _dt(r[28]) else None,
            "status": _s(r[32]), "categoria": _s(r[33]), "comprador": comprador,
            "safra": safra_linha,
        })

    # o resto de build_producao e da safra corrente; a proxima entra so no acumulado
    embrioes_proxima = [e for e in embrioes if e["safra"] == SAFRA_PROXIMA]
    embrioes = [e for e in embrioes if e["safra"] == SAFRA_ATUAL]

    def _in_week(iso):
        return bool(iso and ini <= date.fromisoformat(iso) <= fim)

    def _in_month(iso):
        if not iso:
            return False
        d = date.fromisoformat(iso)
        return d.year == fim.year and d.month == fim.month

    def _split(items):
        return {f: sum(1 for e in items if e["fatia"] == f) for f in ("pg", "socio", "vendido")}

    # ACUMULADO NA ESTAÇÃO (2026-07-31, reunião com o haras):
    #   linhas vivas da planilha "EMBRIÕES E MATRIZES" (PG + sócios) + parições da safra.
    # A planilha do grupo é o retrato do que está em pé — some a linha quando pare ou
    # aborta —, então as parições voltam pra conta e os abortos não. A estação de monta
    # sozinha não serve: só enxerga o que passou pela FPG (56 vs 61 hoje). Fica como
    # conferência em `acumulado_estacao_monta`.
    # ACUMULADO: confirmados da safra na própria aba ESTAÇÃO. Ela não apaga linha —
    # parição e aborto ficam registrados na mesma linha —, então a contagem já é
    # cumulativa. O grupo (EMBRIÕES E MATRIZES) segue sendo lido só para o split por
    # fatia e para a lista arquivada.
    grupo = _acumulado_grupo()
    rep.fontes["embrioes_matrizes"] = grupo["fonte"]
    _LINHAS_BRUTAS["grupo"] = grupo["linhas"]
    acumulado_planejamento = _acumulado_planejamento(wb)

    confirmados = [e for e in embrioes if e["confirmado"]]
    for e in confirmados:
        e["key"] = f"{e['doadora']}|{e['garanhao']}|{e['receptora']}|{e['data_ia']}"
    rep.confirmed = confirmados
    na_semana = []          # confirmados_semana e acumulado_mes: preenchidos por _compute_confirmados_diff
    nascimentos = [e for e in embrioes if _in_week(e["data_paricao"])]
    # aborto = embrião confirmado que não nasceu (data aborto na semana);
    # óbito = nasceu e morreu (data óbito na semana). (absorção = perda pré-60d, não confirmada)
    abortos = [e for e in embrioes if e["confirmado"] and _in_week(e["data_aborto"])]
    obitos = [e for e in embrioes if _in_week(e["data_obito"])]
    # óbito de animal já no plantel não passa pela estação — vem do controle mensal
    mortes_plantel = _mortes_do_plantel(ini, fim)
    if mortes_plantel:
        print(f"  [produção] +{len(mortes_plantel)} óbito(s) da aba "
              f"'CONFIRMAÇÕES, ABORTOS, MORTES' do controle mensal")

    # Parições da safra inteira (não só da semana) — voltam pro acumulado, PORQUE a
    # planilha do grupo apaga a linha de quem pariu. Quando ela ainda não foi apagada,
    # somar as duas conta o mesmo embrião duas vezes: em 17/08/2026 a parição de JAVA
    # DA PAO GRANDE x QUEBRUTO (recep 309) foi lançada na estação às 17:05 e a planilha
    # do grupo estava salva desde 14/08 com a linha viva — o acumulado pulou de 61 pra
    # 62 numa semana sem nenhuma produção nova. Descontar é o certo: enquanto o embrião
    # está nas duas listas, ele é UM.
    paridos_safra = [e for e in embrioes if e["data_paricao"]]
    ainda_no_grupo = [e for e in paridos_safra
                      if _chave_embriao(e["doadora"], e["garanhao"], e["receptora"])
                      in grupo["chaves"]]
    paridos_novos = [e for e in paridos_safra if e not in ainda_no_grupo]
    if ainda_no_grupo:
        print(f"  [acumulado] {len(ainda_no_grupo)} parição(ões) ainda listadas como "
              f"vivas em {grupo['fonte']} — contadas UMA vez, não duas:")
        for e in ainda_no_grupo:
            print(f"    - {e['doadora']} x {e['garanhao']} (recep {e['receptora']}), "
                  f"pariu {e['data_paricao']}")
    # Confirmados da safra na aba ESTAÇÃO = acumulado. Aborto não volta (baixa
    # definitiva, decidido com o haras em 2026-07-31), então sai da conta.
    confirmados_safra = [e for e in embrioes if e["confirmado"]]
    abortados = [e for e in confirmados_safra if e["data_aborto"]]
    acumulado = len(confirmados_safra) - len(abortados)
    split = {"pg": grupo["pg"], "socio": grupo["socio"], "vendido": grupo["vendido"]}
    for e in paridos_novos:
        split[e["fatia"]] = split.get(e["fatia"], 0) + 1
    if acumulado != acumulado_planejamento:
        print(f"  [produção] acumulado {acumulado} "
              f"(planilha do grupo {grupo['total']} + {len(paridos_novos)} parições) "
              f"vs {acumulado_planejamento} na estação de monta")

    # Mesma regra da safra corrente, aplicada na que comeca: vivos na planilha do
    # grupo + paricoes que ja sairam de la. Enquanto nao ha lancamento, da 0 — e 0
    # aqui e o '--' do relatorio, nao um numero perdido.
    grupo_prox = _acumulado_grupo(SAFRA_PROXIMA)
    paridos_prox = [e for e in embrioes_proxima if e["data_paricao"]
                    and _chave_embriao(e["doadora"], e["garanhao"], e["receptora"])
                    not in grupo_prox["chaves"]]
    acumulado_prox = grupo_prox["total"] + len(paridos_prox)
    if acumulado_prox:
        print(f"  [acumulado] safra {SAFRA_PROXIMA} ja tem {acumulado_prox} "
              f"(grupo {grupo_prox['total']} + {len(paridos_prox)} parições)")

    rep.producao = {
        "acumulado_estacao": acumulado,
        "acumulado_estacao_proxima": acumulado_prox,
        "safra_atual": SAFRA_ATUAL,
        "safra_proxima": SAFRA_PROXIMA,
        "safra_atual_rotulo": _rotulo_safra(SAFRA_ATUAL),
        "safra_proxima_rotulo": _rotulo_safra(SAFRA_PROXIMA),
        "acumulado_estacao_split": split,
        "acumulado_estacao_monta": acumulado_planejamento,   # conferência
        "acumulado_grupo_vivos": grupo["total"],
        "acumulado_paricoes_safra": len(paridos_novos),
        "acumulado_paricoes_ainda_no_grupo": len(ainda_no_grupo),
        "acumulado_estacao_split_confirmados": _split(confirmados),  # split antigo (estação)
        "confirmados_semana": None,   # _compute_confirmados_diff
        "acumulado_mes": None,        # _compute_confirmados_diff
        "nascimentos": None,          # _nascimentos_do_roster, abaixo
        "abortos_obitos": len(abortos) + len(obitos) + len(mortes_plantel),
    }
    def _publica_nascimentos():
        rep.producao["nascimentos"] = len(rep.detalhe.get("nascimentos_semana") or [])

    def _produto(e):   # nome do animal nascido (ou descrição sexo — doadora × garanhão)
        base = f"{e.get('doadora') or ''} × {e.get('garanhao') or ''}".strip(" ×")
        sx = {"M": "Macho", "F": "Fêmea"}.get((e.get("sexo_potro") or "").upper(), e.get("sexo_potro") or "")
        return e.get("nome_potro") or (f"{sx} — {base}" if base else sx) or "--"
    rep.detalhe["confirmados_semana"] = na_semana
    _SOCIO_POR_RECEP.clear()
    for e in embrioes:
        soc = _limpa_socio(e.get("comprador")) or _limpa_socio(e.get("socio"))
        if soc and e.get("receptora"):
            _SOCIO_POR_RECEP[_norm(e["receptora"])] = soc
    # Nascimento vem do roster mensal (data + filiação). A parição lançada na aba
    # ESTAÇÃO fica como conferência: divergir significa lançamento faltando em um dos
    # dois lados, e isso tem de aparecer em vez de escolher um número calado.
    nasc_roster = _nascimentos_do_roster(ini, fim)
    rep.producao["nascimentos_estacao"] = len(nascimentos)
    if len(nasc_roster) != len(nascimentos):
        print(f"  [nascimentos] roster mensal {len(nasc_roster)} x aba ESTAÇÃO "
              f"{len(nascimentos)} — publicando o roster (é onde o potro entra); "
              f"a diferença é parição não lançada em um dos dois")
    rep.detalhe["nascimentos_semana"] = nasc_roster
    rep.detalhe["nascimentos_estacao"] = [
        dict(e, produto=_produto(e),
             socio=_limpa_socio(e.get("comprador")) or _limpa_socio(e.get("socio")))
        for e in nascimentos]
    _publica_nascimentos()
    rep.detalhe["abortos_obitos_semana"] = abortos + obitos + mortes_plantel
    rep.detalhe["embrioes_confirmados"] = confirmados
    # listas COMPLETAS datadas (o dashboard filtra por semana no cliente)
    rep.eventos["confirmados"] = [dict(e, data=e["data_confirmacao"]) for e in confirmados if e["data_confirmacao"]]
    rep.eventos["nascimentos"] = [dict(e, data=e["data_paricao"]) for e in embrioes if e["data_paricao"]]
    rep.eventos["abortos_obitos"] = [
        dict(e, data=(e["data_aborto"] or e["data_obito"]))
        for e in embrioes if (e["confirmado"] and e["data_aborto"]) or e["data_obito"]
    ]
    wb.close()


# ------------------------------------------------------------------
# Seção 2 — RECEPTORAS (censo + candidatos de filtro)
# ------------------------------------------------------------------
RECEPTORAS_LOCAIS_ATIVOS = ("PAO GRANDE", "ARRENDAMENTO CESAR FURTADO")


def build_receptoras(rep: Report):
    """Rebanho ATIVO = aba ANIMAIS do PLANTEL ARRENDAMENTOS E RECEPTORAS, filtrando
    LOCAL em (PAO GRANDE, ARRENDAMENTO CESAR FURTADO) e STATUS prenha/vazia.
    NÃO usa mais a aba 'ATUALIZAÇÃO SEMANAL' (aba a ser aposentada). Validado vs
    docx 24/07: prenhas 34 / vazias 28 / total 62 — bate EXATO.
    Colunas ANIMAIS (linha 3 header, dados r4+): 1 ANIMAL, 2 STATUS, 3 LOCAL."""
    src = _latest_no_plantel("*PLANTEL ARRENDAMENTOS E RECEPTORAS.xlsx", "receptoras")
    rep.fontes["receptoras"] = src.name
    wb = _load(src)
    ws = wb["ANIMAIS"]
    pren = vaz = 0
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 4 or r[1] is None:
            continue
        if _norm(r[3]) not in RECEPTORAS_LOCAIS_ATIVOS:
            continue
        st = _norm(r[2])
        if st.startswith("PRENHA"):
            pren += 1
        elif st.startswith("VAZIA"):
            vaz += 1
    wb.close()
    doadoras_plantel = _count_doadoras()             # CATEGORIA='DOADORA' no plantel
    doadoras_fpg = _count_doadoras("FAZENDA PAO GRANDE")   # só referência
    doadoras = DOADORAS_INDICE or doadoras_plantel
    rep.receptoras = {
        "total": pren + vaz,
        "prenhas": pren,
        "vazias": vaz,
        # Índice = vazias ÷ doadoras (CATEGORIA='DOADORA' no roster). Confirmado
        # contra os dois pontos do histórico (21/08: doadoras=12, ciclando=10,
        # índice=2,5 — só 30÷12 fecha; 30÷10 dá 3,0) depois de eu ter "corrigido"
        # isto errado em 28/08/2026 pra usar ciclando, sem checar contra semana
        # nenhuma. Revertido. 'doadoras_ciclando' é indicador PRÓPRIO no relatório
        # (card ao lado), não entra nesta conta.
        "doadoras": doadoras,
        "doadoras_fonte": "fixo" if DOADORAS_INDICE else "plantel",
        "doadoras_plantel": doadoras_plantel,
        "doadoras_plantel_fpg": doadoras_fpg,
        "indice_eficiencia": round(vaz / doadoras, 1) if doadoras else None,
        # preenchido em bases/semanal_manual.json (ver _manual): sem fonte de dado
        "doadoras_ciclando": None,
    }
    if DOADORAS_INDICE and DOADORAS_INDICE != doadoras_plantel:
        print(f"  [receptoras] índice usa {DOADORAS_INDICE} doadoras (fixo); "
              f"no plantel há {doadoras_fpg} na FPG e {doadoras_plantel} no total")


# Denominador do índice de eficiência.
# Era FIXO em 10 pela reunião de 2026-07-31, porque nenhuma leitura do plantel dava
# esse número. Em 07/08/2026 o relatório oficial trocou o divisor: 29 vazias com
# índice 2,4 = 29/12, e 12 é exatamente CATEGORIA='DOADORA' no plantel (nas 4 semanas
# anteriores o divisor implícito no docx era 10). Ou seja, o haras passou a usar o
# contado — então paramos de fixar e contamos.
# None = usar as doadoras contadas no plantel. Só voltar a pôr número aqui se o
# haras decidir travar o divisor de novo.
DOADORAS_INDICE = None


def _count_doadoras(local: str | None = None) -> int:
    """Doadoras no plantel = CATEGORIA 'DOADORA', STATUS ativo, no roster mensal.

    NÃO usa o roster já deduplicado do headcount — achado em 28/08/2026: a fonte
    tem 'XARDA DO SALTO (CARLA)' e 'XARDA DO SALTO (EDUARDO)' como DUAS linhas
    (duas cotistas da mesma égua), e o relatório oficial conta as duas (12
    doadoras). O dedup por cotista do headcount existe porque lá é 1 animal físico
    só; aqui o relatório conta LINHA, não animal — usar o roster deduplicado
    derrubava 12 pra 11. Por isso relê a planilha direto, sem passar por
    `_plantel_por_status()`."""
    src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    wb = _load(src)
    ws = wb["PLANTEL"]
    L = PLANTEL_LAYOUT_MENSAL
    n = 0
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < L["linha1"] or r[L["nome"]] is None:
            continue
        if _norm(r[L["categoria"]]) != "DOADORA":
            continue
        if _norm(r[L["status"]]) not in STATUS_NO_PLANTEL:
            continue
        # mesma exclusão do roster principal: saiu do haras e a PG não tem cota
        # nenhuma = acabou, mesmo com STATUS ainda dizendo 'VENDIDO'. Pegou a
        # CANCAO DA ILHA (SAIU DO HARAS, cota 0) que ficaria contada sem isto.
        cota = r[COL_MENSAL_COTAS] if len(r) > COL_MENSAL_COTAS else None
        condicao = _norm(r[COL_MENSAL_CONDICAO]) if len(r) > COL_MENSAL_CONDICAO else ""
        if condicao == CONDICAO_SAIU and (cota is None or not cota):
            continue
        if local is not None and _norm(r[L["local"]]) != local:
            continue
        n += 1
    wb.close()
    return n


# ------------------------------------------------------------------
# Seção 3 — HEADCOUNT (calculado do roster, espelhando a aba CONTAGEM)
# ------------------------------------------------------------------
# A CONTAGEM é COUNTIF puro sobre o LOCAL do roster:
#     C3 =COUNTIF(PLANTEL!E:E;"FAZENDA PAO GRANDE")      D3 = 25   (digitado)
#     C4 =COUNTIF(PLANTEL!E:E;"ARRENDAMENTO CESAR FURTADO")  D4 = 37   (digitado)
#     C5 =COUNTIF(PLANTEL!E:E;"OUTROS")                  D5 = 0     (rótulo "CTE")
#     C6 =COUNTIF(PLANTEL!E:E;"SOCIO")                   D6 = 0
#     E7 =SUM(E3:E6)
# Reproduzimos a MESMA regra (conta linha por LOCAL, sem filtrar STATUS — por isso
# os vendidos pendentes de saída entram) e o MESMO conjunto de buckets. Só muda de
# onde vem o número das receptoras: contado da fonte de receptoras (a mesma da
# seção 2) em vez de digitado na mão. Confere — PAO GRANDE 25 e ARRENDAMENTO CESAR
# FURTADO 37, idênticos ao que estava fixo em D3/D4.
#
# MATO GROSSO fica FORA do headcount por decisão de negócio (confirmado em
# 2026-07-31): os animais de lá não entram na contagem do plantel, e é por isso
# que a CONTAGEM nunca teve linha para eles.
#
# LOCAL vazio não entra (o COUNTIF também ignora), o que descarta de graça a
# pseudo-linha "RECEPTORAS 67" que mora no meio do roster.

# LOCAL no roster -> chave no relatório (rótulo que a CONTAGEM usa)
HEADCOUNT_BUCKETS = {
    "FAZENDA PAO GRANDE": ("FAZENDA", "fazenda_pg"),
    "ARRENDAMENTO CESAR FURTADO": ("ARRENDAMENTO", "arrendamento"),
    "OUTROS": ("CTE", "cte"),
    "SOCIO": ("SOCIO", "socio"),
}
# LOCAL que existe no roster mas não conta no headcount.
HEADCOUNT_LOCAIS_FORA = ("MATO GROSSO",)
# LOCAL da fonte de receptoras -> LOCAL do roster
RECEPTORAS_PARA_BUCKET = {
    "PAO GRANDE": "FAZENDA PAO GRANDE",
    "ARRENDAMENTO CESAR FURTADO": "ARRENDAMENTO CESAR FURTADO",
}


def _slug_local(local: str) -> str:
    return local.lower().replace(" ", "_")


def _receptoras_por_local() -> dict:
    """{LOCAL do roster: nº de receptoras}. Mesma regra da seção 2 (prenha/vazia)."""
    src = _latest_no_plantel("*PLANTEL ARRENDAMENTOS E RECEPTORAS.xlsx", "receptoras")
    wb = _load(src)
    ws = wb["ANIMAIS"]
    out = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 4 or r[1] is None:
            continue
        loc = _norm(r[3])
        if loc not in RECEPTORAS_LOCAIS_ATIVOS:
            continue
        st = _norm(r[2])
        if st.startswith("PRENHA") or st.startswith("VAZIA"):
            out[RECEPTORAS_PARA_BUCKET[loc]] = out.get(RECEPTORAS_PARA_BUCKET[loc], 0) + 1
    wb.close()
    return out


def build_headcount(rep: Report):
    # 1) animais por LOCAL, do MESMO roster que o resto do fechamento usa.
    #    Relia a planilha aqui dentro e contava linha a linha sem filtro nenhum —
    #    funcionava porque a planilha semanal já vinha curada. Com o roster mensal
    #    isso contaria vendido, morto, embrião e a linha repetida por cotista.
    animais: dict[str, int] = {}
    for linha in _plantel_por_status()["linhas"]:
        local = _norm(linha["local"])
        if not local:
            continue
        animais[local] = animais.get(local, 0) + 1

    # 2) receptoras: contadas da fonte de receptoras
    receptoras = _receptoras_por_local()

    # 3) monta os buckets. LOCAL fora da contagem é ignorado; LOCAL desconhecido
    #    também não entra (a CONTAGEM não o teria), mas vira aviso — assim um
    #    local novo no roster aparece pra alguém decidir, em vez de sumir calado.
    detalhe, chaves, fora = {}, {}, {}
    for local in sorted(set(animais) | set(receptoras)):
        a = animais.get(local, 0)
        rc = receptoras.get(local, 0)
        if local in HEADCOUNT_LOCAIS_FORA or local not in HEADCOUNT_BUCKETS:
            fora[local] = a + rc
            continue
        rotulo, chave = HEADCOUNT_BUCKETS[local]
        detalhe[rotulo] = {"animais": a, "receptoras": rc, "total": a + rc}
        chaves[chave] = a + rc

    total = sum(v["total"] for v in detalhe.values())
    detalhe["TOTAL GERAL"] = {
        "animais": sum(v["animais"] for k, v in detalhe.items() if k != "TOTAL GERAL"),
        "receptoras": sum(v["receptoras"] for k, v in detalhe.items() if k != "TOTAL GERAL"),
        "total": total,
    }

    rep.headcount = {"total": total, **chaves, "detalhe": detalhe, "fora_da_contagem": fora}

    # A conferência contra a aba CONTAGEM saiu: aquela aba é um COUNTIF dentro do
    # arquivo de DIVULGAÇÃO, e conferir o cálculo contra o que foi divulgado não
    # confere nada — é o próprio número que se quer auditar. Quem confere o headcount
    # é o relatório oficial, no placar do fechamento.
    desconhecidos = [l for l in fora if l not in HEADCOUNT_LOCAIS_FORA]
    if desconhecidos:
        print("  [headcount] LOCAL novo no roster, FORA da contagem — conferir:")
        for l in desconhecidos:
            print(f"    - {l}: {fora[l]}")


# ------------------------------------------------------------------
# Seção 5 — SAÍDAS / ENTRADAS / TRANSFERÊNCIAS  (CONTROLE_DE_PLANTEL aba MOVIMENTAÇÕES)
# ------------------------------------------------------------------
def _categorize_mov(obs: str):
    if "MUDOU O LOCAL PARA FAZENDA PAO GRANDE" in obs or "MUDOU O LOCAL PARA ARRENDAMENTO" in obs:
        return "TRANSFERENCIA"
    if "SAIU DO HARAS" in obs:
        return "SAIDA"
    if "CHEGOU NO HARAS" in obs:
        return "ENTRADA"
    return None


# Aba SAIDAS-ENTRADAS do controle mensal (reunião 2026-07-31): fonte oficial de
# saída/entrada. O haras começou a preencher em 12/08/2026.
# Colunas: B animal, C local saída, D local entrada, E data, F classificação.
# Δ combinado: entrada = nascimento + compra; saída = venda + morte.
#
# A classificação vem digitada com o sentido colado no motivo — a primeira linha real
# foi 'SAIDA-SOCIO'. Só as palavras de motivo abaixo não bastavam: 'SAIDA-SOCIO' não
# casava com nada, caía no `else None` e era descartada em silêncio. Como a aba já
# tinha linha, ela vencia como fonte oficial e a saída da semana virava 0 — foi assim
# que LINDEZA DA PAO GRANDE (12/08/2026) sumiu do fechamento de 14/08.
CLASSIF_ENTRADA = ("NASCIMENTO", "COMPRA")
CLASSIF_SAIDA = ("VENDA", "MORTE", "SOCIO")
# SAIDA-SOCIO só fica fora do Δ quando quem sai é RECEPTORA: ela só é contada em
# PAO GRANDE/ARRENDAMENTO, então ir pro sócio a tira de lá mas não mexe no Δ de
# ANIMAIS (é outro total, tratado à parte — ver _refina_afeta_headcount). Era regra
# geral pra qualquer 'SOCIO' e estava errada: em 28/08/2026 a GIM MATIZA (GARANHÃO,
# destino nomeado 'VALTER LIMA') saiu de verdade, física — Δ oficial contou as DUAS
# saídas da semana (-02), não só a venda. Physical is physical: se o bicho deixou a
# fazenda, conta, sócio ou não.
CLASSIF_FORA_DO_DELTA = ("SOCIO",)


def _classificar_se(classif: str, animal: str = ""):
    """(sentido, afeta_headcount) da classificação da aba SAIDAS-ENTRADAS.
    (None, None) = vocabulário desconhecido; quem chama tem de avisar, não engolir."""
    def _sai():
        eh_socio_exempt = (any(c in classif for c in CLASSIF_FORA_DO_DELTA)
                           and _norm(animal).startswith("RECEPTORA"))
        return "SAIDA", not eh_socio_exempt
    if classif.startswith("ENTRADA"):
        return "ENTRADA", True
    if classif.startswith("SAIDA"):
        return _sai()
    if any(c in classif for c in CLASSIF_ENTRADA):
        return "ENTRADA", True
    if any(c in classif for c in CLASSIF_SAIDA):
        return _sai()
    return None, None


def _saidas_entradas_planilha(wb, ini: date, fim: date):
    """Lê a aba SAIDAS-ENTRADAS. Devolve None se ela não existir ou estiver vazia,
    pra que o cálculo caia na diferença de roster (comportamento atual)."""
    if "SAIDAS-ENTRADAS" not in wb.sheetnames:
        return None
    evs = {"SAIDA": [], "ENTRADA": []}
    achou = False
    desconhecidas = []
    for i, r in enumerate(wb["SAIDAS-ENTRADAS"].iter_rows(values_only=True), start=1):
        if i < 3 or r[1] is None or not str(r[1]).strip():
            continue
        achou = True
        d = _dt(r[4])
        if not d or not (ini <= d <= fim):
            continue
        classif = _norm(r[5])
        alvo, afeta = _classificar_se(classif)
        if alvo is None:
            # classificação nova na planilha: avisar, nunca virar zero calado
            desconhecidas.append(f"linha {i} ({_s(r[1])}): {_s(r[5])!r}")
            continue
        evs[alvo].append({"animal": _s(r[1]), "data": d.isoformat(),
                          "classificacao": _s(r[5]), "afeta_headcount": afeta,
                          "local_saida": _s(r[2]), "local_entrada": _s(r[3])})
    if desconhecidas:
        print(f"  [SAIDAS-ENTRADAS] classificação não reconhecida em "
              f"{len(desconhecidas)} linha(s) da janela — NÃO entraram na conta: "
              + "; ".join(desconhecidas))
    return evs if achou else None


# TRANSFERÊNCIA INTERNA = animal que trocou de LOCAL entre os dois locais próprios
# (FPG <-> arrendamento) na aba ANIMAIS do PLANTEL ARRENDAMENTOS E RECEPTORAS.
# A aba MOVIMENTAÇÕES do controle mensal NÃO serve: a última transferência lançada
# lá é de setembro/2025. Em 07/08/2026 o relatório oficial disse 14 transferências e
# o diff de LOCAL dá exatamente 14 (5 arrendamento->FPG, 9 FPG->arrendamento).
# Igual a saídas/entradas, é diff: precisa do mapa da semana anterior. Sem ele
# (primeira captura), fica em branco — nunca zero.
def _receptoras_arquivos() -> list:
    """Arquivos de receptoras, do mais recente pro mais antigo (por mtime)."""
    cands = [f for d in _estacao_dirs()
             for f in d.glob("*PLANTEL ARRENDAMENTOS E RECEPTORAS.xlsx")
             if not f.name.startswith("~$")]
    return sorted(cands, key=lambda f: f.stat().st_mtime, reverse=True)


def _receptoras_info(src: Path | None = None) -> dict:
    """{ANIMAL: {local, status, embriao, obs}} da aba ANIMAIS — TODAS as linhas,
    inclusive fora dos nossos locais, pra saber pra onde o animal foi."""
    if src is None:
        src = _latest_no_plantel("*PLANTEL ARRENDAMENTOS E RECEPTORAS.xlsx", "receptoras")
    wb = _load(src)
    ws = wb["ANIMAIS"]
    out = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 4 or r[1] is None:
            continue
        out[_norm(r[1])] = {"local": _norm(r[3]), "status": _s(r[2]),
                            "embriao": _s(r[4]), "obs": _s(r[5])}
    wb.close()
    return out


def _receptoras_locais(src: Path | None = None) -> dict:
    """{ANIMAL: LOCAL} da aba ANIMAIS, só quem está num local nosso."""
    return {k: v["local"] for k, v in _receptoras_info(src).items()
            if v["local"] in RECEPTORAS_LOCAIS_ATIVOS}


def _transferencias_internas(rep: Report) -> list | None:
    """Diff do LOCAL vs o mapa da semana anterior. None = sem base de comparação."""
    hist = _load_hist()
    prev = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("receptoras_locais"):
            prev = hist[wid]["receptoras_locais"]
    cur = rep.receptoras_locais or {}
    if not cur:
        return None
    if not prev:
        # Bootstrap: nenhuma semana anterior guardou o mapa (o campo é novo). Cai no
        # arquivo de receptoras anterior, que é o retrato de onde os animais estavam.
        anteriores = _receptoras_arquivos()[1:]
        if not anteriores:
            print("  [transferências] sem mapa de LOCAL da semana anterior e sem "
                  "arquivo anterior de receptoras — fica EM BRANCO, não zero")
            return None
        prev = _receptoras_locais(anteriores[0])
        print(f"  [transferências] primeira semana com mapa de LOCAL: comparando com "
              f"{anteriores[0].name} (da próxima em diante, compara com o snapshot)")
    return [{"animal": k, "local_saida": prev[k], "local_entrada": cur[k]}
            for k in cur if k in prev and prev[k] != cur[k]]


def build_movimentacao(rep: Report, ini: date, fim: date):
    src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    rep.fontes["controle_plantel_mensal"] = src.name
    wb = _load(src)
    rep.saidas_planilha = _saidas_entradas_planilha(wb, ini, fim)
    ws = wb["MOVIMENTAÇÕES"]
    evs = {"SAIDA": [], "ENTRADA": [], "TRANSFERENCIA": []}
    ultimo = None  # último lançamento DATADO da aba, de qualquer tipo
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 3 or r[3] is None:
            continue
        d = _dt(r[3])
        if not d:
            continue
        if ultimo is None or d > ultimo:
            ultimo = d
        tipo = _categorize_mov(str(r[4] or "").upper())
        if tipo is None:
            continue
        evs[tipo].append({"animal": _s(r[2]), "data": d.isoformat(), "ocorrencia": _s(r[4])})
    inw = lambda x: ini <= date.fromisoformat(x["data"]) <= fim
    # A aba só mede a semana se ela chegou na semana. Parada antes da janela, contar
    # zero afirmaria "não houve movimentação" quando o que existe é falta de
    # lançamento — foi o que fez transferências virar 0 com 14 no relatório oficial.
    rep.saidas = {
        "saidas_semana": sum(1 for x in evs["SAIDA"] if inw(x)),
        "entradas_semana": sum(1 for x in evs["ENTRADA"] if inw(x)),
    }
    # transferências não vêm daqui (ver comentário acima de _receptoras_locais)
    rep.receptoras_locais = _receptoras_locais()
    transf = _transferencias_internas(rep)
    rep.saidas["transferencias_semana"] = len(transf) if transf is not None else None
    rep.detalhe["transferencias_internas"] = transf
    rep.saidas["movimentacao_ultimo_lancamento"] = ultimo.isoformat() if ultimo else None
    if ultimo is None or ultimo < ini:
        rep.saidas["movimentacao_defasada"] = True
        print(f"  [movimentação] {src.name}: último lançamento datado é "
              f"{ultimo.strftime('%d/%m/%Y') if ultimo else 'nenhum'}, antes da janela "
              f"({ini.strftime('%d/%m')}) — a aba MOVIMENTAÇÕES não mede esta semana")
    # listas COMPLETAS datadas p/ filtro client-side
    rep.eventos["saidas"] = evs["SAIDA"]
    rep.eventos["entradas"] = evs["ENTRADA"]
    rep.eventos["transferencias"] = evs["TRANSFERENCIA"]
    wb.close()


# ------------------------------------------------------------------
# Seção 4b — PENDENTES DE SAÍDA / TERCEIROS  (CONTROLE PLANTEL aba PLANTEL)
# ------------------------------------------------------------------
# Reunião com o haras (2026-07-31): vendidos pendentes e terceiros passam a sair do
# roster, pela coluna STATUS PLANTEL.
#
# O haras CUMPRIU — só que no CONTROLE_DE_PLANTEL_PAO_GRANDE mensal (pasta PLANTEL),
# não na cópia semanal. Vocabulário de STATUS PLANTEL em 14/08/2026:
#   semanal (CONTROLE PLANTEL.xlsx) : PLANTEL 143, VENDIDO 6
#   mensal  (CONTROLE_DE_PLANTEL...) : PLANTEL 224, VENDIDO E ENTREGUE 108, OBITO 26,
#                                      DOADO 16, DE TERCEIRO 8, VENDIDO PENDENTE SAIDA 6,
#                                      VENDIDO 3
# Lendo o arquivo errado, os dois indicadores morriam em silêncio: vendidos caía no
# 'Animais para sair', congelado em 24/07 (2 em vez de 6, faltando PATRIMONIO, PODIO e
# PAETE) e terceiros saía 2 em vez de 8. O relatório oficial de 14/08 traz 08 terceiros
# e 06 animais vendidos pendentes — os números do MENSAL.
STATUS_VENDIDO_PENDENTE = "VENDIDO PENDENTE"
STATUS_TERCEIRO = "TERCEIRO"
# Sociedade pendente de animal: até 28/08/2026 não tinha marca nenhuma (comentário
# antigo em build_pendentes: "sociedade nunca recebe marca"), então soc_animais
# ficava sempre vazio. O haras passou a marcar na coluna OBS (Y, índice 24) — mesma
# ideia do STATUS_VENDIDO_PENDENTE, coluna diferente.
STATUS_SOCIEDADE_PENDENTE = "EM SOCIEDADE PENDENTE DE SAIDA"
COL_MENSAL_OBS = 24

# Status que significam "o animal AINDA ESTÁ AQUI".
#
# Vendido conta: a venda foi fechada mas o animal não saiu da fazenda, e enquanto
# não sai ele é headcount. O que tira da conta é a ENTREGA — por isso 'VENDIDO E
# ENTREGUE' fica de fora, junto de DOADO, OBITO e DE TERCEIRO.
#
# A comparação é EXATA, não por substring: 'VENDIDO' como pedaço de texto casaria
# com 'VENDIDO E ENTREGUE' e traria de volta justamente quem já foi embora.
STATUS_NO_PLANTEL = ("PLANTEL", "VENDIDO", "VENDIDO PENDENTE SAIDA")
# Categoria que não é animal do headcount: embrião não nasceu; receptora é contada
# pela planilha de receptoras, e somar aqui duplicaria.
CATEGORIAS_FORA_DO_HEADCOUNT = ("EMBRIAO", "RECEPTORA")
# Colunas do mensal usadas só aqui. COTAS (%) é a fatia que a PG ainda tem; CONDICAO
# ATUAL registra a movimentação física do animal.
COL_MENSAL_COTAS = 16
COL_MENSAL_CONDICAO = 18
CONDICAO_SAIU = "SAIU DO HARAS"
# LOCAL que significa "está aqui". 'OUTROS' fica de fora de propósito: apesar de a
# CONTAGEM usar esse rótulo para o Centro de Treinamento, na prática ele é o destino
# de quem sai ("MUDOU O LOCAL PARA OUTROS" na aba MOVIMENTAÇÕES). Se o haras confirmar
# que OUTROS é só o CTE, é aqui que se acrescenta.
LOCAIS_NA_PROPRIEDADE = ("FAZENDA PAO GRANDE", "ARRENDAMENTO CESAR FURTADO")
# Layout da aba PLANTEL nos dois arquivos (0-based). O mensal é o mesmo roster com 3
# colunas a mais na frente e cabeçalho 3 linhas abaixo — por isso não dá pra apontar
# o mesmo leitor pros dois sem parametrizar.
PLANTEL_LAYOUT_SEMANAL = {"linha1": 2, "nome": 0, "categoria": 2, "status": 3, "local": 4}
PLANTEL_LAYOUT_MENSAL = {"linha1": 5, "nome": 3, "categoria": 5, "status": 6, "local": 7}


ROSTER_FONTE = "controle_mensal"      # gravado no snapshot; ver _conferir_delta


# "ainda está aqui" para o embrião = onde a receptora dele está. LOCAL SOCIO ou
# COMPRADOR significa que já saiu.
LOCAIS_RECEPTORA_NA_PG = ("PAO GRANDE", "ARRENDAMENTO CESAR FURTADO")


# Embrião de sociedade: 100% do sócio. Na aba ESTAÇÃO isso é COTAS EMBRIÃO vazia
# (ou zero) com SÓCIO EMBRIÃO = 1. Cota parcial (0,25 / 0,5) é embrião da PG COM
# sócio, que não é pendência de saída — a planilha do grupo dizia o mesmo separando
# em duas abas.
def _e_sociedade(cota, socio) -> bool:
    c = _to_num(cota)
    s = _to_num(socio)
    return (c is None or c == 0) and s is not None and s >= 1


def _embrioes_sociedade_pendentes() -> list:
    """Embriões em sociedade que ainda estão na PG.

    Regra do haras: a `EMBRIOES A ENTREGAR` NÃO serve aqui — ali só tem embrião
    vendido e não gestado, sem sociedade e sem vendido já gestado. O que vale é
    esta aba, com STATUS SOCIO e o embrião ainda em terra nossa."""
    locais = _receptoras_locais()
    wb = _load(_latest_estacao_master())
    ws = wb["ESTAÇÃO"]
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 3 or r[0] is None or _s(r[35]) != SAFRA_ATUAL:
            continue
        if not _e_sociedade(r[5], r[6]):
            continue
        if _norm(r[17]) != "OK":            # não confirmado, não é pendência
            continue
        if r[21] is not None or r[23] is not None:   # abortou ou já pariu
            continue
        local = _norm(locais.get(_norm(r[11])))
        if local not in LOCAIS_RECEPTORA_NA_PG:      # já está com o sócio: saiu
            continue
        out.append({
            "nome": f"{_s(r[2])} x {_s(r[3])}",
            "receptora": _s(r[11]),
            "local": local,
            "estacao": _s(r[35]),
            "socio": _s(r[34]),
            "tipo": "SOCIEDADE",
            "especie": "EMBRIAO",
        })
    wb.close()
    return out


def _plantel_por_status() -> dict:
    """Roster do plantel a partir do CONTROLE_DE_PLANTEL mensal, na pasta PLANTEL.

    Antes vinha do `CONTROLE PLANTEL.xlsx` da pasta ATUALIZACAO SEMANAL, que é de
    divulgação. As regras estão no cabeçalho do commit e resumidas abaixo; cada uma
    saiu de comparar nome a nome os dois rosters, não de suposição.

    O nome do animal NÃO é identidade aqui: o mensal batiza o potro
    (`PRINCIPE MN DA PAO GRANDE`) enquanto o semanal o descrevia pelo cruzamento
    (`MACHO LIBRA DA PAO GRANDE X OLIMPO DO MH`). Quem identifica é nome+MAE+PAI."""
    src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    wb = _load(src)
    ws = wb["PLANTEL"]
    L = PLANTEL_LAYOUT_MENSAL
    vistos, linhas = {}, []
    fora = {"status": 0, "categoria": 0, "duplicado": 0}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < L["linha1"] or r[L["nome"]] is None:
            continue
        nome = _s(r[L["nome"]])
        if not nome:
            continue
        if _norm(r[L["status"]]) not in STATUS_NO_PLANTEL:
            fora["status"] += 1
            continue
        categoria = _norm(r[L["categoria"]])
        if categoria in CATEGORIAS_FORA_DO_HEADCOUNT:
            fora["categoria"] += 1
            continue
        # Saiu do haras E a PG não tem cota nenhuma: acabou. Uma coisa só não basta —
        # animal no sócio segue no plantel enquanto a PG tem parte dele (29 estão
        # nessa situação), e vendido de cota zero que ainda não saiu continua aqui
        # (os 5 'vendido pendente saída', o PRADO). Juntas, as duas dizem que o
        # animal não é mais da casa nem está mais nela.
        cota = r[COL_MENSAL_COTAS] if len(r) > COL_MENSAL_COTAS else None
        condicao = _norm(r[COL_MENSAL_CONDICAO]) if len(r) > COL_MENSAL_CONDICAO else ""
        if condicao == CONDICAO_SAIU and (cota is None or not cota):
            fora["saiu_sem_cota"] = fora.get("saiu_sem_cota", 0) + 1
            continue
        # STATUS='VENDIDO' cobre dois sentidos opostos: venda NOSSA ainda não
        # entregue (conta — animal continua aqui) e COMPRA ainda não entregue pelo
        # vendedor (NÃO conta — animal ainda não chegou). Achado em 28/08/2026: a
        # ELEITA tinha STATUS='VENDIDO' e entrava no headcount como se já estivesse
        # na fazenda, mas a OBS diz "o vendedor entregará" — é compra, não saiu daqui
        # coisa nenhuma porque nunca chegou. Único marcador vivo hoje é essa frase
        # na OBS; se aparecer outro caso com texto diferente, ajustar aqui.
        obs = _norm(r[COL_MENSAL_OBS]) if len(r) > COL_MENSAL_OBS else ""
        if _norm(r[L["status"]]) == "VENDIDO" and "VENDEDOR ENTREGARA" in obs:
            fora["compra_nao_entregue"] = fora.get("compra_nao_entregue", 0) + 1
            continue
        mae = _s(r[COL_MENSAL_MAE]) if len(r) > COL_MENSAL_MAE else None
        pai = _s(r[COL_MENSAL_PAI]) if len(r) > COL_MENSAL_PAI else None
        chave = _chave_animal(nome, mae, pai)
        if chave in vistos:
            fora["duplicado"] += 1
            continue
        # o nome sem cotista é o que vai para o roster: com o cotista, o mesmo animal
        # apareceria como saída de uma semana e entrada na outra ao mudar de sócio
        limpo = _sem_cotista(nome)
        vistos[chave] = limpo
        linhas.append({"nome": limpo, "categoria": _s(r[L["categoria"]]),
                       "status_plantel": _s(r[L["status"]]), "local": _s(r[L["local"]]),
                       "mae": mae, "pai": pai})
    wb.close()
    compra_pend = fora.get("compra_nao_entregue", 0)
    extra = f", {compra_pend} compra(s) ainda não entregue(s)" if compra_pend else ""
    print(f"  [roster] {len(vistos)} animais em {src.name} "
          f"(fora: {fora['status']} por status, {fora['categoria']} embrião/receptora, "
          f"{fora['duplicado']} linha(s) repetida(s) por cotista{extra})")
    return {"roster": sorted(set(vistos.values())), "linhas": linhas,
            "fonte": src.name, "roster_fonte": ROSTER_FONTE}


# CONTROLE_DE_PLANTEL mensal, aba PLANTEL: colunas que não estão no layout mínimo.
# MAE/PAI/NASCIMENTO são o que permite achar o potro sem depender do nome dele.
COL_MENSAL_MAE = 8
COL_MENSAL_PAI = 9
COL_MENSAL_NASCIMENTO = 10


def _nascimentos_do_roster(ini: date, fim: date) -> list:
    """Nascimentos da janela pela coluna NASCIMENTO do roster mensal.

    Por data e filiação, nunca por nome: o potro entra no roster com nome próprio
    (`PRINCIPE MN DA PAO GRANDE`) ou com o cruzamento (`MACHO LIBRA x OLIMPO`),
    e as duas formas convivem. Data e MAE/PAI existem nas duas."""
    src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    wb = _load(src)
    ws = wb["PLANTEL"]
    L = PLANTEL_LAYOUT_MENSAL
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < L["linha1"] or r[L["nome"]] is None:
            continue
        d = _dt(r[COL_MENSAL_NASCIMENTO]) if len(r) > COL_MENSAL_NASCIMENTO else None
        if not d or not (ini <= d <= fim):
            continue
        nome = _s(r[L["nome"]])
        m = RE_PRODUTO.search(_norm(nome))
        out.append({
            "produto": nome,
            "mae": _s(r[COL_MENSAL_MAE]), "pai": _s(r[COL_MENSAL_PAI]),
            "receptora": m.group(1) if m else None,
            "socio": _limpa_socio(r[COL_MENSAL_NOME_SOCIO])
                     if len(r) > COL_MENSAL_NOME_SOCIO else None,
            "data": d.isoformat(),
            "local": _s(r[L["local"]]),
        })
    wb.close()
    return out


def _status_plantel_mensal() -> dict:
    """Vendidos pendentes e terceiros pela coluna STATUS PLANTEL do CONTROLE_DE_PLANTEL
    mensal. Devolve as listas cruas; quem chama decide o que é zero e o que é ausência."""
    src = _latest_no_plantel("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx", "controle mensal")
    wb = _load(src)
    ws = wb["PLANTEL"]
    L = PLANTEL_LAYOUT_MENSAL
    vendidos_pend, terceiros, soc_pend, marcado = [], [], [], False
    # NOME SOCIO / COTAS (%) ficam fora do layout mínimo porque só este trecho usa.
    # Indexado pela receptora do fim do nome: é o que o roster semanal e o mensal
    # têm em comum (o mensal escreve a data no meio e acentua o garanhão).
    socio_por_recep = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < L["linha1"] or r[L["nome"]] is None:
            continue
        nome = _s(r[L["nome"]])
        if not nome:
            continue
        status_plantel = _norm(r[L["status"]])
        categoria, local = _norm(r[L["categoria"]]), _s(r[L["local"]])
        m_rec = RE_PRODUTO.search(_norm(nome))
        soc = _limpa_socio(r[COL_MENSAL_NOME_SOCIO]) if len(r) > COL_MENSAL_NOME_SOCIO else None
        if m_rec and soc:
            socio_por_recep[m_rec.group(1)] = soc
        if STATUS_VENDIDO_PENDENTE in status_plantel:
            marcado = True
            vendidos_pend.append({"nome": nome, "local": local, "cota": None,
                                  "comprador": None, "tipo": "VENDA",
                                  "obs": _s(r[L["status"]]), "reposicao": False,
                                  "categoria": categoria,
                                  "especie": "EMBRIAO" if categoria == "EMBRIAO" else None})
        if STATUS_TERCEIRO in status_plantel:
            marcado = True
            terceiros.append({"nome": nome, "local": local, "categoria": categoria,
                              "status_plantel": _s(r[L["status"]]),
                              "especie": "EMBRIAO" if categoria == "EMBRIAO" else None})
        obs = _norm(r[COL_MENSAL_OBS]) if len(r) > COL_MENSAL_OBS else ""
        if STATUS_SOCIEDADE_PENDENTE in obs:
            soc_pend.append({"nome": nome, "local": local, "categoria": categoria,
                             "obs": _s(r[COL_MENSAL_OBS]),
                             "especie": "EMBRIAO" if categoria == "EMBRIAO" else None})
    wb.close()
    return {"fonte": src.name, "marcado": marcado,
            "vendidos_pendentes": vendidos_pend, "terceiros": terceiros,
            "sociedade_pendentes": soc_pend, "socio_por_recep": socio_por_recep}


# Embrião comercial pendente de saída: aba ENTREGAR do "EMBRIOES A ENTREGAR - A
# RECEBER". 'Cota PG' < 1 = sociedade; = 1 = venda 100%.
#
# PRONTO tem DOIS estados e só um deles é pendência de saída:
#   PRONTO - AGUARDANDO ENTREGA  (4)  o embrião vai embora  -> PENDENTE
#   PRONTO - NASCE NA PG         (3)  o produto nasce aqui  -> NÃO é pendência
# Casar só o prefixo 'PRONTO' misturava os dois e inflava o card de sociedade em 3.
# Os demais estados — A FAZER, ENTREGUE, NASCIDO, CANCELADO, REPOSIÇÃO — já ficavam
# de fora. A aba EMBRIOES VENDIDOS do "Animais para sair" está vazia e não é usada.
EMB_STATUS_PENDENTE = "AGUARDANDO ENTREGA"


def _embrioes_pendentes() -> list:
    """Embriões prontos e aguardando entrega, tipo SOCIEDADE (cota parcial) ou VENDA."""
    wb = _load(EMB_COMERCIAIS)
    ws = wb["ENTREGAR"]
    out, cols, ficam = [], None, []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 3:
            cols = {n: _col_idx(r, n) for n in
                    ("ID Embrião", "Doadora", "Garanhão", "Comprador", "Cota PG",
                     "Status embrião", "Observação")}
            continue
        if cols is None or r[cols["ID Embrião"]] is None:
            continue
        status = _norm(r[cols["Status embrião"]])
        if EMB_STATUS_PENDENTE not in status:
            if status.startswith("PRONTO"):
                ficam.append(f'{_s(r[cols["ID Embrião"]])} ({_s(r[cols["Status embrião"]])})')
            continue
        cota = r[cols["Cota PG"]]
        try:
            parcial = cota is not None and float(cota) < 1
        except (TypeError, ValueError):
            parcial = False
        out.append({
            "nome": f'{_s(r[cols["Doadora"]])} x {_s(r[cols["Garanhão"]])}',
            "id": _s(r[cols["ID Embrião"]]), "local": None, "cota": cota,
            "comprador": _s(r[cols["Comprador"]]),
            "tipo": "SOCIEDADE" if parcial else "VENDA",
            "obs": _s(r[cols["Status embrião"]]), "reposicao": False,
            "especie": "EMBRIAO",
        })
    if ficam:
        print(f"  [embriões] {len(ficam)} pronto(s) que NÃO saem, fora da pendência: "
              + "; ".join(ficam))
    wb.close()
    return out


def build_pendentes(rep: Report):
    plantel = _plantel_por_status()
    rep.roster = plantel["roster"]
    _LINHAS_BRUTAS["roster"] = plantel["linhas"]
    rep.fontes["roster_plantel"] = plantel["fonte"]

    # O "Animais para sair" saiu do pipeline: mora na pasta de divulgação (o que foi
    # enviado ao grupo), está congelado em 24/07/2026 e só alimentava a marcação de
    # reposição, que já era apenas um aviso. Vendido pendente vem do STATUS PLANTEL do
    # controle mensal; sociedade, da aba ESTAÇÃO.
    pend = []
    pend_emb = _embrioes_pendentes()

    # VENDIDOS PENDENTES: fonte é o STATUS PLANTEL do controle mensal. O "Animais para
    # sair" só entra se ninguém estiver marcado lá — e aí com aviso, porque ele está
    # congelado desde 24/07/2026 e perde os marcados depois disso.
    mensal = _status_plantel_mensal()
    rep.fontes["status_plantel"] = mensal["fonte"]
    _SOCIO_ROSTER.clear()
    _SOCIO_ROSTER.update(mensal.get("socio_por_recep") or {})
    mensal_terceiros_embrioes = [t for t in mensal["terceiros"]
                                 if t.get("especie") == "EMBRIAO"]
    # O embrião dos vendidos pendentes vem do ROSTER MENSAL, não do EMBRIOES A
    # ENTREGAR. O relatório abre "07 (05 animais e 02 embriões)" e o roster tem
    # exatamente 5 com VENDIDO PENDENTE SAIDA e 2 com DE TERCEIRO + CATEGORIA=EMBRIAO
    # — mesma fonte, mesmo número, mesma abertura, e é o mesmo conjunto que ele
    # publica em "Total terceiros: 07 (vendidos pendentes)".
    # Antes puxávamos o embrião de cota integral do ENTREGAR (1 em vez de 2): aquilo
    # é a fila de entrega comercial, outra coisa.
    vend_embrioes = [t for t in mensal_terceiros_embrioes]
    # REPOSIÇÃO não é venda pendente: o animal está saindo para repor outro, não para
    # um comprador. O STATUS PLANTEL não tem essa marca — ela vive na coluna de obs do
    # Animais para sair —, então cruzamos os dois pelo núcleo do nome. Essa regra
    # existia antes da migração para o STATUS PLANTEL e se perdeu no caminho: era o que
    # separava os nossos 5 dos 4 do relatório.
    # REPOSIÇÃO CONTA. Cheguei a excluir, apoiado num rascunho de relatorio que dizia
    # "04 animais"; a conferencia por movimentacao derrubou isso — o STATUS PLANTEL tem
    # 5 marcados e o relatorio fechado diz "05 animais", reposicao inclusa. Fica so o
    # aviso, porque a natureza da saida e diferente e alguem pode querer separar.
    reposicoes = {_nucleo_nome(x["nome"]) for x in pend if x["reposicao"]}
    if mensal["vendidos_pendentes"]:
        marcados = mensal["vendidos_pendentes"]
        repostos = [x for x in marcados if _nucleo_nome(x["nome"]) in reposicoes]
        if repostos:
            print("  [pendentes] entre os vendidos pendentes ha reposição (sai para "
                  "repor outro animal, nao para comprador): "
                  + "; ".join(x["nome"] for x in repostos))
        vendidos = marcados + vend_embrioes
        fonte_vendidos = "status_plantel"
    else:
        vendidos = [p for p in pend if p["tipo"] == "VENDA" and not p["reposicao"]]
        fonte_vendidos = "animais_para_sair"
        print(f"  [terceiros] nenhum '{STATUS_VENDIDO_PENDENTE}' no STATUS PLANTEL de "
              f"{mensal['fonte']}; vendidos pendentes caindo no Animais para sair "
              f"(congelado em 24/07/2026)")

    # SOCIEDADE pendente = animais + embriões (regra do relatório desde 07/08/2026).
    # Até 28/08/2026 não havia marca viva pra animal em sociedade — só o "Animais
    # para sair", congelado em 24/07 — e soc_animais ficava sempre vazio (o
    # relatório de 21/08 confirmava: "01 (embrião)", nenhum animal). O haras passou
    # a marcar direto na coluna OBS do roster mensal com a frase
    # STATUS_SOCIEDADE_PENDENTE, mesma ideia do VENDIDO PENDENTE SAIDA — então a
    # fonte agora é viva e o teste vira leitura direta, igual aos vendidos.
    soc_animais = mensal["sociedade_pendentes"]
    # Embrião de sociedade vem da aba de sócios do grupo, não do 'EMBRIOES A
    # ENTREGAR' — ver _embrioes_sociedade_pendentes.
    soc_embrioes = _embrioes_sociedade_pendentes()
    sociedade = soc_animais + soc_embrioes
    rep.fontes["embrioes_pendentes"] = EMB_COMERCIAIS.name

    # TERCEIROS NA PROPRIEDADE = o que é de terceiro e ainda está aqui, ou seja, os
    # PENDENTES DE SAÍDA — é o que o próprio relatório diz no rótulo da linha:
    # "Total terceiros: 08 (vendidos pendentes)".
    #
    # NÃO é a contagem de STATUS PLANTEL 'DE TERCEIRO'. Aquilo são doadoras, matriz,
    # receptora e embriões de terceiro que passaram pela estação: em 14/08/2026 dava 8
    # e batia com o relatório por coincidência de número, não por ser a mesma coisa.
    # Nenhuma das 8 linhas está no roster do plantel, e 7 das 8 estão fora da fazenda.
    #
    # "NA PROPRIEDADE" é presença física — embrião não ocupa espaço, então não conta
    # aqui mesmo sendo 'vendido pendente'. Em 14/08 os dois totais empatavam (8 e 8,
    # sem embrião no lote), o que escondia a diferença; em 28/08 o relatório abriu os
    # dois: linha 4 "05 (vendidos pendentes)" só animal, linha 5 "07 (05 animais e 02
    # embriões)" com embrião. `terceiros = vendidos` tratava as duas linhas como a
    # mesma contagem e ficou errado assim que apareceu embrião no lote.
    terc_embrioes = [t for t in vendidos if t.get("especie") == "EMBRIAO"]
    terc_animais = [t for t in vendidos if t.get("especie") != "EMBRIAO"]
    terceiros = terc_animais

    # DOADORAS DE TERCEIROS: doadora de terceiro que está NA PROPRIEDADE — fazenda ou
    # arrendamento. As marcadas em LOCAL 'OUTROS' não contam: 'OUTROS' é para onde o
    # animal vai quando deixa o haras (a aba MOVIMENTAÇÕES registra "MUDOU O LOCAL PARA
    # OUTROS" nas saídas), e nenhuma delas aparece no roster do plantel. Contá-las dava
    # 4 onde o relatório escreve "--".
    de_terceiro = mensal["terceiros"]
    doadoras_terc = [t for t in de_terceiro
                     if t["categoria"] == "DOADORA"
                     and _norm(t["local"]) in LOCAIS_NA_PROPRIEDADE]
    fora = [t for t in de_terceiro
            if t["categoria"] == "DOADORA" and _norm(t["local"]) not in LOCAIS_NA_PROPRIEDADE]
    if fora:
        print(f"  [terceiros] {len(fora)} doadora(s) de terceiro fora da propriedade, "
              f"não contadas: " + "; ".join(f"{t['nome']} ({t['local']})" for t in fora))
    if not mensal["marcado"]:
        print(f"  [terceiros] nenhuma linha marcada com {STATUS_TERCEIRO} ou "
              f"{STATUS_VENDIDO_PENDENTE} em {mensal['fonte']}")

    rep.terceiros.update({
        "vendidos_pendentes": len(vendidos),
        "vendidos_pendentes_animais": len(terc_animais),
        "vendidos_pendentes_embrioes": len(terc_embrioes),
        "vendidos_pendentes_fonte": fonte_vendidos,
        "sociedade_pendentes": len(sociedade),
        "sociedade_pendentes_animais": len(soc_animais),
        "sociedade_pendentes_embrioes": len(soc_embrioes),
        "terceiros_propriedade": len(terceiros),
        "terceiros_animais": len(terc_animais),
        "terceiros_embrioes": len(terc_embrioes),
        "doadoras_terceiros": len(doadoras_terc) if doadoras_terc else None,
        "outros_terceiros": None,
    })
    rep.detalhe["doadoras_terceiros"] = doadoras_terc
    rep.detalhe["terceiros_propriedade"] = terceiros
    rep.detalhe["terceiros_vendidos"] = vendidos          # vendidos pendentes (KPI seção 5)
    # A lista de vendidos pendentes vai logo abaixo da seção 4 (é o mesmo conjunto do
    # "Total terceiros", já que o relatório oficial escreve "05 (vendidos pendentes)"
    # — as duas linhas SÃO a mesma coisa). Embrião pendente de venda tem lista própria
    # na seção 5, porque não é "terceiro na propriedade": embrião não ocupa espaço.
    rep.detalhe["terceiros_vendidos_embrioes"] = terc_embrioes
    rep.detalhe["terceiros_sociedade"] = sociedade        # sociedade pendente de saída, listada igual
    # lista completa da seção 5 = o que os dois KPIs contam. Era `pend + pend_emb` (só
    # o "Animais para sair"), então os marcados no STATUS PLANTEL não apareciam.
    rep.detalhe["pendentes_saida"] = vendidos + sociedade


def _latest_animais_sair() -> Path:
    """'Animais para sair*.xlsx' — hoje só sociedade pendente (vendidos migraram pro
    STATUS PLANTEL). Pasta canônica é VENDAS/SAIDA DE ANIMAIS VENDIDOS."""
    return _resolver(ANIMAIS_SAIR_GLOB, ANIMAIS_SAIR_DIRS, "sociedade pendente",
                     requer_aba="ANIMAIS VENDIDOS")


# ------------------------------------------------------------------
# Δ headcount vs run anterior (histórico local leve, automático)
# ------------------------------------------------------------------
def build_headcount_delta(rep: Report, fim: date):
    HIST_HEADCOUNT.parent.mkdir(parents=True, exist_ok=True)
    hist = {}
    if HIST_HEADCOUNT.exists():
        try:
            hist = json.loads(HIST_HEADCOUNT.read_text(encoding="utf-8"))
        except Exception:
            hist = {}
    total = rep.headcount.get("total")
    # run anterior = maior data < fim
    prev = None
    for k in sorted(hist):
        if k < fim.isoformat():
            prev = hist[k]
    if prev is not None and total is not None:
        rep.headcount["delta"] = total - prev.get("total", total)
    else:
        rep.headcount["delta"] = None

    # Δ do TOTAL nao e o mesmo que o Δ do relatorio.
    # O relatorio escreve '+02 / -01' contando so ANIMAIS: 2 potros nascidos, 1
    # vendido. As receptoras que foram pro socio sairam da contagem (-2) e nao
    # aparecem ali. Resultado em 21/08/2026: total 203 -> 202 = -1, animais
    # 143 -> 144 = +1. As duas contas estao certas, medem coisas diferentes — e
    # comparar a nossa do total contra a dele de animais dava divergencia falsa.
    # Nas semanas sem movimento de receptora as duas coincidem, e por isso o
    # problema so apareceu quando duas receptoras sairam na mesma semana.
    det = rep.headcount.get("detalhe") or {}
    tg = det.get("TOTAL GERAL") or {}
    ani, rec = tg.get("animais"), tg.get("receptoras")
    pa, pr = prev.get("animais") if prev else None, prev.get("receptoras") if prev else None
    rep.headcount["delta_animais"] = (ani - pa) if None not in (ani, pa) else None
    rep.headcount["delta_receptoras"] = (rec - pr) if None not in (rec, pr) else None
    atual = {"total": total, "fpg": rep.headcount.get("fazenda_pg"),
             "arr": rep.headcount.get("arrendamento"),
             "cte": rep.headcount.get("cte"), "soc": rep.headcount.get("socio"),
             # abertura animais/receptoras: base do Δ de animais, que e o que o
             # relatorio publica
             "animais": tg.get("animais"), "receptoras": tg.get("receptoras")}
    # CONTAGEM idêntica à da semana passada, local por local, quase sempre significa
    # que a aba não foi atualizada — não que nada mudou. Em 31/07/2026 isso aconteceu:
    # o snapshot repetiu 205 de 24/07, Δ saiu 0, e o relatório oficial dizia 204 / -01.
    if prev is not None and atual == prev:
        print(f"  [headcount] CONTAGEM idêntica à de {max(k for k in hist if k < fim.isoformat())} "
              f"em todos os locais ({total} total) — conferir se a aba foi atualizada; "
              f"Δ desta semana sai 0 por isso")
    # grava snapshot desta run (idempotente por data)
    hist[fim.isoformat()] = atual
    HIST_HEADCOUNT.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")


# ------------------------------------------------------------------
# Seção 4 — TERCEIROS / COMERCIAIS (embriões a entregar / a receber)
# ------------------------------------------------------------------
def build_comerciais(rep: Report):
    wb = _load(EMB_COMERCIAIS)
    painel = {}
    ws = wb["PAINEL"]
    grid = list(ws.iter_rows(values_only=True))
    for r in grid:
        cells = [_s(c) for c in r]
        for j, c in enumerate(cells):
            if c and c.upper() in ("A FAZER", "EM ANDAMENTO", "TOTAL VENDIDOS", "TOTAL COMPRADOS"):
                # o número costuma estar mais à direita na mesma linha
                nums = [x for x in r if isinstance(x, (int, float))]
                if nums:
                    painel[c] = nums[-1]
    rep.terceiros = {"painel": painel}
    # tabelas detalhe
    def _tab(sheet, hdr_row=3):
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        hdr = [_s(c) for c in data[hdr_row - 1]]
        rows = []
        for r in data[hdr_row:]:
            if all(c is None for c in r):
                continue
            if _s(r[0]) is None and _s(r[1]) is None:
                continue
            rows.append({(hdr[i] or f"c{i}"): _s(v) for i, v in enumerate(r) if hdr[i]})
        return rows
    rep.detalhe["a_entregar"] = _tab("ENTREGAR")
    rep.detalhe["a_receber"] = _tab("RECEBER")
    wb.close()


# ------------------------------------------------------------------
# Orquestração
# ------------------------------------------------------------------
def _calendario_dos_snapshots(hist: dict) -> list:
    """Seletor = semanas com snapshot (docx-semente ou capturadas pelo script).
    Chave = data de referência (fim da semana). Janela = docx anterior+1 .. este."""
    keys = sorted(k for k in hist if _is_iso(k))
    semanas = []
    for i, wid in enumerate(keys):
        ref = date.fromisoformat(wid)
        ini = (date.fromisoformat(keys[i - 1]) + timedelta(days=1)) if i > 0 else (ref - timedelta(days=7))
        semanas.append({
            "id": wid,
            "ini": ini.isoformat(),
            "fim": wid,
            "iso": ref.isocalendar()[1],
            "source": hist[wid].get("source"),
        })
    return semanas


def _is_iso(s: str) -> bool:
    try:
        date.fromisoformat(s)
        return True
    except (ValueError, TypeError):
        return False


def build_report(ini: date, fim: date) -> Report:
    rep = Report(semana_inicio=ini.isoformat(), semana_fim=fim.isoformat())
    # ANTES dos builds: build_movimentacao -> _transferencias_internas compara com o
    # snapshot da semana anterior usando `wid < rep.semana_atual`. Atribuído depois,
    # semana_atual era "" ali, nenhuma semana passava no teste e o diff caía sempre no
    # bootstrap contra o arquivo anterior de receptoras — em 14/08/2026 isso recontou
    # as 14 transferências de 07/08 (o diff real dos snapshots é 0).
    rep.semana_atual = fim.isoformat()           # semana de referência = data do fechamento
    build_producao(rep, ini, fim)
    build_receptoras(rep)
    build_headcount(rep)
    build_headcount_delta(rep, fim)
    build_movimentacao(rep, ini, fim)
    build_comerciais(rep)
    build_pendentes(rep)
    rep.docx_ref = _load_docx_ref()               # relatórios oficiais (validação + seed do 1º caso)
    _compute_movimento(rep)                       # saídas/entradas = diff da população contada
    _paricoes_do_roster(rep)                      # potro no roster sem parição na ESTAÇÃO
    _registra_caminhos(rep)                       # pasta de cada fonte, p/ auditoria
    _aplica_manual(rep)                           # campos sem fonte de planilha
    _acumulado_nunca_cai(rep)                     # agregador da safra, nao cai
    # UMA vez, no fim: chamado no meio do caminho ele via as entradas ainda sem os
    # nascimentos e acusava movimentacao fantasma que se resolvia duas linhas depois
    _conferir_delta(rep)
    _compute_confirmados_diff(rep)                # confirmados na semana = diff de confirmados (forward)
    _avisar_pasta_de_saida()
    _avisar_fontes_velhas(ini, fim)                # BLOQUEIA se a fonte for velha
    _arquivar_linhas(rep)                         # historico linha a linha
    _persist_snapshot(rep)                        # congela snapshot CALCULADO desta semana
    rep.calendario = _calendario_dos_snapshots(rep.snapshots)
    return rep


def _load_hist() -> dict:
    if HIST_SNAPSHOTS.exists():
        try:
            return json.loads(HIST_SNAPSHOTS.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _populacao_contada(roster, receptoras_locais) -> list:
    """Conjunto que o headcount conta: animais do plantel + receptoras nos NOSSOS
    locais. Receptora que vai pro sócio sai da contagem (CONTAGEM só tem receptora
    na fazenda e no arrendamento), então tem de contar como saída."""
    return sorted(set(roster or []) | set(receptoras_locais or {}))


def _descreve_mov(nome: str, info_ant: dict, info_atual: dict) -> dict:
    """Uma saída/entrada com contexto: '309' virou 'receptora 309, prenha de JAVA x
    QUEBRUTO, Pao Grande -> sócio'."""
    ant, atual = info_ant.get(nome), info_atual.get(nome)
    ref = atual or ant or {}
    return {
        "animal": nome,
        "tipo": "RECEPTORA" if (ant or atual) else "ANIMAL",
        "local_saida": (ant or {}).get("local"),
        "local_entrada": (atual or {}).get("local"),
        "status": ref.get("status"),
        "embriao": ref.get("embriao"),
        "obs": ref.get("obs"),
    }


def _mapa_receptoras_anterior(semana: str) -> dict:
    """Mapa {receptora: local} da ultima semana congelada antes desta."""
    hist = _load_hist()
    prev = {}
    for wid in sorted(hist):
        if wid < semana and hist[wid].get("receptoras_locais"):
            prev = hist[wid]["receptoras_locais"]
    return prev


def _refina_afeta_headcount(rep: Report):
    """SAIDA-SOCIO nao e igual pra todo mundo.

    Animal que vai pro socio CONTINUA no headcount — a aba CONTAGEM tem bucket
    SOCIO. Receptora, nao: ela so e contada em PAO GRANDE e ARRENDAMENTO, entao ir
    pro socio a TIRA da conta. A regra unica marcava as duas como 'nao afeta', e o
    Δ so nao denunciou porque os erros se cancelavam (0-1 e 2-3 dao -1).
    """
    prev = _mapa_receptoras_anterior(rep.semana_atual)
    if not prev:
        return
    for e in rep.detalhe.get("saidas_diff") or []:
        if "SOCIO" not in _norm(e.get("classificacao")):
            continue
        chave = re.sub(r"^RECEPTORA\s+", "", _norm(e.get("animal")))
        if chave in prev:
            e["afeta_headcount"] = True
            e["era_receptora_contada"] = True
    rep.saidas["saidas_no_headcount"] = sum(
        1 for x in (rep.detalhe.get("saidas_diff") or []) if x.get("afeta_headcount"))


def _compute_movimento(rep: Report):
    """Saídas/entradas na semana.

    Ordem de preferência:
      1. aba SAIDAS-ENTRADAS do controle mensal, classificada — entrada = nascimento
         ou compra, saída = venda ou morte. É a fonte oficial quando o haras preencher.
      2. diff da POPULAÇÃO CONTADA vs o snapshot anterior.
      3. bootstrap do relatório oficial em Word, na primeira captura.

    O diff era só do roster do plantel e por isso vivia dando 0: o plantel (aba
    PLANTEL) não mexe quando a movimentação é de receptora. Em 07/08/2026 a única
    saída da semana foi a receptora 309 indo pro sócio — diff de roster: 0; diff da
    população contada: 1 saída, 0 entradas, que é o '+00 / -01' do relatório e fecha
    com o Δ do headcount (204 -> 203).
    """
    if rep.saidas_planilha:
        ent, sai = rep.saidas_planilha["ENTRADA"], rep.saidas_planilha["SAIDA"]
        # A aba SAIDAS-ENTRADAS registra a VENDA quando ela é fechada, não quando o
        # animal fisicamente sai — e o STATUS PLANTEL pode continuar 'VENDIDO
        # PENDENTE SAIDA' depois disso. Achado em 28/08/2026: INUSITADA DA PAO GRANDE
        # tinha SAIDA-VENDA lançada nesta semana E status ainda pendente — contada
        # como saída E como pendente ao mesmo tempo. O relatório oficial só conta
        # como saída quando ela DE FATO sai (03 saídas, não 04) — então quem ainda
        # está pendente sai da conta de saídas e fica só na de pendentes.
        pendentes_nomes = {_norm(p["nome"]) for p in rep.detalhe.get("terceiros_propriedade") or []}
        sai_pendente = [e for e in sai if _norm(e.get("animal")) in pendentes_nomes]
        if sai_pendente:
            print(f"  [saídas] {len(sai_pendente)} lançamento(s) de venda com STATUS "
                  f"ainda 'VENDIDO PENDENTE SAIDA' — venda fechada mas animal não "
                  f"saiu de fato, fora da conta de saídas: "
                  + "; ".join(e["animal"] for e in sai_pendente))
            sai = [e for e in sai if _norm(e.get("animal")) not in pendentes_nomes]
        rep.saidas["saidas_semana"] = len(sai)
        rep.saidas["entradas_semana"] = len(ent)
        rep.saidas["fonte"] = "SAIDAS-ENTRADAS"
        # O Δ do headcount só pode ser conferido contra quem entra/sai da CONTAGEM:
        # saída pro sócio deixa a fazenda e continua contada (ver CLASSIF_FORA_DO_DELTA).
        rep.saidas["saidas_no_headcount"] = sum(1 for x in sai if x.get("afeta_headcount"))
        rep.saidas["entradas_no_headcount"] = sum(1 for x in ent if x.get("afeta_headcount"))
        rep.detalhe["saidas_diff"] = sai
        rep.detalhe["entradas_diff"] = ent
        _refina_afeta_headcount(rep)
        return

    rep.populacao = _populacao_contada(rep.roster, rep.receptoras_locais)
    hist = _load_hist()
    prev_pop = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("populacao"):
            prev_pop = hist[wid]["populacao"]
    if prev_pop is None:
        # Bootstrap: nenhuma semana anterior guardou a população (campo novo). Monta a
        # anterior com o roster daquele snapshot + o arquivo de receptoras anterior.
        prev_snap = None
        for wid in sorted(hist):
            if wid < rep.semana_atual and hist[wid].get("roster"):
                prev_snap = hist[wid]
        anteriores = _receptoras_arquivos()[1:]
        if prev_snap and anteriores:
            prev_pop = _populacao_contada(prev_snap["roster"],
                                          _receptoras_locais(anteriores[0]))
            print(f"  [saídas/entradas] primeira semana com população guardada: "
                  f"receptoras da semana anterior vindas de {anteriores[0].name}")

    cur = set(rep.populacao)
    if prev_pop and cur:
        rep.saidas["fonte"] = "diff_populacao"
        prev = set(prev_pop)
        saidas = sorted(prev - cur)
        entradas = sorted(cur - prev)
        rep.saidas["saidas_semana"] = len(saidas)
        rep.saidas["entradas_semana"] = len(entradas)
        # nome sozinho não diz nada — receptora é número ("309"). Anexa o que ela é,
        # de onde saiu, pra onde foi e a observação da planilha.
        info_atual = _receptoras_info()
        info_ant = _receptoras_info(_receptoras_arquivos()[1]) if len(_receptoras_arquivos()) > 1 else {}
        rep.detalhe["saidas_diff"] = [_descreve_mov(n, info_ant, info_atual) for n in saidas]
        rep.detalhe["entradas_diff"] = [_descreve_mov(n, info_ant, info_atual) for n in entradas]
    else:
        # BOOTSTRAP: 1ª captura, sem semana anterior p/ diff → semeia do relatório oficial
        dx = (rep.docx_ref or {}).get(rep.semana_atual, {}).get("saidas", {})
        rep.saidas["saidas_semana"] = dx.get("saidas_semana")
        rep.saidas["entradas_semana"] = dx.get("entradas")
        rep.saidas["fonte"] = "docx"
        rep.saidas["_seed"] = "docx" if dx else None
    _conferir_delta(rep)


# receptora -> "50% FULANO", montado em build_producao a partir da ESTACAO DE MONTA
_SOCIO_POR_RECEP: dict = {}
# receptora -> "50% FULANO" pela coluna NOME SOCIO do roster mensal (fonte primária:
# todo potro nascido entra no roster, mesmo antes de a parição ir para a estação)
_SOCIO_ROSTER: dict = {}


def _limpa_socio(v) -> str | None:
    """'50% ELIANE ANDRADE/vendido' -> '50% ELIANE ANDRADE'.

    O sufixo '/vendido' e marca de controle da planilha, nao parte do nome."""
    t = _s(v)
    if not t:
        return None
    t = re.sub(r"\s*/\s*vendid[oa]\s*$", "", t, flags=re.IGNORECASE).strip()
    return t or None


# CONTROLE_DE_PLANTEL mensal, aba PLANTEL: 'NOME SOCIO' ('50% RENATA CAZZANI DE
# CARVALHO'). Fora do PLANTEL_LAYOUT_MENSAL de proposito — só as parições usam.
COL_MENSAL_NOME_SOCIO = 17

# Dado que NENHUMA planilha tem e que muda toda semana. Fica versionado, por
# semana, para congelar no snapshot e aparecer na auditoria como o que é: input
# humano. Não usar os overrides do dashboard para isto — eles vivem no localStorage
# de um navegador só.
MANUAL = BASE_DIR / "_cache" / "semanal_manual.json"


def _manual(semana: str) -> dict:
    """Campos manuais da semana. Semana sem entrada devolve {} — e o campo fica
    vazio no dashboard, nunca herdado da semana anterior: 'ciclando' de outra semana
    é um número errado com cara de certo."""
    if not MANUAL.exists():
        return {}
    try:
        return (json.loads(MANUAL.read_text(encoding="utf-8")) or {}).get(semana) or {}
    except Exception as exc:
        print(f"  [manual] {MANUAL.name} ilegível ({exc!r}) — campos manuais vazios")
        return {}


PARICOES_EXTRA = BASE_DIR / "_cache" / "paricoes_extra.json"
# assinatura de nome de produto no roster: "... RECEP 309", "MACHO ... RECEP 258"
RE_PRODUTO = re.compile(r"RECEP\w*\s*([A-Z0-9]+)\s*$")


# Arquivo das LINHAS das fontes, uma pasta por semana. FICA FORA do
# semanal_snapshots.json de proposito: aquele JSON e embutido no dashboard, e linha a
# linha ele pesaria centenas de KB por semana. Aqui e so historico consultavel.
FONTES_DIR = BASE_DIR / "_cache" / "fontes"
# As linhas brutas ficam AQUI, nao em rep.detalhe: o semanal_data.json e embutido
# inteiro no dashboard, e linha a linha engordaria o HTML sem servir a ninguem lendo.
_LINHAS_BRUTAS: dict = {}


def _arquivar_linhas(rep: Report):
    """Congela as linhas lidas nesta semana: sem isso, linha apagada na origem leva a
    informacao embora — foi assim que a cota das parições de 21/08/2026 se perdeu."""
    FONTES_DIR.mkdir(parents=True, exist_ok=True)
    dados = {
        "semana": rep.semana_atual,
        "fontes": dict(rep.fontes),
        "grupo_embrioes": _LINHAS_BRUTAS.get("grupo") or [],
        "receptoras": [{"animal": k, **v} for k, v in sorted(_receptoras_info().items())],
        "roster": _LINHAS_BRUTAS.get("roster") or [],
        "pendentes_saida": rep.detalhe.get("pendentes_saida") or [],
        "terceiros_de_terceiro": rep.detalhe.get("terceiros_propriedade") or [],
    }
    alvo = FONTES_DIR / f"{rep.semana_atual}.json"
    alvo.write_text(json.dumps(dados, ensure_ascii=False, indent=1), encoding="utf-8")
    n = sum(len(v) for v in dados.values() if isinstance(v, list))
    print(f"  [fontes] {n} linhas arquivadas em _cache/fontes/{alvo.name}")


def _linhas_da_semana(wid: str) -> dict:
    f = FONTES_DIR / f"{wid}.json"
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _arquivo_anterior(semana: str) -> dict:
    """Ultimo arquivo de linhas antes desta semana."""
    if not FONTES_DIR.exists():
        return {}
    ant = sorted(f.stem for f in FONTES_DIR.glob("*.json") if f.stem < semana)
    return _linhas_da_semana(ant[-1]) if ant else {}


ACUMULADO_PISO = BASE_DIR / "_cache" / "acumulado_piso.json"


def _acumulado_nunca_cai(rep: Report):
    """O acumulado da estacao e um agregador: soma ocorrencias ate o fim da estacao e
    nao volta atras. A formula 'vivos no grupo + parições' e derivada, e por isso fragil
    — linha apagada na origem sem parição lançada fazia o numero CAIR (61 -> 60 em
    21/08/2026). Guardamos o maior valor ja visto na safra e usamos como piso, avisando
    quando a derivacao vem abaixo dele."""
    ac = rep.producao.get("acumulado_estacao")
    if ac is None:
        return
    piso = {}
    if ACUMULADO_PISO.exists():
        try:
            piso = json.loads(ACUMULADO_PISO.read_text(encoding="utf-8"))
        except Exception:
            piso = {}
    ant = piso.get(SAFRA_ATUAL)
    if ant is not None and ac < ant:
        print(f"  [acumulado] a derivacao deu {ac}, abaixo do maior valor ja registrado "
              f"na safra ({ant}). Acumulado nao cai — publicando {ant} e mantendo a "
              f"diferenca visivel. Alguma linha saiu da planilha do grupo sem parição "
              f"nem aborto lançado.")
        rep.producao["acumulado_estacao_derivado"] = ac
        rep.producao["acumulado_estacao"] = ant
        ac = ant
    if ant is None or ac > ant:
        piso[SAFRA_ATUAL] = ac
        ACUMULADO_PISO.parent.mkdir(parents=True, exist_ok=True)
        ACUMULADO_PISO.write_text(json.dumps(piso, ensure_ascii=False, indent=2),
                                  encoding="utf-8")


def _produto_do_roster(nome: str) -> str:
    """'MACHO FACEIRA MAPEJO X IMPERIO SAPECADO RECEP 258' -> 'Macho — Faceira
    Mapejo × Imperio Sapecado'. O roster grava em caixa alta com a receptora
    colada; a tabela do dashboard mostra lado a lado com os que vem da ESTACAO,
    entao o formato tem de ser o mesmo."""
    t = re.sub(r"\s+RECEP\w*\s*[A-Z0-9]+\s*$", "", _norm(nome)).strip()
    sexo = ""
    for pref, rot in (("FEMEA", "Fêmea"), ("FEMA", "Fêmea"), ("MACHO", "Macho"),
                      ("POTRA", "Fêmea"), ("POTRO", "Macho")):
        if t.startswith(pref + " "):
            sexo, t = rot, t[len(pref) + 1:]
            break
    t = t.title().replace(" X ", " × ").replace(" Da ", " da ").replace(" De ", " de ")
    return f"{sexo} — {t}" if sexo else t


def _paricoes_do_roster(rep: Report):
    """Parição que o roster conhece e a aba ESTAÇÃO não.

    O roster é prova de que o potro nasceu: ele entra lá com nome de produto. A aba
    ESTAÇÃO é a fonte oficial, mas só vê o que passou pela estação de monta — em
    21/08/2026 nasceram DOIS potros e só um tinha parição lançada. O que faltava
    derrubava DUAS linhas de uma vez, porque o acumulado é 'vivos no grupo + parições':
    a planilha do grupo já tinha apagado a linha do embrião (ele pariu), mas sem
    parição correspondente o número não voltava. Resultado: nascimentos 1 em vez de 2
    E acumulado 60 em vez de 61, pela mesma causa.

    O registro é CUMULATIVO em disco. Se contasse só os novos da semana, o acumulado
    subiria nesta semana e cairia na próxima, quando o potro deixa de ser novidade no
    roster.
    """
    # NOME SOCIO do roster mensal manda: e a coluna que o relatorio copia, e cobre
    # tambem o potro cuja paricao ainda nao foi lancada na estacao (recep 258).
    for row in rep.detalhe.get("nascimentos_semana") or []:
        do_roster = _SOCIO_ROSTER.get(_norm(row.get("receptora")))
        if do_roster:
            row["socio"] = do_roster

    hist = _load_hist()
    prev = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("roster"):
            prev = hist[wid]["roster"]
    reg = {}
    if PARICOES_EXTRA.exists():
        try:
            reg = json.loads(PARICOES_EXTRA.read_text(encoding="utf-8"))
        except Exception:
            reg = {}
    if prev and rep.roster:
        # receptoras das parições que a ESTAÇÃO já entregou — evita contar duas vezes
        na_estacao = {_norm(e.get("receptora")) for e in rep.detalhe.get("nascimentos_semana", [])}
        for nome in sorted(set(rep.roster) - set(prev)):
            m = RE_PRODUTO.search(_norm(nome))
            if not m or m.group(1) in na_estacao or nome in reg:
                continue
            reg[nome] = {"receptora": m.group(1), "semana": rep.semana_atual,
                         "safra": SAFRA_ATUAL}
        PARICOES_EXTRA.parent.mkdir(parents=True, exist_ok=True)
        PARICOES_EXTRA.write_text(json.dumps(reg, ensure_ascii=False, indent=2),
                                  encoding="utf-8")

    da_safra = {k: v for k, v in reg.items() if v.get("safra") == SAFRA_ATUAL}
    if not da_safra:
        return
    desta = [k for k, v in da_safra.items() if v["semana"] == rep.semana_atual]
    rep.producao["acumulado_estacao"] = (rep.producao.get("acumulado_estacao") or 0) + len(da_safra)
    rep.producao["acumulado_paricoes_so_no_roster"] = len(da_safra)

    # FATIA: a cota do embrião vive na planilha do grupo e vai embora quando a linha é
    # apagada na parição. O arquivo de linhas da semana anterior ainda tem — é o que
    # permite fechar o split PG/sócio/vendido em vez de deixá-lo abaixo do acumulado.
    ant = _arquivo_anterior(rep.semana_atual).get("grupo_embrioes") or []
    por_recep = {_norm(l.get("receptora")): l for l in ant}
    split = rep.producao.get("acumulado_estacao_split") or {}
    sem_fatia = []
    for k, v in da_safra.items():
        linha = por_recep.get(_norm(v.get("receptora")))
        if linha and linha.get("fatia"):
            split[linha["fatia"]] = split.get(linha["fatia"], 0) + 1
            v["fatia"] = linha["fatia"]
            v["cota"] = linha.get("cota")
        else:
            sem_fatia.append(k)
    rep.producao["acumulado_estacao_split"] = split
    if sem_fatia:
        print(f"  [nascimentos] fatia nao recuperada para {len(sem_fatia)} parição(ões) "
              f"— o split fica abaixo do acumulado. Sem arquivo de linhas da semana "
              f"anterior nao ha de onde tirar a cota:")
        for k in sem_fatia:
            print(f"    - {k}")
    if desta:
        # NÃO soma em nascimentos: a contagem publicada vem do roster mensal, por data
        # de nascimento, e o potro desta lista já está lá. Somar aqui contaria duas
        # vezes. Este bloco existe só para o ACUMULADO da safra, que precisa da parição
        # que a aba ESTAÇÃO não tem.
        def _socio_da_recep(rec):
            r = _norm(rec)
            if r in _SOCIO_ROSTER:
                return _SOCIO_ROSTER[r]
            if r in _SOCIO_POR_RECEP:
                return _SOCIO_POR_RECEP[r]
            linha = por_recep.get(r) or {}
            return _limpa_socio(linha.get("socio"))

        sem_socio = [k for k in desta if not _socio_da_recep(da_safra[k]["receptora"])]
        if sem_socio:
            print(f"  [nascimentos] sem sócio na estação nem no arquivo do grupo "
                  f"({len(sem_socio)}) — o relatório publica esse nome, aqui fica vazio:")
            for k in sem_socio:
                print(f"    - {k} (recep {da_safra[k]['receptora']})")
        rep.detalhe["nascimentos_so_roster"] = [
            {"produto": _produto_do_roster(k), "receptora": da_safra[k]["receptora"],
             "socio": _socio_da_recep(da_safra[k]["receptora"]),
             "origem": "roster"} for k in desta]

        # Potro que nasce ENTRA no plantel: o roster cresce e o headcount sobe —
        # então ele conta no Δ ('+02 / -01' = 2 nascimentos, 1 venda), e por isso vai
        # para `entradas_no_headcount`, que é o que o Δ usa.
        #
        # Mas NÃO é "entrada na semana": entrada é animal que chega de fora. O potro
        # já aparece em 'Nascimentos'; somá-lo às entradas colocava o mesmo animal em
        # duas linhas do relatório. Fica fora de `entradas_semana` e da lista da
        # seção 5, e continua registrado aqui para quem precisar da abertura do Δ.
        rep.saidas["entradas_no_headcount"] = (
            rep.saidas.get("entradas_no_headcount") or 0) + len(desta)
        rep.saidas["entradas_nascimento"] = len(desta)
    print(f"  [nascimentos] {len(da_safra)} parição(ões) da safra conhecidas só pelo "
          f"roster, somadas ao acumulado (a aba ESTAÇÃO não as tem):")
    for k in sorted(da_safra):
        marca = "  <- nesta semana" if k in desta else ""
        print(f"    - {k} (recep {da_safra[k]['receptora']}){marca}")


def _registra_caminhos(rep: Report):
    """rep.fontes tem o NOME do arquivo; aqui vai o caminho, para a auditoria dizer
    em qual pasta clicar. Um dict à parte para não mexer no formato de rep.fontes,
    que os snapshots antigos já gravaram."""
    rep.fontes_caminhos = {r: caminho_curto(f) for r, f in _FONTES_USADAS.items()}
    rep.fontes_fora_de_lugar = sorted(FONTES_FORA_DE_LUGAR)


def _aplica_manual(rep: Report):
    """Traz os campos manuais da semana para o relatório.

    Roda depois de a semana estar definida, porque o arquivo é indexado por semana —
    e de propósito NÃO cai para a semana anterior quando falta: número de outra
    semana passaria por atual sem ninguém notar."""
    m = _manual(rep.semana_atual)
    ciclando = m.get("doadoras_ciclando")
    rep.receptoras["doadoras_ciclando"] = ciclando
    if ciclando is None:
        print(f"  [manual] doadoras ciclando não preenchida para {rep.semana_atual} "
              f"— escreva em {MANUAL.name}; o card fica vazio")


def _conferir_delta(rep: Report):
    """Δ do headcount = entradas - saídas. As duas contas vêm de fontes diferentes
    (CONTAGEM vs diff da população), então uma confere a outra. Divergir significa
    movimentação que não passou pelas planilhas — tem de aparecer, não sumir."""
    ent, sai = rep.saidas.get("entradas_semana"), rep.saidas.get("saidas_semana")
    ent = rep.saidas.get("entradas_no_headcount", ent)
    sai = rep.saidas.get("saidas_no_headcount", sai)

    # O badge '+02 / -01' do dashboard tem de dizer a MESMA coisa que o relatorio, e
    # o relatorio abre o movimento de ANIMAIS: potro que nasceu entrou, animal
    # vendido saiu. Receptora nao entra nessa abertura — ela e contada a parte, e
    # receptora que vai pro socio sai da contagem sem aparecer ali.
    # A conta vem do roster, que e a lista de animais: quem entrou e quem saiu dele.
    # Usar o movimento que afeta o headcount dava '+2 / -3' em 21/08/2026, somando as
    # duas receptoras, contra o '+02 / -01' do relatorio.
    hist = _load_hist()
    prev_roster = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("roster"):
            prev_roster = hist[wid]["roster"]
    # Fonte do roster mudou de uma semana para a outra? O diff não vale: o mensal
    # batiza o potro e o semanal o descrevia pelo cruzamento, então TODO potro
    # apareceria como uma saída mais uma entrada. Pular é o certo — inventar
    # movimentação é pior que não ter diff nesta semana.
    prev_fonte = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("roster"):
            prev_fonte = hist[wid].get("roster_fonte")
    if prev_roster and prev_fonte != ROSTER_FONTE:
        print(f"  [roster] fonte mudou de {prev_fonte or 'controle_semanal'} para "
              f"{ROSTER_FONTE} — diff de roster PULADO nesta semana (nome de potro é "
              f"diferente nas duas planilhas). A partir da próxima o diff volta.")
        prev_roster = None
    if prev_roster and rep.roster:
        entraram = set(rep.roster) - set(prev_roster)
        sairam = set(prev_roster) - set(rep.roster)

        # O diff CRU do roster nao serve: renome aparece como uma saida mais uma
        # entrada. Em 21/08/2026 'PERSIA ING DA PAO GRANDE' virou 'PERSIA DA PAO
        # GRANDE' e o badge saiu '+3 / -2' no lugar de '+02 / -01'. Contamos por
        # CAUSA CONHECIDA — nascimento entra, saida lancada sai — e o que sobra do
        # diff vira aviso, em vez de virar numero.
        nasc = rep.producao.get("nascimentos") or 0
        lancadas = {_norm(e.get("animal")) for e in (rep.detalhe.get("saidas_diff") or [])}
        saiu_com_lancamento = {n for n in sairam if _norm(n) in lancadas}
        rep.headcount["delta_entradas"] = nasc
        rep.headcount["delta_saidas"] = len(saiu_com_lancamento)

        sem_causa_saida = sorted(sairam - saiu_com_lancamento)
        sem_causa_entrada = sorted(
            n for n in entraram if not RE_PRODUTO.search(_norm(n)))
        if sem_causa_saida or sem_causa_entrada:
            print(f"  [roster] movimento sem causa lançada — provavel renome, "
                  f"conferir: saiu {sem_causa_saida or '—'}; entrou "
                  f"{sem_causa_entrada or '—'}")
        rep.detalhe["animais_entraram"] = sorted(entraram)
        rep.detalhe["animais_sairam"] = sorted(sairam)
    else:
        rep.headcount["delta_entradas"] = ent
        rep.headcount["delta_saidas"] = sai
    delta = rep.headcount.get("delta")
    if None in (ent, sai) or delta is None:
        return
    if ent - sai != delta:
        print(f"  [Δ] headcount variou {delta:+d} mas o diff da população dá "
              f"+{ent}/-{sai} (líquido {ent - sai:+d}) — conferir: uma das duas "
              f"fontes não registrou alguma movimentação")


def _compute_confirmados_diff(rep: Report):
    """Confirmados na semana = embriões que viraram +/-=OK vs o snapshot anterior
    (novos no conjunto de confirmados). Forward: precisa de 2 semanas capturadas."""
    hist = _load_hist()
    prev_keys = None
    for wid in sorted(hist):
        if wid < rep.semana_atual and hist[wid].get("confirmed_keys") is not None:
            prev_keys = hist[wid]["confirmed_keys"]
    cur = {e["key"]: e for e in rep.confirmed}
    if prev_keys is not None:
        candidatos = [e for k, e in cur.items() if k not in set(prev_keys)]
        # Uma cobrição confirmada não pode ter IA no futuro — confirmação é IA+60d.
        # Achado em 28/08/2026: FACEIRA MAPEJO x IMPERIO SAPECADO só existe na cópia
        # do master na pasta da safra NOVA (a antiga nunca teve a linha), com IA
        # 26/09/2026 e parição JÁ LANÇADA em 15/08/2026 — nasceu antes de cobrir.
        # É erro de digitação na planilha do haras (ano da cobrição), não confirmação
        # nova; contar isso como "confirmado esta semana" é publicar lixo de dado.
        hoje = date.today()
        novos, suspeitos = [], []
        for e in candidatos:
            ia = date.fromisoformat(e["data_ia"]) if e.get("data_ia") else None
            if ia and ia > hoje:
                suspeitos.append(e)
            else:
                novos.append(e)
        if suspeitos:
            print(f"  [confirmados] {len(suspeitos)} confirmação(ões) com IA no futuro, "
                  f"fora da contagem (provável erro de digitação na fonte): " +
                  "; ".join(f"{e['doadora']} x {e['garanhao']} (IA {e['data_ia']})"
                            for e in suspeitos))
        rep.producao["confirmados_semana"] = len(novos)
        rep.detalhe["confirmados_semana"] = novos
    else:
        # BOOTSTRAP: 1ª captura → semeia do relatório oficial
        dx = (rep.docx_ref or {}).get(rep.semana_atual, {}).get("producao", {})
        rep.producao["confirmados_semana"] = dx.get("confirmados_semana")
        rep.detalhe["confirmados_semana"] = []

    # ACUMULADO NO MÊS = novos confirmados desde o último snapshot ANTES do mês corrente
    # (docx "--" = 0; acumulado parado no mês → 0). Não usa IA+60 (proxy ruim).
    month_start = rep.semana_atual[:8] + "01"          # YYYY-MM-01
    prev_month_keys = None
    for wid in sorted(hist):
        if wid < month_start and hist[wid].get("confirmed_keys") is not None:
            prev_month_keys = hist[wid]["confirmed_keys"]
    if prev_month_keys is not None:
        pm = set(prev_month_keys)
        rep.producao["acumulado_mes"] = sum(1 for k in cur if k not in pm)
    else:
        dxp = (rep.docx_ref or {}).get(rep.semana_atual, {}).get("producao", {})
        rep.producao["acumulado_mes"] = dxp.get("acumulado_mes") or 0   # "--" = 0


def _snap_from_rep(rep: Report) -> dict:
    """Snapshot completo desta run (mesmo schema do _map_docx_to_snap)."""
    return {
        "source": "extractor",
        "acumulado_estacao": rep.producao.get("acumulado_estacao"),
        # Mesmo esquecimento do fontes_caminhos: computado em rep.producao, nunca
        # copiado pro snapshot. É a acumulada da SAFRA NOVA — conceito diferente de
        # 'acumulado no mês' (esse é confirmação nova no mês corrente, qualquer
        # safra; aquele é confirmação nova desde que a safra nova começou).
        "acumulado_estacao_proxima": rep.producao.get("acumulado_estacao_proxima"),
        # Mesmo esquecimento, quarta vez: sem isto o card do dashboard nunca aparecia
        # — o skip é 'sem safra_proxima_rotulo, esconde', e o rótulo nunca chegava
        # no snapshot pra além de ser calculado em rep.producao.
        "safra_atual_rotulo": rep.producao.get("safra_atual_rotulo"),
        "safra_proxima_rotulo": rep.producao.get("safra_proxima_rotulo"),
        "confirmados_semana": rep.producao.get("confirmados_semana"),
        "acumulado_mes": rep.producao.get("acumulado_mes"),
        "nascimentos": rep.producao.get("nascimentos"),
        "abortos_obitos": rep.producao.get("abortos_obitos"),
        "acumulado_estacao_split": rep.producao.get("acumulado_estacao_split"),
        "receptoras": rep.receptoras,
        "headcount": {k: v for k, v in rep.headcount.items() if k != "detalhe"},
        # a abertura animais/receptoras por local ia fora do snapshot, então quem quisesse
        # a contagem de um mês passado só tinha a aba CONTAGEM — que é retrato AO VIVO e
        # fazia o slide de junho do comitê exibir a contagem de agosto.
        "headcount_detalhe": rep.headcount.get("detalhe"),
        "terceiros": {k: v for k, v in rep.terceiros.items() if k != "painel"},
        "movimento": {"saidas": rep.saidas.get("saidas_semana"),
                      "entradas": rep.saidas.get("entradas_semana"),
                      "transferencias": rep.saidas.get("transferencias_semana")},
        "detalhe": {
            "confirmados": rep.detalhe.get("confirmados_semana"),
            "nascimentos": rep.detalhe.get("nascimentos_semana"),
            "abortos_obitos": rep.detalhe.get("abortos_obitos_semana"),
            "saidas": rep.detalhe.get("saidas_diff"),
            "entradas": rep.detalhe.get("entradas_diff"),
            "pendentes_saida": rep.detalhe.get("pendentes_saida"),
            # Faltava esta — mesmo esquecimento do fontes_caminhos e do
            # acumulado_estacao_proxima: existe em rep.detalhe, nunca chegava no
            # dict congelado. É a lista que a seção 4 (Terceiros na propriedade /
            # vendidos pendentes) mostra; sem isto ela renderiza vazia mesmo com o
            # KPI certo ao lado, porque o KPI lê rep.terceiros e a tabela lê o snapshot.
            "terceiros_propriedade": rep.detalhe.get("terceiros_propriedade"),
            "terceiros_vendidos": rep.detalhe.get("terceiros_vendidos"),
            "terceiros_vendidos_embrioes": rep.detalhe.get("terceiros_vendidos_embrioes"),
            "terceiros_sociedade": rep.detalhe.get("terceiros_sociedade"),
            "transferencias": rep.detalhe.get("transferencias_internas"),
        },
        "roster": rep.roster,
        "roster_fonte": ROSTER_FONTE,
        "receptoras_locais": rep.receptoras_locais,
        "populacao": rep.populacao,
        "confirmed_keys": [e["key"] for e in rep.confirmed],
        # Sem isto o snapshot esquece de onde cada número saiu — achado em
        # 28/08/2026 quando a auditoria publicou zero caminho de arquivo pra
        # semana inteira: rep.fontes_caminhos existia, _registra_caminhos rodava,
        # mas _snap_from_rep nunca copiava pro dict que de fato é congelado.
        "fontes_caminhos": rep.fontes_caminhos,
        "fontes_fora_de_lugar": rep.fontes_fora_de_lugar,
    }


def _persist_snapshot(rep: Report):
    """Congela o snapshot CALCULADO desta semana. O que vai pro dash é sempre calculado
    aqui — o docx nunca vira dado, só valida (ver rep.docx_ref)."""
    HIST_SNAPSHOTS.parent.mkdir(parents=True, exist_ok=True)
    hist = _load_hist()
    hist[rep.semana_atual] = _snap_from_rep(rep)
    HIST_SNAPSHOTS.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")
    rep.snapshots = hist


def _load_docx_ref() -> dict:
    """Números dos relatórios oficiais (semanal_docx.json), keyed by data de referência.
    Usado SÓ como overlay de validação no dash — nunca como dado."""
    fp = BASES_DIR / "semanal_docx.json"
    if not fp.exists():
        return {}
    out = {}
    for w in json.loads(fp.read_text(encoding="utf-8")):
        out[w["ref"]] = w
    return out


def _cmp(label, got, target):
    flag = "" if target is None else ("  OK" if got == target else f"  (docx: {target})")
    print(f"    {label:38} {got}{flag}")


def print_report(rep: Report):
    print("=" * 66)
    print(f"ATUALIZAÇÃO SEMANAL — {rep.semana_inicio} a {rep.semana_fim}")
    print("=" * 66)
    print("Fontes:")
    for k, v in rep.fontes.items():
        print(f"    {k}: {v}")
    print("\n1) PRODUÇÃO")
    _cmp("Acumulado na estação", rep.producao["acumulado_estacao"], DOCX_1707["acumulado_estacao"])
    print(f"        split PG/sócio/vendido: {rep.producao['acumulado_estacao_split']}")
    _cmp("Embriões confirmados na semana", rep.producao["confirmados_semana"], DOCX_1707["confirmados_semana"])
    _cmp("Acumulado no mês", rep.producao["acumulado_mes"], None)
    _cmp("Nascimentos na semana", rep.producao["nascimentos"], DOCX_1707["nascimentos"])
    _cmp("Abortos / óbitos na semana", rep.producao["abortos_obitos"], DOCX_1707["abortos_obitos"])
    print("\n2) RECEPTORAS (rebanho ativo = FPG + ARRENDAMENTO)")
    _cmp("Total receptoras", rep.receptoras["total"], DOCX_1707["receptoras_total"])
    _cmp("Prenhas", rep.receptoras["prenhas"], DOCX_1707["receptoras_prenhas"])
    _cmp("Vazias", rep.receptoras["vazias"], DOCX_1707["receptoras_vazias"])
    _cmp("Índice eficiência (vazias/doadoras)", rep.receptoras["indice_eficiencia"], None)
    print("\n3) HEADCOUNT")
    _cmp("Total geral", rep.headcount["total"], DOCX_1707["headcount_total"])
    _cmp("Fazenda Pão Grande", rep.headcount["fazenda_pg"], DOCX_1707["headcount_fpg"])
    _cmp("Arrendamento", rep.headcount["arrendamento"], DOCX_1707["headcount_arr"])
    _cmp("Centro de Treinamento", rep.headcount["cte"], DOCX_1707["headcount_cte"])
    _cmp("Sócios", rep.headcount["socio"], DOCX_1707["headcount_soc"])
    _cmp("Δ vs semana anterior", rep.headcount.get("delta"), None)
    print("\n4) TERCEIROS / COMERCIAIS")
    _cmp("Vendidos pendentes de saída", rep.terceiros.get("vendidos_pendentes"), DOCX_1707["vendidos_pendentes"])
    _cmp("Em sociedade pendentes", rep.terceiros.get("sociedade_pendentes"), DOCX_1707["sociedade_pendentes"])
    _cmp("Terceiros na propriedade", rep.terceiros.get("terceiros_propriedade"), None)
    print(f"    painel comerciais: {rep.terceiros.get('painel')}")
    print("\n5) SAÍDAS")
    _cmp("Saídas na semana", rep.saidas["saidas_semana"], DOCX_1707["saidas_semana"])
    _cmp("Entradas na semana", rep.saidas["entradas_semana"], None)
    _cmp("Transferências internas", rep.saidas["transferencias_semana"], None)


def _parse_d(s: str) -> date:
    s = s.strip()
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        d, m, y = s.split("/"); return date(int(y), int(m), int(d))
    return date.fromisoformat(s)


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    args = sys.argv[1:]
    if len(args) >= 2:
        ini, fim = _parse_d(args[0]), _parse_d(args[1])
    elif len(args) == 1:
        fim = _parse_d(args[0]); ini = fim - timedelta(days=14)
    else:
        fim = date.today(); ini = fim - timedelta(days=7)
    rep = build_report(ini, fim)
    print_report(rep)
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(asdict(rep), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n-> {JSON_OUT.name} gravado ({JSON_OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
