"""
Helpers compartilhados pelos scripts do projeto HPG.

Responsabilidades:
  - parse e prompt de mês de referência no formato MM/AAAA
  - localizar planilha mensal no Drive (com tolerância a "REAVALIAÇÃO")
  - cache local da fonte para evitar leituras lentas via shortcut do Drive
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent  # raiz do projeto (scripts/ fica 1 nível abaixo)

# Raiz do Drive
DRIVE_ROOT = Path(
    r"G:\.shortcut-targets-by-id\1mBrSeztRwtBnMlkOMnq6aO4LQUkNjiTb"
    r"\PLANILHAS DE CONTROLE"
)
# Comitê mensal: CONTROLE_DE_PLANTEL
SOURCE_ROOT = DRIVE_ROOT / "PLANTEL"
# Atualização semanal: outras 3 fontes
ESTACAO_MONTA_ROOT = DRIVE_ROOT / "Estação de Monta"  # case-insensitive — search
VENDAS_MAPAS_ROOT = DRIVE_ROOT / "VENDAS" / "MAPAS DE VENDAS"

CACHE_DIR = BASE_DIR / "_cache"
PARQUET_DIR = CACHE_DIR / "parquet"
PARQUET_SEMANAL_DIR = CACHE_DIR / "parquet_semanal"

# Comitê mensal
BASE_BI_XLSX = BASE_DIR / "bases" / "base_bi.xlsx"
BASE_BI_PARQUET = BASE_DIR / "bases" / "base_bi.parquet"
# Atualização semanal
BASE_BI_SEMANAL_XLSX = BASE_DIR / "base_bi_semanal.xlsx"
BASE_BI_SEMANAL_PARQUET = BASE_DIR / "base_bi_semanal.parquet"

PT_MONTHS = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR",
    5: "MAI", 6: "JUN", 7: "JUL", 8: "AGO",
    9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}
PT_TO_MM = {v: k for k, v in PT_MONTHS.items()}

# Cabeçalho esperado da sheet PLANTEL (linha 4). Usamos uma amostra mínima
# como assinatura de compatibilidade — planilhas antigas (2021-2023) têm
# layout diferente e devem ser puladas automaticamente.
PLANTEL_HEADER_SIG = ["QTDE.", "LETRA", "SUFIXO", "NOME", "SEXO", "CATEGORIA", "STATUS PLANTEL", "LOCAL"]

# Regex para arquivos com nomenclatura mensal MES_YY ou MES_YYYY
_FILE_RE = re.compile(
    r".*_CONTROLE_DE_PLANTEL_PAO_GRANDE_(?P<mes>[A-Za-zÇç]{3,})_(?P<ano>\d{2}|\d{4})(?:\b|[^0-9]).*\.xlsx$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class MesRef:
    mm: int
    yyyy: int

    @property
    def tag(self) -> str:
        """'2026-04' — sufixo cronologicamente ordenável para nomes de arquivo."""
        return f"{self.yyyy:04d}-{self.mm:02d}"

    @property
    def mes_pt(self) -> str:
        return PT_MONTHS[self.mm]

    @property
    def yy(self) -> int:
        return self.yyyy % 100

    def __str__(self) -> str:
        return f"{self.mm:02d}/{self.yyyy:04d}"


def parse_mes_aaaa(s: str) -> MesRef:
    s = s.strip()
    if not re.fullmatch(r"\d{1,2}/\d{4}", s):
        raise ValueError(f"Formato inválido. Use MM/AAAA. Recebido: {s!r}")
    mm_s, yyyy_s = s.split("/")
    mm, yyyy = int(mm_s), int(yyyy_s)
    if not (1 <= mm <= 12):
        raise ValueError(f"Mês fora de 1..12: {mm}")
    if yyyy < 2000:
        raise ValueError(f"Ano implausível: {yyyy}")
    return MesRef(mm=mm, yyyy=yyyy)


def prompt_mes_aaaa(label: str = "Mês de referência (MM/AAAA): ") -> MesRef:
    while True:
        try:
            return parse_mes_aaaa(input(label))
        except ValueError as exc:
            print(f"  ! {exc}")


def _parse_filename(path: Path) -> MesRef | None:
    """Extrai MesRef de um nome de arquivo. None se não bater o padrão."""
    if path.name.startswith("~$"):
        return None
    m = _FILE_RE.match(path.name)
    if not m:
        return None
    mes_pt = m.group("mes").upper()[:3]
    if mes_pt not in PT_TO_MM:
        return None
    ano_raw = m.group("ano")
    yyyy = int(ano_raw) if len(ano_raw) == 4 else 2000 + int(ano_raw)
    if yyyy < 2000 or yyyy > 2100:
        return None
    return MesRef(mm=PT_TO_MM[mes_pt], yyyy=yyyy)


def _iter_source_files(source_root: Path = SOURCE_ROOT):
    """Yields (MesRef, Path) para todo arquivo CONTROLE_DE_PLANTEL com naming mensal,
    em todas as pastas Estação YYYY-YYYY."""
    for season_dir in source_root.glob("Estação *"):
        if not season_dir.is_dir():
            continue
        for f in season_dir.glob("*CONTROLE_DE_PLANTEL_PAO_GRANDE_*.xlsx"):
            mes = _parse_filename(f)
            if mes is None:
                continue
            yield mes, f


def _score_candidate(path: Path) -> tuple:
    """Ordenação para escolher entre múltiplos arquivos do mesmo mês.
    Maior é melhor:
      1) tem REAVALIAÇÃO no nome
      2) mtime mais recente
    """
    return ("REAVALIA" in path.name.upper(), path.stat().st_mtime)


def find_source_files(mes: MesRef, source_root: Path = SOURCE_ROOT) -> list[Path]:
    """Lista de candidatos para o mês, ordenada da mais preferida para a menos
    (REAVALIAÇÃO primeiro, depois mtime mais recente)."""
    candidates = [f for m, f in _iter_source_files(source_root) if m == mes]
    if not candidates:
        raise FileNotFoundError(f"Nenhuma planilha CONTROLE_DE_PLANTEL para {mes}")
    candidates.sort(key=_score_candidate, reverse=True)
    return candidates


def find_source_file(mes: MesRef, source_root: Path = SOURCE_ROOT) -> Path:
    """Retorna o candidato top (sem checar compatibilidade)."""
    return find_source_files(mes, source_root)[0]


def find_compatible_source(mes: MesRef, source_root: Path = SOURCE_ROOT) -> tuple[Path, list[str]]:
    """Encontra o primeiro candidato com layout compatível. Tenta cada em ordem
    de preferência. Retorna (path_compatível, lista_de_motivos_dos_rejeitados).
    Levanta FileNotFoundError se nenhum candidato passar.
    """
    motivos: list[str] = []
    for cand in find_source_files(mes, source_root):
        cached = ensure_cache(cand)
        ok, motivo = check_format_compatible(cached)
        if ok:
            return cand, motivos
        motivos.append(f"{cand.name}: {motivo}")
    raise FileNotFoundError(
        f"Nenhuma planilha compatível para {mes}. Rejeitadas:\n  - "
        + "\n  - ".join(motivos)
    )


def list_available_meses(source_root: Path = SOURCE_ROOT) -> list[MesRef]:
    """Retorna todos os meses únicos com pelo menos uma planilha (ordenado)."""
    seen: set[MesRef] = {m for m, _ in _iter_source_files(source_root)}
    return sorted(seen, key=lambda m: (m.yyyy, m.mm))


def check_format_compatible(path: Path) -> tuple[bool, str]:
    """Abre a planilha e valida que a sheet PLANTEL tem o cabeçalho esperado.
    Retorna (True, '') se compatível, (False, motivo) caso contrário.
    Usa read_only=True — não modifica nem trava o arquivo.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return False, f"erro abrindo: {exc!r}"
    try:
        if "PLANTEL" not in wb.sheetnames:
            return False, "sheet 'PLANTEL' ausente"
        ws = wb["PLANTEL"]
        header_row = None
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r > 6:
                break
            if r == 4:
                header_row = row
                break
        if header_row is None:
            return False, "não conseguiu ler linha 4"
        first8 = [str(v).strip() if v is not None else "" for v in header_row[:8]]
        if first8 != PLANTEL_HEADER_SIG:
            return False, f"header linha 4 incompatível: {first8}"
        return True, ""
    finally:
        wb.close()


def ensure_cache(source: Path, cache_dir: Path = CACHE_DIR) -> Path:
    """Copia source → cache local; re-copia só se mtime mudou."""
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_dir / source.name
    src_mtime = source.stat().st_mtime
    if cached.exists() and abs(cached.stat().st_mtime - src_mtime) < 1.0:
        return cached
    print(f"  copiando {source.name} para cache local...")
    shutil.copy2(source, cached)
    return cached
